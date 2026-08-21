#!/usr/bin/env python3
"""Workbooks linexcel is expected to find hard, and what it actually does.

The unit tests answer "is the lineage right?". This answers a different
question, the one that decides whether someone keeps using the tool: *does it
come back at all*, and how long does it take to say so. A user whose file
hangs the analysis does not file a bug — they close the terminal.

Each scenario builds a workbook, and sometimes a folder of the files it links
to, then runs the analysis **in a subprocess under a wall-clock timeout**. A
hang is therefore a result like any other rather than something that takes the
harness down with it, and a segfault in the Rust engine is reported instead of
killing the run.

    uv run python scripts/stress_scenarios.py              # the quick set
    uv run python scripts/stress_scenarios.py --all        # + the heavy ones
    uv run python scripts/stress_scenarios.py --only links # by name fragment
    uv run python scripts/stress_scenarios.py --timeout 30

Exit status is 1 if any scenario crashed or timed out, so this can be a gate.
"""

from __future__ import annotations

import argparse
import json
import shutil
import subprocess
import sys
import tempfile
import time
import zipfile
from collections.abc import Callable
from dataclasses import dataclass
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

from openpyxl import Workbook  # noqa: E402

#: What a scenario hands back: the workbook to analyse, and the folder of
#: linked files to resolve against — ``None`` when the point of the scenario
#: is that the folder is not given.
Built = tuple[Path, Path | None]


@dataclass(frozen=True)
class Scenario:
    name: str
    #: What this attacks, and what a good answer looks like.
    attacks: str
    build: Callable[[Path], Built]
    #: Heavy scenarios are skipped unless --all: they are about cost, not
    #: correctness, and each one takes tens of seconds by design.
    heavy: bool = False


# ──────────────────────────────────────────────
# Building blocks
# ──────────────────────────────────────────────


def _save(wb: Workbook, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def _book(cells: dict[str, object], sheet: str = "S") -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for address, value in cells.items():
        ws[address] = value
    return wb


def _reader(directory: Path, name: str = "main.xlsx", target: str = "Ref.xlsx") -> Path:
    """A workbook whose every interesting cell reads another file."""
    return _save(
        _book(
            {
                "A1": 2,
                "B1": f"='[{target}]Data'!B2 * A1",
                "B2": f"='[{target}]Data'!B3",
                "B3": f"=SUM('[{target}]Data'!B2:B3) + B1",
                "B4": "=B3 * 1.2",
            }
        ),
        directory / name,
    )


def _referenced(path: Path, rows: int = 3) -> Path:
    """The file on the other side of the link, with values worth reading."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["B2"] = 21
    ws["B3"] = "Contoso"
    for r in range(4, rows + 4):
        ws.cell(row=r, column=2, value=r)
    return _save(wb, path)


# ──────────────────────────────────────────────
# Scenarios: the workbooks this one depends on
# ──────────────────────────────────────────────


def links_resolved(d: Path) -> Built:
    refs = d / "refs"
    _referenced(refs / "Ref.xlsx")
    return _reader(d), refs


def links_missing(d: Path) -> Built:
    refs = d / "refs"
    refs.mkdir()
    return _reader(d), refs


def links_no_dir(d: Path) -> Built:
    return _reader(d), None


def links_corrupt(d: Path) -> Built:
    """The linked file is there, and is not readable."""
    refs = d / "refs"
    refs.mkdir()
    good = _referenced(d / "source.xlsx").read_bytes()
    (refs / "Ref.xlsx").write_bytes(good[: len(good) // 2])  # truncated zip
    return _reader(d), refs


def links_not_a_workbook(d: Path) -> Built:
    refs = d / "refs"
    refs.mkdir()
    (refs / "Ref.xlsx").write_text("Region,Amount\nNorth,12\n", encoding="utf-8")
    return _reader(d), refs


def links_legacy_xls(d: Path) -> Built:
    """An OLE2 file wearing an .xlsx name — what a re-saved .xls looks like."""
    refs = d / "refs"
    refs.mkdir()
    (refs / "Ref.xlsx").write_bytes(
        b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 4096
    )
    return _reader(d), refs


def links_empty_file(d: Path) -> Built:
    refs = d / "refs"
    refs.mkdir()
    (refs / "Ref.xlsx").write_bytes(b"")
    return _reader(d), refs


def links_cycle(d: Path) -> Built:
    """A reads B, B reads A. Nothing in the format forbids it."""
    refs = d / "refs"
    _save(_book({"B2": "='[main.xlsx]S'!B1", "B3": 7}, "Data"), refs / "Ref.xlsx")
    main = _reader(d)
    shutil.copy(main, refs / "main.xlsx")
    return main, refs


def links_self(d: Path) -> Built:
    """The workbook links to a file of its own name, sitting in the folder."""
    refs = d / "refs"
    refs.mkdir()
    main = _reader(d, target="main.xlsx")
    shutil.copy(main, refs / "main.xlsx")
    return main, refs


def links_chain(d: Path) -> Built:
    """A reads B, B reads C: resolution is one hop, and should say so."""
    refs = d / "refs"
    _save(_book({"B2": "='[C.xlsx]Data'!B2", "B3": 3}, "Data"), refs / "Ref.xlsx")
    _save(_book({"B2": 99}, "Data"), refs / "C.xlsx")
    return _reader(d), refs


def links_many_files(d: Path) -> Built:
    """A reference folder that is really someone's whole Finance share."""
    refs = d / "refs"
    _referenced(refs / "Ref.xlsx")
    for i in range(300):
        _save(_book({"A1": i}), refs / f"other_{i:03d}.xlsx")
    return _reader(d), refs


def links_deep_tree(d: Path) -> Built:
    """Nested folders: the walk is bounded, and the bound should hold."""
    refs = d / "refs"
    _referenced(refs / "a" / "b" / "c" / "d" / "e" / "Ref.xlsx")
    for i in range(20):
        _save(_book({"A1": i}), refs / f"lvl{i}" / f"deep_{i}.xlsx")
    return _reader(d), refs


def links_corner_ref(d: Path) -> Built:
    """The linked file has a stray cell in the far corner.

    Nothing about it looks unusual: three values, and a cell someone touched
    once at XFD1048576. Reading it densely asks for half a terabyte.
    """
    refs = d / "refs"
    refs.mkdir()
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["B2"] = 21
    ws["B3"] = "Contoso"
    ws["XFD1048576"] = "corner"
    _save(wb, refs / "Ref.xlsx")
    return _reader(d), refs


def links_big_ref(d: Path) -> Built:
    """The linked workbook is large — every cell of it is read for one value."""
    refs = d / "refs"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["B2"] = 21
    ws["B3"] = "Contoso"
    for r in range(4, 40_004):
        for c in range(1, 11):
            ws.cell(row=r, column=c, value=r * c)
    _save(wb, refs / "Ref.xlsx")
    return _reader(d), refs


# ──────────────────────────────────────────────
# Scenarios: the workbook itself
# ──────────────────────────────────────────────


def main_not_a_workbook(d: Path) -> Built:
    path = d / "main.xlsx"
    path.write_text("this is a CSV someone renamed\n", encoding="utf-8")
    return path, None


def main_empty_zip(d: Path) -> Built:
    """A valid ZIP with none of the parts a workbook is made of."""
    path = d / "main.xlsx"
    with zipfile.ZipFile(path, "w") as zf:
        zf.writestr("hello.txt", "not a workbook")
    return path, None


def main_declared_used_range(d: Path) -> Built:
    """One stray cell at the far corner makes the used range 17 billion."""
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = 1
    ws["A2"] = "=A1+1"
    ws["XFD1048576"] = "corner"
    return _save(wb, d / "main.xlsx"), None


def main_deep_chain(d: Path) -> Built:
    """5,000 cells, each reading the one before it: one very long chain."""
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = 1
    for r in range(2, 5_001):
        ws.cell(row=r, column=1, value=f"=A{r - 1}+1")
    return _save(wb, d / "main.xlsx"), None


def main_many_sheets(d: Path) -> Built:
    wb = Workbook()
    wb.active.title = "S0"
    wb.active["A1"] = 1
    for i in range(1, 150):
        ws = wb.create_sheet(f"S{i}")
        ws["A1"] = f"=S{i - 1}!A1 + 1"
    return _save(wb, d / "main.xlsx"), None


def main_huge_strings(d: Path) -> Built:
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    for r in range(1, 60):
        ws.cell(row=r, column=1, value="x" * 30_000)
        ws.cell(row=r, column=2, value=f"=LEN(A{r})")
    return _save(wb, d / "main.xlsx"), None


def main_many_formulas(d: Path) -> Built:
    """150,000 formulas over ten sheets — an ordinary big corporate file."""
    wb = Workbook()
    for s in range(10):
        ws = wb.create_sheet(f"Sheet{s}") if s else wb.active
        ws.title = f"Sheet{s}"
        for r in range(1, 1_501):
            ws.cell(row=r, column=1, value=r)
            for c in range(2, 12):
                ws.cell(
                    row=r,
                    column=c,
                    value=f"={ws.cell(row=r, column=c - 1).coordinate}*1.1",
                )
    return _save(wb, d / "main.xlsx"), None


SCENARIOS = [
    Scenario(
        "links-resolved", "the folder is given and holds the file", links_resolved
    ),
    Scenario(
        "links-missing", "the folder is given and does not hold it", links_missing
    ),
    Scenario("links-no-dir", "no folder at all — the default path", links_no_dir),
    Scenario("links-corrupt", "the linked file is a truncated zip", links_corrupt),
    Scenario(
        "links-not-a-workbook",
        "the linked file is a CSV in disguise",
        links_not_a_workbook,
    ),
    Scenario(
        "links-legacy-xls", "the linked file is OLE2 (a real .xls)", links_legacy_xls
    ),
    Scenario("links-empty-file", "the linked file is zero bytes", links_empty_file),
    Scenario("links-cycle", "A reads B, B reads A", links_cycle),
    Scenario("links-self", "the workbook links to its own name", links_self),
    Scenario("links-chain", "A reads B, B reads C", links_chain),
    Scenario(
        "links-deep-tree", "the linked file is five folders down", links_deep_tree
    ),
    Scenario(
        "links-corner-ref",
        "the linked file declares 17 billion cells",
        links_corner_ref,
    ),
    Scenario("main-not-a-workbook", "a text file named .xlsx", main_not_a_workbook),
    Scenario("main-empty-zip", "a zip with no workbook parts", main_empty_zip),
    Scenario(
        "main-declared-used-range",
        "a stray cell at XFD1048576",
        main_declared_used_range,
    ),
    Scenario("main-deep-chain", "5,000 cells in one dependency chain", main_deep_chain),
    Scenario("main-many-sheets", "150 sheets, each reading the last", main_many_sheets),
    Scenario("main-huge-strings", "cells holding 30 kB of text", main_huge_strings),
    Scenario(
        "links-many-files",
        "300 workbooks in the reference folder",
        links_many_files,
        heavy=True,
    ),
    Scenario(
        "links-big-ref",
        "the linked workbook holds 400,000 cells",
        links_big_ref,
        heavy=True,
    ),
    Scenario(
        "main-many-formulas",
        "150,000 formulas over ten sheets",
        main_many_formulas,
        heavy=True,
    ),
]


# ──────────────────────────────────────────────
# Running one, without letting it take us with it
# ──────────────────────────────────────────────

RUNNER = """
import json, sys, time
sys.path.insert(0, {src!r})
import linexcel
start = time.perf_counter()
try:
    result = linexcel.analyze({workbook!r}, refs_dir={refs!r})
except BaseException as exc:
    print("@@" + json.dumps({{
        "outcome": "error",
        "error": type(exc).__name__,
        "message": str(exc)[:200],
        "seconds": round(time.perf_counter() - start, 2),
    }}))
else:
    print("@@" + json.dumps({{
        "outcome": "ok",
        "seconds": round(time.perf_counter() - start, 2),
        "nodes": len(result.nodes),
        "warnings": [w[:150] for w in result.warnings],
        "external": result.stats.get("externalWorkbooks"),
        "externalRead": result.stats.get("externalWorkbooksRead"),
    }}))
"""


def run(scenario: Scenario, timeout: float) -> dict:
    with tempfile.TemporaryDirectory(prefix=f"linexcel-{scenario.name}-") as tmp:
        directory = Path(tmp)
        try:
            workbook, refs = scenario.build(directory)
        except Exception as exc:  # a scenario that cannot even be built
            return {"outcome": "build-failed", "error": f"{type(exc).__name__}: {exc}"}
        size = workbook.stat().st_size / 1_048_576
        code = RUNNER.format(
            src=str(ROOT / "src"),
            workbook=str(workbook),
            refs=str(refs) if refs else None,
        )
        start = time.perf_counter()
        try:
            done = subprocess.run(
                [sys.executable, "-c", code],
                capture_output=True,
                text=True,
                timeout=timeout,
            )
        except subprocess.TimeoutExpired:
            return {"outcome": "TIMEOUT", "seconds": timeout, "mb": size}
        elapsed = round(time.perf_counter() - start, 2)
        line = next(
            (ln for ln in done.stdout.splitlines() if ln.startswith("@@")), None
        )
        if line is None:
            tail = (done.stderr.strip().splitlines() or ["no output"])[-1]
            return {
                "outcome": "CRASH",
                "returncode": done.returncode,
                "error": tail[:200],
                "seconds": elapsed,
                "mb": size,
            }
        payload = json.loads(line[2:])
        payload["mb"] = size
        payload["wall"] = elapsed
        return payload


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--all", action="store_true", help="include the heavy scenarios"
    )
    parser.add_argument(
        "--only", default="", help="run scenarios whose name contains this"
    )
    parser.add_argument(
        "--timeout", type=float, default=120.0, help="seconds per scenario"
    )
    parser.add_argument(
        "--json", type=Path, default=None, help="also write the results as JSON"
    )
    args = parser.parse_args()

    chosen = [s for s in SCENARIOS if (args.all or not s.heavy) and args.only in s.name]
    print(f"{len(chosen)} scenario(s), {args.timeout:g}s each\n")

    results: dict[str, dict] = {}
    bad = 0
    for scenario in chosen:
        print(f"  {scenario.name:26} …", end="", flush=True)
        outcome = run(scenario, args.timeout)
        results[scenario.name] = {"attacks": scenario.attacks, **outcome}
        verdict = outcome["outcome"]
        if verdict in ("TIMEOUT", "CRASH", "build-failed"):
            bad += 1
        detail = ""
        if verdict == "ok":
            detail = f"{outcome['nodes']} nodes, {len(outcome['warnings'])} warning(s)"
        elif verdict == "error":
            detail = f"{outcome['error']}: {outcome['message'][:80]}"
        elif verdict == "CRASH":
            detail = outcome["error"][:80]
        seconds = outcome.get("wall", outcome.get("seconds", 0))
        print(f"\r  {scenario.name:26} {verdict:12} {seconds:6.2f}s  {detail}")

    if args.json:
        args.json.write_text(json.dumps(results, indent=2), encoding="utf-8")
        print(f"\nWritten to {args.json}")

    print(f"\n{len(chosen) - bad}/{len(chosen)} came back on their own.")
    return 1 if bad else 0


if __name__ == "__main__":
    raise SystemExit(main())
