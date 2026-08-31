"""Builds the lineage graph of an Excel workbook.

Steps:
1. structure (sheets, dimensions, defined names) via openpyxl read-only;
2. formulas + computed values via the Rust engine formualizer;
3. grouping of stretched formulas by R1C1 canonicalization —
   a column of 50,000 copied formulas becomes ONE node;
4. resolution of precedents (cells, ranges, names, other sheets);
5. decomposition of each composite formula into individually evaluated steps
   in a scratch sheet of the engine;
6. lineage of extracted VBA code (oletools);
7. Power Query — the queries that fill a range no formula writes to.
"""

from __future__ import annotations

import datetime
import io
import itertools
import re
import sys
import time
import uuid
import zipfile
from collections import defaultdict
from collections.abc import Callable
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

import formualizer as fz
from openpyxl import load_workbook

from linexcel.external import (
    ExternalBook,
    ExternalRef,
    find_workbooks,
    macro_files,
    parse_external_refs,
    read_external_links,
    read_workbook_values,
    resolve_books,
)
from linexcel.loader import (
    MAX_CELLS_PER_SHEET,
    MAX_DENSE_CELLS,
    CachedValues,
    declared_cells,
    load_cached_values,
)
from linexcel.powerquery import Query, QuerySource, read_queries
from linexcel.progress import Reporter
from linexcel.refs import (
    Rect,
    num_to_col,
    parse_ref,
    parse_ref_detailed,
    stretch_ref,
)
from linexcel.rewrite import canonical_r1c1, qualify_sheet
from linexcel.values import (
    EXCEL_EPOCH_1900,
    EXCEL_ERRORS,
    _date_text_of,
    _fmt_value,
    _is_uncomputed,
    _jsonable,
    readings_agree,
    serial_to_date_text,
)
from linexcel.vba import VbaProc, analyze_vba, extract_vba_modules

# Guards to stay responsive on large workbooks.
SCAN_CHUNK_ROWS = 20_000
#: Ceiling on one ``get_formulas`` call, in cells. The engine hands back a
#: dense grid of Python strings, so a 16,384-column sheet read 20,000 rows at
#: a time would materialize 327 million of them in one go.
SCAN_CHUNK_CELLS = 1_000_000
SMALL_RANGE_CELLS = 20_000
MAX_NODES_PER_SHEET = 400
MAX_STEPS_PER_FORMULA = 48
MAX_SCRATCH_EVALS = 4_000
MAX_VALUE_SAMPLE = 5
MAX_VBA_CODE_CHARS = 6_000
MAX_QUERY_CODE_CHARS = 6_000
#: How many query sources one warning line names before it says "and more".
MAX_QUERY_SOURCES_SHOWN = 6
MAX_VALUE_WARNINGS = 25
# Chained recovery: how deep the precedent walk goes, and how wide a referenced
# range may be before its cells are left to the engine. Both only bound the
# work; a cell that is skipped simply keeps the value the engine reports.
MAX_CHAIN_DEPTH = 24
MAX_CHAIN_RANGE_CELLS = 4_096
SCRATCH_SHEET = "__lineage_scratch__"
# Written into the scratch cell before each guarded evaluation: when the engine
# fails to compute an expression it silently keeps the previous cell value
# instead of raising, so an unchanged marker is how we detect that failure.
SCRATCH_SENTINEL = "__linexcel_no_value__"
GUARD_FUNCTIONS = {"IFERROR", "IFNA"}
#: Functions that answer with the clock or with chance. Excel calls a wider set
#: "volatile" — OFFSET, INDIRECT, CELL and INFO recalculate on every edit too —
#: but those return the same value for the same workbook, so linexcel keeps
#: computing them. These cannot be checked against anything: a `=TODAY()`
#: recomputed today never matches a file saved last week, and calling that a
#: divergence blames the workbook for the calendar.
VOLATILE_FUNCTIONS = frozenset({"NOW", "TODAY", "RAND", "RANDBETWEEN", "RANDARRAY"})
_VOLATILE_RE = re.compile(
    r"(?<![\w.])(?:_xlfn\.)?(?:" + "|".join(sorted(VOLATILE_FUNCTIONS)) + r")\s*\(",
    re.IGNORECASE,
)
#: A double-quoted Excel string, doubled quotes included. Blanked before the
#: volatile scan so `="TODAY() is volatile"` is not read as a call.
_STRING_LITERAL_RE = re.compile(r'"(?:[^"]|"")*"')
#: How many uncomputable cells the warning names before it stops listing them.
MAX_UNCOMPUTED_LISTED = 10


@dataclass
class FormulaGroup:
    """A set of cells on a sheet sharing the same R1C1 formula."""

    sheet: str
    r1c1: str
    cells: list[tuple[int, int]] = field(default_factory=list)
    formulas: dict[tuple[int, int], str] = field(default_factory=dict)

    @property
    def rep(self) -> tuple[int, int]:
        return min(self.cells)

    @property
    def bbox(self) -> tuple[int, int, int, int]:
        rows = [r for r, _ in self.cells]
        cols = [c for _, c in self.cells]
        return min(rows), min(cols), max(rows), max(cols)


class _Budget:
    def __init__(self, limit: int):
        self.left = limit

    def take(self) -> bool:
        if self.left <= 0:
            return False
        self.left -= 1
        return True


class _ExternalResolver:
    """Reads and caches values from workbooks referenced from outside the file.

    A file that registered its external links refers to them by index —
    ``[1]`` — and those are known up front (``externals``). A file that did
    not spells the name out, ``[Budget FY26.xlsx]``, and then the reference
    folder (``refs_files``) is the only place to look; such a workbook is
    read once, on first use, and remembered in ``_named_books``.
    """

    def __init__(
        self,
        externals: dict[str, ExternalBook] | None,
        refs_files: dict[str, Path] | None,
        warnings: list[str],
    ):
        self.externals = externals or {}
        self.refs_files = refs_files or {}
        self.warnings = warnings
        self._declared_by_name = {
            book.name.lower(): book for book in self.externals.values() if book.name
        }
        self._named_books: dict[str, ExternalBook] = {}

    def external_books(self, formula: str | None) -> list[dict[str, Any]]:
        """The other workbooks a formula reads, and how far each one was read.

        Named even when unreadable: "this cell depends on Budget FY26.xlsx,
        which linexcel was not given" is the answer a reader needs, and it is
        one a grey ``[1]`` node never gave.
        """
        if not formula:
            return []
        seen: dict[str, dict[str, Any]] = {}
        for ref in parse_external_refs(formula):
            book = self.book_for(ref)
            name = book.name if book else _external_name(ref)
            if name in seen:
                continue
            entry: dict[str, Any] = {"name": name}
            if book is not None and book.target:
                entry["path"] = book.target
            elif ref.directory:
                entry["path"] = ref.directory + name
            if book is not None and book.resolved:
                entry["read"] = "folder"
                entry["file"] = str(book.path)
            elif book is not None and book.cached:
                entry["read"] = "cache"
            else:
                entry["read"] = "none"
            seen[name] = entry
        return list(seen.values())

    def external_value(self, ref: ExternalRef) -> tuple[Any, str | None]:
        """``(value, source)`` for one external reference, or ``(None, None)``."""
        book = self.book_for(ref)
        position = _a1_position(ref.cell)
        if book is None or position is None:
            return None, None
        return book.value(ref.sheet, *position)

    def external_workbooks(self) -> list[ExternalBook]:
        """Every workbook seen, whether the file declared the link or not."""
        return list(self.externals.values()) + list(self._named_books.values())

    def substitute_externals(self, expr: str) -> str:
        """Replace external references by the values they resolve to.

        The engine cannot follow a link to another workbook — nothing in the
        file it loaded points there — so the reference is turned into the
        literal it stands for before the expression is evaluated. A reference
        that resolves to nothing is left alone: the evaluation then fails the
        way it did before, which is the honest outcome.
        """
        refs = parse_external_refs(expr)
        if not refs:
            return expr
        for ref in refs:
            book = self.book_for(ref)
            if book is None:
                continue
            position = _a1_position(ref.cell)
            if position is None:
                continue
            value, source = book.value(ref.sheet, *position)
            if source is None:
                continue
            expr = expr.replace(ref.text, _as_literal(value))
        return expr

    def book_for(self, ref: ExternalRef) -> ExternalBook | None:
        """The workbook a reference points at, read on first use.

        A file that registered its links refers to them by index — ``[1]`` —
        and those are known up front. A file that did not spells the name out,
        ``[Budget FY26.xlsx]``, and then the reference folder is the only place
        to look; the workbook is read once and remembered.
        """
        declared = self.externals.get(ref.book) or self._declared_by_name.get(
            ref.book.lower()
        )
        if declared is not None:
            return declared
        if ref.book.isdigit():
            return None  # an index whose link part the file does not carry
        key = ref.book.lower()
        book = self._named_books.get(key)
        if book is None:
            book = ExternalBook(
                key=ref.book, target=ref.directory + ref.book, name=ref.book
            )
            path = self.refs_files.get(key)
            if path is not None:
                try:
                    book.values = read_workbook_values(path)
                    book.path = path
                except Exception as exc:
                    self.warnings.append(
                        f"External workbook '{ref.book}' could not be read: {exc}"
                    )
            self._named_books[key] = book
        return book


class _CacheReader:
    """Reads the values cached in the file's XML and flags divergences.

    Isolated from ``_ValueResolver`` because it only ever touches the cache
    (``cached``) and the shared ``warnings`` list — never the engine, the
    budget, or the recovery/decomposition state that the rest of the resolver
    is built around.
    """

    def __init__(self, cached: CachedValues, warnings: list[str]):
        self.cached = cached
        self.warnings = warnings
        self._compared: set[tuple[str, int, int]] = set()
        self._n_mismatches = 0
        self._uncomputed: list[str] = []

    def cached_value(self, sheet: str | None, row: int, col: int) -> Any:
        if sheet is None:
            return None
        raw = self.cached.get(sheet, row, col)
        # dates stay dates in the graph: a midnight datetime serializes as the
        # bare ``YYYY-MM-DD`` so it compares cleanly against ``valueDate``
        if isinstance(raw, datetime.datetime):
            if raw.time() == datetime.time.min:
                return raw.date().isoformat()
            return raw.isoformat(sep=" ")
        if isinstance(raw, datetime.date):
            return raw.isoformat()
        return _jsonable(raw)

    def from_cache(
        self, sheet: str, row: int, col: int
    ) -> tuple[Any, str | None, str | None]:
        raw = self.cached.get(sheet, row, col)
        if raw is None:
            return None, None, None
        date_text = _date_text_of(raw)
        return _jsonable(raw), "file", date_text

    def date_text(self, sheet: str, row: int, col: int, raw: Any) -> str | None:
        if isinstance(raw, bool) or not isinstance(raw, (int, float)):
            return _date_text_of(raw)
        cached = self.cached.get(sheet, row, col)
        is_date = isinstance(cached, (datetime.datetime, datetime.date))
        if not is_date and not self.cached.is_date(sheet, row, col):
            return None
        return serial_to_date_text(raw, self.cached.epoch_1904)

    def check_mismatch(
        self, sheet: str, row: int, col: int, raw: Any, date_text: str | None
    ) -> None:
        key = (sheet, row, col)
        if key in self._compared:
            return
        self._compared.add(key)
        cached = self.cached.get(sheet, row, col)
        if cached is None or readings_agree(raw, cached, date_text) != "differ":
            return
        self._n_mismatches += 1
        if self._n_mismatches > MAX_VALUE_WARNINGS:
            if self._n_mismatches == MAX_VALUE_WARNINGS + 1:
                self.warnings.append(
                    "More recalculated values differ from the file "
                    f"(only the first {MAX_VALUE_WARNINGS} are listed)"
                )
            return
        self.warnings.append(
            f"{sheet}!{a1(row, col)}: recalculated {_fmt_value(raw)} "
            f"differs from file value {_fmt_value(cached)}"
        )

    def note_uncomputed(self, sheet: str, row: int, col: int) -> None:
        """Record a cell the engine could not compute, once."""
        label = f"{sheet}!{a1(row, col)}"
        if label not in self._uncomputed:
            self._uncomputed.append(label)

    def uncomputed_warning(self) -> str | None:
        """One line naming the cells left to the value stored in the file.

        Silence would be the wrong report here: those cells show a figure the
        panel attributes to the file, and nothing else would say that linexcel
        met a formula it cannot evaluate.
        """
        if not self._uncomputed:
            return None
        listed = ", ".join(self._uncomputed[:MAX_UNCOMPUTED_LISTED])
        rest = len(self._uncomputed) - MAX_UNCOMPUTED_LISTED
        if rest > 0:
            listed += f", and {rest} more"
        return (
            f"{len(self._uncomputed)} cell(s) could not be computed by the engine "
            f"and keep the value stored in the file: {listed}"
        )


class _StepEvaluator:
    """Scratch-sheet evaluation of individual step expressions, with a budget.

    A step the engine cannot compute reads back as *not evaluated* rather
    than as an error: an error under a step is the spreadsheet's own verdict
    on the formula, and the engine hitting its own limit is not that.
    """

    def __init__(
        self,
        engine,
        scratch_ready: bool,
        budget: _Budget,
        substitute_externals: Callable[[str], str],
    ):
        self.engine = engine
        self.scratch_ready = scratch_ready
        self.budget = budget
        self._substitute_externals = substitute_externals
        self._step_cache: dict[str, tuple[Any, bool]] = {}

    def eval_raw(self, expr: str, sheet: str) -> tuple[Any, bool]:
        """Scratch evaluation keeping the engine value as-is.

        Errors come back as ``{"type": "Error", ...}``; they only become the
        string the graph carries at the very end, because an error fed back
        into the engine is what lets a guard downstream still fire.
        """
        if not self.scratch_ready or not self.budget.take():
            return None, False
        cached = self._step_cache.pop(expr, None)
        if cached is not None:
            return cached
        # Keyed on the expression as written, evaluated on the one the engine
        # can take: a link to another workbook becomes the value it stands for.
        return _scratch_eval(self.engine, self._substitute_externals(expr), sheet)

    def eval_expr(self, expr: str, sheet: str) -> tuple[Any, bool]:
        raw, ok = self.eval_raw(expr, sheet)
        if _is_uncomputed(raw):
            return None, False
        return _jsonable(raw), ok

    def preload_steps(self, exprs: list[str], sheet: str, engine_alive: bool) -> None:
        """Batch-evaluate step expressions in one ``evaluate_cells`` call.

        Each ``evaluate_cell`` on a large workbook recalculates the whole
        dependency graph (~77 ms).  Batching N expressions into a single
        ``evaluate_cells`` pays that cost once instead of N times.  When
        ``engine_alive`` is False the engine state mutates during
        decomposition (``_remember`` calls ``set_value``), so batching is
        unsafe — the pre-computed values would be stale.  In that case the
        cache stays empty and every expression falls back to
        ``_scratch_eval`` one at a time, exactly as before.
        """
        self._step_cache.clear()
        if not self.scratch_ready or not engine_alive or not exprs:
            return
        # ponytail: dedup preserves order — identical sub-expressions share
        # one scratch cell and one cache entry; eval_raw pops on first hit
        # so the second occurrence falls through to _scratch_eval, matching
        # the old per-call behaviour.
        seen: set[str] = set()
        unique: list[str] = []
        for e in exprs:
            if e not in seen:
                seen.add(e)
                unique.append(e)
        targets: list[tuple[str, int, int]] = []
        valid: list[str] = []
        for i, e in enumerate(unique):
            try:
                qualified = qualify_sheet(self._substitute_externals(e), sheet)
            except Exception:
                continue
            try:
                # Primed with the sentinel first, exactly as ``_scratch_eval``
                # does: an expression the engine will not take — a structured
                # reference it cannot resolve, say — makes ``set_formula``
                # no-op rather than raise, and the cell then still holds the
                # step value *another node* left in that column. Reported as
                # this step's own result, that is a value from elsewhere.
                self.engine.set_formula(
                    SCRATCH_SHEET, 2, i + 1, f'="{SCRATCH_SENTINEL}"'
                )
                self.engine.set_formula(SCRATCH_SHEET, 2, i + 1, qualified)
            except Exception:
                continue
            targets.append((SCRATCH_SHEET, 2, i + 1))
            valid.append(e)
        if not targets:
            return
        try:
            results = self.engine.evaluate_cells(targets)
        except Exception:
            return  # a broken reference poisons the batch — fall back
        for e, val in zip(valid, results):
            if val is not None and val != SCRATCH_SENTINEL and not _is_uncomputed(val):
                self._step_cache[e] = (_jsonable(val), True)
            else:
                self._step_cache[e] = (None, False)


class _RecoveryResolver:
    """Per-cell recovery of values ``evaluate_all`` could not compute.

    The engine is all-or-nothing: one cell pointing at a missing sheet makes
    ``evaluate_all`` raise, and from then on every formula cell reads back as
    ``None``. Each read then falls back, in order: the memoized recovery, the
    engine's own per-cell recalculation, and finally an isolated
    re-evaluation of the formula in the scratch sheet (with the IFERROR/IFNA
    fallback branch when the guarded expression itself cannot be computed).

    A scratch evaluation only sees constants: a reference to another
    *formula* cell reads back as blank. Recovery therefore resolves the
    precedents of a formula first, deepest one first, and feeds each
    recovered value back into the engine as a constant so the next
    evaluation — direct reference or range — reads a number instead of a
    formula it cannot compute.
    """

    def __init__(
        self,
        engine,
        engine_sheets: set[str],
        sheet_dims: dict[str, tuple[int, int]],
        budget: _Budget,
        eval_formula: Callable[[str, str, int], tuple[Any, str | None]],
        engine_alive: bool = True,
    ):
        self.engine = engine
        self.engine_sheets = engine_sheets
        self.sheet_dims = sheet_dims
        self.budget = budget
        self._eval_formula = eval_formula
        self.engine_alive = engine_alive
        self._resolved: dict[tuple[str, int, int], tuple[Any, str | None]] = {}
        self._resolving: set[tuple[str, int, int]] = set()
        self.n_recovered = 0
        self.n_unrecovered = 0

    def formula_at(self, sheet: str, row: int, col: int) -> str | None:
        try:
            return self.engine.get_formula(sheet, row, col)
        except Exception:
            return None

    def engine_read(
        self, sheet: str, row: int, col: int, formula: str | None
    ) -> tuple[Any, str | None]:
        memo = self._resolved.get((sheet, row, col))
        if memo is not None:
            return memo
        try:
            raw = self.engine.get_value(sheet, row, col)
        except Exception:
            raw = None
        if raw is not None:
            return raw, "engine"
        if formula is None:
            formula = self.formula_at(sheet, row, col)
        if not formula:
            return None, None
        return self.recover(sheet, row, col, formula)

    def recover(
        self, sheet: str, row: int, col: int, formula: str
    ) -> tuple[Any, str | None]:
        uncomputed = None
        if self.engine_alive:
            try:
                raw = self.engine.evaluate_cell(sheet, row, col)
                if raw is not None:
                    if _is_uncomputed(raw):
                        uncomputed = raw
                    else:
                        return raw, "engine"
            except Exception:
                # The first failure poisons the engine for good: every later
                # whole-graph evaluation would raise the same way.
                self.engine_alive = False
        expr = formula if formula.startswith("=") else "=" + formula
        result = self._eval_formula(sheet, expr, 0)
        if result[0] is None and uncomputed is not None:
            result = uncomputed, None
        return self.remember(sheet, row, col, result)

    def resolve_precedents(self, sheet: str, expr: str, depth: int) -> None:
        """Recover every formula cell the expression reads, deepest first."""
        if depth >= MAX_CHAIN_DEPTH:
            return
        try:
            ast_dict = fz.parse(expr).to_dict()
        except Exception:
            return
        for ref in _collect_ref_strings(ast_dict):
            detail = parse_ref_detailed(ref, default_sheet=sheet)
            if detail is None or detail.rect.sheet not in self.engine_sheets:
                continue
            rect = detail.rect
            dims = self.sheet_dims.get(rect.sheet or "")
            if dims is not None:
                clipped = rect.clipped(*dims)
                if clipped is None:
                    continue
                rect = clipped
            if rect.ncells > MAX_CHAIN_RANGE_CELLS or rect.sheet is None:
                continue
            for r in range(rect.r1, rect.r2 + 1):
                for c in range(rect.c1, rect.c2 + 1):
                    self.resolve_chain(rect.sheet, r, c, depth + 1)

    def resolve_chain(self, sheet: str, row: int, col: int, depth: int) -> None:
        key = (sheet, row, col)
        # ``_resolving`` breaks the self-referencing formula (=B1+B2 in B2):
        # the cell stays unresolved and reads as blank, as it did before.
        if key in self._resolved or key in self._resolving:
            return
        if depth >= MAX_CHAIN_DEPTH or self.budget.left <= 0:
            return
        try:
            if self.engine.get_value(sheet, row, col) is not None:
                return  # constant: the engine already resolves it on its own
        except Exception:
            pass
        formula = self.formula_at(sheet, row, col)
        if not formula:
            return
        expr = formula if formula.startswith("=") else "=" + formula
        self._resolving.add(key)
        try:
            result = self._eval_formula(sheet, expr, depth)
        finally:
            self._resolving.discard(key)
        self.remember(sheet, row, col, result)

    def remember(
        self, sheet: str, row: int, col: int, result: tuple[Any, str | None]
    ) -> tuple[Any, str | None]:
        """Memoize a recovered value and feed it back into the engine."""
        self._resolved[(sheet, row, col)] = result
        raw, _source = result
        if raw is None or _is_uncomputed(raw):
            self.n_unrecovered += 1
            return result
        self.n_recovered += 1
        if not self.engine_alive:
            # Only ever done on the engine rebuilt after a failed global
            # evaluation, which holds no computed value anyway. It drops the
            # formula of the cell, hence the memo read first in ``engine_read``.
            try:
                self.engine.set_value(sheet, row, col, raw)
            except Exception:
                pass
        return result


class _ValueResolver:
    """Single entry point for reading the value of a cell.

    The engine is all-or-nothing: one cell pointing at a missing sheet makes
    ``evaluate_all`` raise, and from then on every formula cell reads back as
    ``None`` — the whole workbook loses its values because of one bad
    reference. So each read is resolved in order: engine value, per-cell
    recalculation, isolated re-evaluation of the formula in the scratch sheet
    (with the IFERROR/IFNA fallback branch when the guarded expression itself
    cannot be computed), and finally the value cached in the file.

    A scratch evaluation only sees constants: a reference to another *formula*
    cell reads back as blank, so ``=A3+A4`` silently yielded 0 and
    ``=SUM(A1:A7)`` counted the constants only. Recovery therefore resolves the
    precedents of a formula first, deepest one first, and feeds each recovered
    value back into the engine as a constant so the next evaluation — direct
    reference or range — reads a number instead of a formula it cannot compute.
    """

    def __init__(
        self,
        engine,
        engine_sheets: set[str],
        cached: CachedValues,
        warnings: list[str],
        budget: _Budget,
        scratch_ready: bool,
        engine_alive: bool = True,
        sheet_dims: dict[str, tuple[int, int]] | None = None,
        externals: dict[str, ExternalBook] | None = None,
        refs_files: dict[str, Path] | None = None,
    ):
        self.engine = engine
        self.engine_sheets = engine_sheets
        self.cached = cached
        self.warnings = warnings
        self.budget = budget
        self.scratch_ready = scratch_ready
        self.sheet_dims = sheet_dims or {}
        self._external = _ExternalResolver(externals, refs_files, warnings)
        self._cache = _CacheReader(cached, warnings)
        self._steps = _StepEvaluator(
            engine, scratch_ready, budget, self.substitute_externals
        )
        self._recovery = _RecoveryResolver(
            engine,
            engine_sheets,
            self.sheet_dims,
            budget,
            self._eval_formula,
            engine_alive,
        )

    @property
    def n_recovered(self) -> int:
        return self._recovery.n_recovered

    @property
    def n_unrecovered(self) -> int:
        return self._recovery.n_unrecovered

    # -- public API --------------------------------------------------------
    def value(
        self, sheet: str | None, row: int, col: int, formula: str | None = None
    ) -> tuple[Any, str | None, str | None]:
        """Return ``(value, source, date_text)`` for one cell."""
        if sheet is None:
            return None, None, None
        if sheet not in self.engine_sheets:
            return self._cache.from_cache(sheet, row, col)
        if formula is None:
            formula = self._recovery.formula_at(sheet, row, col)
        # A volatile formula answers differently every time it is computed, so
        # recomputing it says nothing about the workbook: `=TODAY()` recalculated
        # today cannot agree with a file saved last week, and reporting that as a
        # divergence blames the file for the calendar. The stored value is kept
        # and labelled as the only reading there is.
        if formula and _is_volatile(formula):
            value, _source, date_text = self._cache.from_cache(sheet, row, col)
            return value, "volatile", date_text
        raw, source = self._recovery.engine_read(sheet, row, col, formula)
        if _is_uncomputed(raw):
            self._cache.note_uncomputed(sheet, row, col)
            raw = None
        if raw is None:
            return self._cache.from_cache(sheet, row, col)
        date_text = self._cache.date_text(sheet, row, col, raw)
        self._cache.check_mismatch(sheet, row, col, raw, date_text)
        return _jsonable(raw), source, date_text

    def describe(
        self, sheet: str | None, row: int, col: int, formula: str | None = None
    ) -> dict[str, Any]:
        """Graph fields for a cell: value, provenance, cache and date."""
        value, source, date_text = self.value(sheet, row, col, formula)
        fields: dict[str, Any] = {"value": value}
        if source is not None:
            fields["valueSource"] = source
        cached = self.cached_value(sheet, row, col)
        if cached is not None:
            fields["cachedValue"] = cached
            # Decided here rather than in the viewer's JavaScript: the same
            # question was being answered twice, with two different answers.
            fields["cachedAgreement"] = readings_agree(value, cached, date_text)
        if date_text is not None:
            fields["valueDate"] = date_text
        return fields

    def external_books(self, formula: str | None) -> list[dict[str, Any]]:
        """The other workbooks a formula reads, and how far each one was read."""
        return self._external.external_books(formula)

    def external_value(self, ref: ExternalRef) -> tuple[Any, str | None]:
        """``(value, source)`` for one external reference, or ``(None, None)``."""
        return self._external.external_value(ref)

    def external_workbooks(self) -> list[ExternalBook]:
        """Every workbook seen, whether the file declared the link or not."""
        return self._external.external_workbooks()

    def substitute_externals(self, expr: str) -> str:
        """Replace external references by the values they resolve to."""
        return self._external.substitute_externals(expr)

    def cached_value(self, sheet: str | None, row: int, col: int) -> Any:
        return self._cache.cached_value(sheet, row, col)

    def eval_expr(self, expr: str, sheet: str) -> tuple[Any, bool]:
        """Evaluate an expression in the scratch sheet, if budget allows."""
        return self._steps.eval_expr(expr, sheet)

    def preload_steps(self, exprs: list[str], sheet: str) -> None:
        """Batch-evaluate step expressions in one ``evaluate_cells`` call."""
        self._steps.preload_steps(exprs, sheet, self._recovery.engine_alive)

    # -- internals ---------------------------------------------------------
    def _eval_raw(self, expr: str, sheet: str) -> tuple[Any, bool]:
        return self._steps.eval_raw(expr, sheet)

    def _eval_formula(
        self, sheet: str, expr: str, depth: int
    ) -> tuple[Any, str | None]:
        """Evaluate one formula on its own, precedents resolved first.

        An expression the engine cannot compute counts as a failure, not as a
        result: the IFERROR/IFNA fallback branch is then tried, exactly as it is
        when the evaluation itself does not come back.
        """
        if not self._recovery.engine_alive:
            self._recovery.resolve_precedents(sheet, expr, depth)
        # A value that came out of another workbook is not the engine's own
        # reading of this file, and the card has to be able to say so.
        computed = "external" if parse_external_refs(expr) else "engine"
        raw, ok = self._eval_raw(expr, sheet)
        uncomputed = raw if ok and raw is not None and _is_uncomputed(raw) else None
        if ok and raw is not None and not _is_uncomputed(raw):
            return raw, computed
        fallback = _guard_fallback_expr(expr)
        if fallback is not None:
            raw, ok = self._eval_raw(fallback, sheet)
            if uncomputed is None and ok and raw is not None and _is_uncomputed(raw):
                uncomputed = raw
            if ok and raw is not None and not _is_uncomputed(raw):
                return raw, "fallback"
        if uncomputed is not None:
            return uncomputed, None
        return None, None

    def uncomputed_warning(self) -> str | None:
        """One line naming the cells left to the value stored in the file.

        Silence would be the wrong report here: those cells show a figure the
        panel attributes to the file, and nothing else would say that linexcel
        met a formula it cannot evaluate.
        """
        return self._cache.uncomputed_warning()


def _external_warning(
    workbooks: list[ExternalBook], refs_dir: str | Path | None
) -> str | None:
    """One line naming every workbook this file reads, and how far each got.

    A workbook that depends on files nobody handed over is the common case in
    practice, and the values above such a link are only as good as the cache
    Excel left behind. Saying so is the point.
    """
    books = {b.name: b for b in workbooks if b.name}
    if not books:
        return None
    read = [b.name for b in books.values() if b.resolved]
    cached = [b.name for b in books.values() if not b.resolved and b.cached]
    missing = [b.name for b in books.values() if not b.resolved and not b.cached]
    parts = [f"This workbook reads {len(books)} external workbook(s)."]
    if read:
        parts.append(f"Read from the reference folder: {', '.join(sorted(read))}.")
    if cached:
        parts.append(
            f"Not read, values taken from the cache Excel left in the file: "
            f"{', '.join(sorted(cached))}."
        )
    if missing:
        parts.append(
            f"Neither read nor cached, so cells reading them have no value: "
            f"{', '.join(sorted(missing))}."
        )
        if refs_dir is None:
            parts.append("Pass refs_dir= (CLI: --refs-dir) to resolve them.")
    return " ".join(parts)


def _query_warning(queries: list[Query]) -> str | None:
    """One line for the queries that feed the workbook from outside it.

    A query whose source is a file, a URL or a server is a dependency of the
    same nature as a link to another workbook: the values it produced are in
    the file, what produced them is not. The graph shows the query and names
    its source; nobody should read that as the source having been checked.
    """
    if not queries:
        return None
    loaded = sum(1 for query in queries if query.loaded)
    if len(queries) == 1:
        head = "1 Power Query query feeds this workbook" + (
            ", loaded onto a sheet."
            if loaded
            else ", loaded nowhere (connection only)."
        )
    else:
        head = (
            f"{len(queries)} Power Query queries feed this workbook, "
            f"{loaded} of them loaded onto a sheet."
        )
    parts = [head]
    outside = sorted(
        {source.target for query in queries for source in query.outside_sources()}
    )
    if outside:
        shown = ", ".join(outside[:MAX_QUERY_SOURCES_SHOWN])
        if len(outside) > MAX_QUERY_SOURCES_SHOWN:
            shown += f", … (+{len(outside) - MAX_QUERY_SOURCES_SHOWN})"
        parts.append(
            f"Their data comes from outside the file and was not read: {shown}."
        )
    return " ".join(parts)


def _external_name(ref: ExternalRef) -> str:
    """The file name an external reference names, index or path alike."""
    return ref.book if not ref.book.isdigit() else f"[{ref.book}]"


def _a1_position(cell: str) -> tuple[int, int] | None:
    """``B4`` → ``(4, 2)``; a range takes its top-left cell."""
    rect = parse_ref(cell.split(":")[0])
    return (rect.r1, rect.c1) if rect is not None else None


def _as_literal(value: Any) -> str:
    """A value as the engine would read it back in a formula."""
    if value is None:
        return "0"
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        return repr(value)
    if isinstance(value, datetime.datetime):
        return repr(_serial_of(value))
    text = str(value)
    if text in EXCEL_ERRORS:
        return text
    return '"' + text.replace('"', '""') + '"'


def _serial_of(moment: datetime.datetime) -> float:
    """A datetime as the 1900-epoch serial a formula computes with."""
    delta = moment - EXCEL_EPOCH_1900
    return delta.days + delta.seconds / 86_400


def _is_volatile(formula: str) -> bool:
    """Whether a formula answers with the clock or with chance.

    Text is blanked first: a cell spelling out ``="uses TODAY()"`` holds a
    string, not a call.
    """
    return bool(_VOLATILE_RE.search(_STRING_LITERAL_RE.sub('""', formula)))


def a1(row: int, col: int) -> str:
    return f"{num_to_col(col)}{row}"


#: Seconds per megabyte of uncompressed sheet XML. Measured across workbooks
#: from a thousand cells to two hundred thousand formulas, where the cost per
#: byte stays near enough constant to be worth quoting: what an analysis
#: really costs tracks how much formula there is to read, and that is what the
#: sheet parts weigh. A file of mostly values is faster than this says, which
#: is the right direction for a warning to be wrong in.
SECONDS_PER_SHEET_MB = 0.5
#: Below this the estimate is noise and nobody was going to wait anyway.
WORTH_MENTIONING_SECONDS = 5.0

_SHEET_PART_RE = re.compile(r"xl/worksheets/sheet\d+\.xml")


def sheet_bytes(data: bytes) -> int:
    """Uncompressed weight of the sheet parts, without unpacking one.

    The zip central directory carries each entry's real size, so this costs a
    read of the index — microseconds on a file that takes minutes to analyse.
    That is the whole point: an estimate nobody waits for.
    """
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            return sum(
                entry.file_size
                for entry in zf.infolist()
                if _SHEET_PART_RE.fullmatch(entry.filename)
            )
    except Exception:
        return 0


def inspect_workbook(data: bytes) -> dict[str, Any]:
    """What the file says about itself, before anything analyses it.

    Everything here is read from the package headers — sheet dimensions and
    external link declarations — so it costs milliseconds on a file that would
    take minutes to analyse. That is the point: it answers "is this going to
    be long, and will anything be left out?" *before* someone commits to
    finding out the slow way.

    Declared sizes, not real ones. A sheet claiming 17 billion cells holds
    nothing of the sort, and saying so is exactly the warning worth having.
    """
    owb = load_workbook(io.BytesIO(data), read_only=True, data_only=False)
    try:
        sheets = []
        for ws in owb.worksheets:
            max_row, max_col = ws.max_row, ws.max_column
            if not max_row or not max_col:
                max_row, max_col = _force_dimensions(ws)
            rows, cols = max_row or 1, max_col or 1
            sheets.append(
                {
                    "name": ws.title,
                    "rows": rows,
                    "cols": cols,
                    "cells": rows * cols,
                    "state": ws.sheet_state,
                    "truncated": rows * cols > MAX_CELLS_PER_SHEET,
                }
            )
    finally:
        owb.close()
    books = read_external_links(data)
    weight = sheet_bytes(data)
    return {
        "bytes": len(data),
        "sheetBytes": weight,
        "estimatedSeconds": round(weight / 1_048_576 * SECONDS_PER_SHEET_MB, 1),
        "sheets": sheets,
        "declaredCells": sum(s["cells"] for s in sheets),
        "externalWorkbooks": [b.name for b in books.values()],
        "densePathRefused": declared_cells(data) > MAX_DENSE_CELLS,
        "ceilings": {
            "cellsPerSheet": MAX_CELLS_PER_SHEET,
            "nodesPerSheet": MAX_NODES_PER_SHEET,
            "denseCells": MAX_DENSE_CELLS,
        },
    }


def _load_workbook_structure(
    data: bytes, refs_dir: str | Path | None, warnings: list[str]
) -> tuple[
    dict[str, tuple[int, int]],
    dict[str, list[Rect]],
    dict[str, ExternalBook],
    dict[str, Path],
]:
    """Sheet dimensions, defined names, external links, and the folder's workbooks.

    Read via openpyxl in read-only mode: cheap even on a large file, and the
    only source for defined names and for sheet sizes the engine itself does
    not need. External links are always parsed and named; they are only
    *read*, into ``refs_files``, when the caller points ``refs_dir`` at the
    folder holding them.
    """
    owb = load_workbook(io.BytesIO(data), read_only=True, data_only=False)
    try:
        sheet_dims: dict[str, tuple[int, int]] = {}
        for ws in owb.worksheets:
            max_row, max_col = ws.max_row, ws.max_column
            if not max_row or not max_col:
                max_row, max_col = _force_dimensions(ws)
            sheet_dims[ws.title] = (max_row or 1, max_col or 1)
        defined_names = _collect_defined_names(owb)
    finally:
        owb.close()

    externals = read_external_links(data)
    refs_files: dict[str, Path] = {}
    if refs_dir is not None:
        refs_files = find_workbooks(Path(refs_dir))
        if externals:
            resolve_books(externals, Path(refs_dir), warnings)
    return sheet_dims, defined_names, externals, refs_files


def _init_engine(
    data: bytes,
    sheet_dims: dict[str, tuple[int, int]],
    warnings: list[str],
) -> tuple[Any, set[str], bool, dict[tuple[str, int, int], str], bool]:
    """The Rust engine, evaluated — recovering from an all-or-nothing failure.

    ``evaluate_all`` gives up on the *first* reference it cannot resolve, so a
    single formula pointing at another workbook costs every other cell in the
    file its computed value. When that happens, the offending cells are
    quarantined (their formula set aside) and evaluation is retried; only if
    that retry also fails does the caller fall back to per-cell recovery.

    Returns the engine, its sheet names, whether the global evaluation left it
    usable, the quarantined cells (formula text keyed by cell), and whether a
    scratch sheet for step-by-step decomposition is available.
    """
    engine = fz.Workbook.from_bytes(data)
    engine_sheets = set(engine.sheet_names)
    engine_alive = True
    quarantined: dict[tuple[str, int, int], str] = {}
    try:
        engine.evaluate_all()
    except Exception as exc:  # graph remains useful without values
        # A failed global evaluation does not just drop the values: the engine
        # then reports no formula at all, which would leave the graph empty.
        # Rebuilding from the bytes gives the formulas back.
        engine = fz.Workbook.from_bytes(data)
        quarantined = _quarantine_unresolvable(engine, sheet_dims, engine_sheets)
        retried = False
        if quarantined:
            try:
                engine.evaluate_all()
                retried = True
            except Exception:
                engine = fz.Workbook.from_bytes(data)
        if retried:
            warnings.append(
                f"Global evaluation completed after isolating {len(quarantined)} "
                f"cell(s) whose references the engine cannot resolve; every other "
                f"cell was recomputed. Only those keep the value stored in the "
                f"file, if any. First blocker: {exc}"
            )
        else:
            warnings.append(f"Global evaluation incomplete: {exc}")
            # Values are recovered cell by cell further down.
            engine_alive = False
            quarantined = {}

    scratch_ready = _ensure_scratch(engine)
    return engine, engine_sheets, engine_alive, quarantined, scratch_ready


def _extract_formula_groups(
    engine: Any,
    sheet_dims: dict[str, tuple[int, int]],
    engine_sheets: set[str],
    quarantined: dict[tuple[str, int, int], str],
    warnings: list[str],
    reporter: Reporter,
) -> tuple[
    dict[tuple[str, str], FormulaGroup],
    dict[str, dict[tuple[int, int], str]],
    int,
    list[dict[str, Any]],
]:
    """Sweep every sheet, grouping cells that share the same R1C1 formula.

    A column of 50,000 copied formulas becomes one :class:`FormulaGroup`
    rather than 50,000 nodes; ``cell_owner`` maps each cell back to the group
    (or, later, the "misc" node) that will represent it in the graph.
    """
    groups: dict[tuple[str, str], FormulaGroup] = {}
    cell_owner: dict[str, dict[tuple[int, int], str]] = defaultdict(dict)
    formula_count = 0
    sheet_stats: list[dict[str, Any]] = []

    with reporter.phase("extraction+grouping", total=len(sheet_dims)) as _bar:
        for sheet, (max_row, max_col) in sheet_dims.items():
            if sheet not in engine_sheets:
                warnings.append(f"Sheet '{sheet}' skipped (not loaded by engine)")
                continue
            n_formulas = 0
            scanned = 0
            fsheet = engine.sheet(sheet)
            chunk_rows = _chunk_rows(max_col)
            r0 = 1
            while r0 <= max_row:
                # The ceiling is spent in rows, and the last chunk is clipped to
                # what is left rather than dropped whole: dropping it stopped a
                # 4,000,000-cell budget at 3,600,000 and lost every row of the
                # chunk that would have overshot.
                rows_left = (MAX_CELLS_PER_SHEET - scanned) // max_col
                if rows_left <= 0:
                    warnings.append(
                        f"Sheet '{sheet}' scanned to row {r0 - 1:,} of {max_row:,} "
                        f"({MAX_CELLS_PER_SHEET:,} cell ceiling): formulas below "
                        f"that row are missing from the lineage"
                    )
                    break
                r1 = min(r0 + chunk_rows - 1, max_row, r0 + rows_left - 1)
                ra = fz.RangeAddress(sheet, r0, 1, r1, max_col)
                try:
                    rows = fsheet.get_formulas(ra)
                except Exception as exc:
                    warnings.append(f"Could not read formulas on {sheet}: {exc}")
                    break
                scanned += (r1 - r0 + 1) * max_col
                for i, row_vals in enumerate(rows):
                    r = r0 + i
                    for j, f in enumerate(row_vals):
                        if not f:
                            # A quarantined cell reads back blank: its formula was
                            # removed so the rest of the workbook could evaluate.
                            f = quarantined.get((sheet, r, j + 1))
                        if not f:
                            continue
                        c = j + 1
                        n_formulas += 1
                        key = (sheet, canonical_r1c1(f, r, c))
                        grp = groups.get(key)
                        if grp is None:
                            grp = groups[key] = FormulaGroup(sheet, key[1])
                        grp.cells.append((r, c))
                        # row/col order scan: first cell seen is the representative
                        # (min), keep 3 example formulas
                        if len(grp.formulas) < 3:
                            grp.formulas[(r, c)] = f
                r0 = r1 + 1
            formula_count += n_formulas
            sheet_stats.append(
                {
                    "name": sheet,
                    "rows": max_row,
                    "cols": max_col,
                    "formulaCells": n_formulas,
                }
            )
            _bar.step(f"sweeping {sheet}")

    return groups, cell_owner, formula_count, sheet_stats


def _finalize_formula_groups(
    groups: dict[tuple[str, str], FormulaGroup],
    cell_owner: dict[str, dict[tuple[int, int], str]],
    nodes: dict[str, dict[str, Any]],
    warnings: list[str],
) -> list[tuple[str, FormulaGroup]]:
    """Assign a node id to each formula group, capping nodes per sheet.

    The busiest patterns on a sheet each keep their own node; the rest are
    folded into one ``misc`` node so a sheet with thousands of distinct
    formulas does not turn into thousands of graph nodes. ``cell_owner`` is
    updated in place so later edge-resolution finds the right node for every
    cell, kept or dropped.
    """
    per_sheet_groups: dict[str, list[FormulaGroup]] = defaultdict(list)
    for grp in groups.values():
        per_sheet_groups[grp.sheet].append(grp)

    kept_groups: list[tuple[str, FormulaGroup]] = []
    for sheet, sheet_groups in per_sheet_groups.items():
        sheet_groups.sort(key=lambda g: (-len(g.cells), g.rep))
        kept = sheet_groups[:MAX_NODES_PER_SHEET]
        dropped = sheet_groups[MAX_NODES_PER_SHEET:]
        for grp in kept:
            rep_r, rep_c = grp.rep
            if len(grp.cells) == 1:
                node_id = f"c:{sheet}!{a1(rep_r, rep_c)}"
            else:
                node_id = f"g:{sheet}!{a1(rep_r, rep_c)}#{len(grp.cells)}"
            kept_groups.append((node_id, grp))
            for cell in grp.cells:
                cell_owner[sheet][cell] = node_id
        if dropped:
            n_cells = sum(len(g.cells) for g in dropped)
            misc_id = f"misc:{sheet}"
            nodes[misc_id] = {
                "id": misc_id,
                "kind": "misc",
                "sheet": sheet,
                "label": f"{len(dropped)} other patterns ({n_cells} cells)",
                "count": n_cells,
                "patterns": len(dropped),
            }
            warnings.append(
                f"Sheet '{sheet}': {len(dropped)} formula patterns aggregated "
                f"into a 'misc' node (limit {MAX_NODES_PER_SHEET})"
            )
            for grp in dropped:
                for cell in grp.cells:
                    cell_owner[sheet][cell] = misc_id
    return kept_groups


class _GraphBuilder:
    """Assigns node ids and builds edges for the lineage graph.

    Node ids are content-addressed (by label), so linking the same input
    rect or opaque reference from several formulas reuses one node instead
    of duplicating it. ``nodes``/``edges`` are the caller's own dicts —
    mutated in place, not owned — since ``analyze_workbook`` keeps adding
    to them after this builder's job (defined names, VBA, Power Query).
    """

    def __init__(
        self,
        *,
        sheet_dims: dict[str, tuple[int, int]],
        cell_owner: dict[str, dict[tuple[int, int], str]],
        resolver: _ValueResolver,
        table_index: dict[str, list[dict[str, Any]]],
        kept_groups: list[tuple[str, FormulaGroup]],
        nodes: dict[str, dict[str, Any]],
        edges: dict[tuple[str, str, str], dict[str, Any]],
    ) -> None:
        self.sheet_dims = sheet_dims
        self.cell_owner = cell_owner
        self.resolver = resolver
        self.table_index = table_index
        self.kept_groups = kept_groups
        self.nodes = nodes
        self.edges = edges
        self.input_nodes: dict[str, str] = {}

    def ensure_opaque_node(self, ref: str) -> str:
        """A reference the graph cannot follow, named as precisely as it can be.

        An external reference is one of those, but it is not opaque to the
        *reader*: the file says which workbook it points at, and the node
        carries that — with the value, when the workbook could be read.
        """
        external = parse_external_refs(ref)
        if not external:
            return self.ensure_input_node(Rect(None, 1, 1, 1, 1), opaque_label=ref)
        first = external[0]
        books = self.resolver.external_books(ref)
        name = books[0]["name"] if books else _external_name(first)
        label = f"[{name}]{first.sheet}!{first.cell}" if first.sheet else f"[{name}]"
        node_id = self.input_nodes.get(label)
        if node_id:
            return node_id
        node_id = f"x:{label}"
        node: dict[str, Any] = {
            "id": node_id,
            "kind": "opaque",
            "label": label,
            "sheet": None,
            "ref": ref,
            "externalBooks": books,
        }
        value, source = self.resolver.external_value(first)
        if source is not None:
            node["value"] = _jsonable(value)
            node["valueSource"] = source
        self.nodes[node_id] = node
        self.input_nodes[label] = node_id
        return node_id

    def ensure_input_node(self, rect: Rect, opaque_label: str | None = None) -> str:
        label = opaque_label or rect.to_a1()
        node_id = self.input_nodes.get(label)
        if node_id:
            return node_id
        if opaque_label is not None:
            node_id = f"x:{opaque_label}"
            self.nodes[node_id] = {
                "id": node_id,
                "kind": "opaque",
                "label": opaque_label,
                "sheet": None,
            }
        else:
            node_id = f"i:{label}"
            node: dict[str, Any] = {
                "id": node_id,
                "kind": "input",
                "label": label,
                "sheet": rect.sheet,
                "addr": label.split("!")[-1],
                "count": rect.ncells,
                "values": _sample_range_values(self.resolver, rect),
            }
            if rect.ncells == 1 and rect.sheet is not None:
                node.update(self.resolver.describe(rect.sheet, rect.r1, rect.c1))
                _enrich_with_table(node, self.table_index, rect.sheet, rect.r1, rect.c1)
            self.nodes[node_id] = node
        self.input_nodes[label] = node_id
        return node_id

    def add_edge(self, src: str, dst: str, kind: str, approx: bool = False) -> None:
        if src == dst:
            return
        key = (src, dst, kind)
        e = self.edges.get(key)
        if e is None:
            self.edges[key] = {
                "id": f"e{len(self.edges)}",
                "source": src,
                "target": dst,
                "kind": kind,
                "approx": approx,
            }
        elif not approx:
            e["approx"] = False

    def resolve_rect_edges(self, rect: Rect, target_id: str, kind: str = "dep") -> None:
        """Create precedent → target edges for a referenced range."""
        sheet = rect.sheet
        if sheet not in self.sheet_dims:
            # A sheet this file does not have: another workbook, most often —
            # ``'[1]Annual'!B4`` parses as a sheet name, brackets and all.
            self.add_edge(self.ensure_opaque_node(rect.to_a1()), target_id, kind)
            return
        clipped = rect.clipped(*self.sheet_dims[sheet])
        if clipped is None:
            return
        owners = self.cell_owner.get(sheet, {})
        if clipped.ncells <= SMALL_RANGE_CELLS:
            seen: set[str] = set()
            has_plain = False
            for r in range(clipped.r1, clipped.r2 + 1):
                for c in range(clipped.c1, clipped.c2 + 1):
                    owner = owners.get((r, c))
                    if owner is None:
                        has_plain = True
                    elif owner not in seen:
                        seen.add(owner)
                        self.add_edge(owner, target_id, kind)
            if has_plain:
                self.add_edge(self.ensure_input_node(clipped), target_id, kind)
        else:
            # Huge range: approximate intersection with node bounding boxes.
            for node_id, grp in self.kept_groups:
                if grp.sheet != sheet:
                    continue
                r1, c1, r2, c2 = grp.bbox
                if clipped.intersects(Rect(sheet, r1, c1, r2, c2)):
                    self.add_edge(node_id, target_id, kind, approx=True)
            self.add_edge(self.ensure_input_node(clipped), target_id, kind, approx=True)

    def query_source_node(self, source: QuerySource) -> str:
        """A node for something a query reads that is not in this workbook."""
        node_id = self.ensure_input_node(
            Rect(None, 1, 1, 1, 1), opaque_label=source.target
        )
        self.nodes[node_id].setdefault("sourceKind", source.kind)
        self.nodes[node_id].setdefault("function", source.function)
        return node_id


def _process_vba(
    data: bytes,
    filename: str,
    refs_dir: str | Path | None,
    warnings: list[str],
    builder: _GraphBuilder,
) -> tuple[dict[str, str], list[VbaProc]]:
    """Extract VBA modules/procedures and wire their call and cell edges.

    Code a workbook calls often does not live in it: an .xlam add-in holds
    the functions, and the workbook only names them. Given ``refs_dir``, that
    code is read too, and each module says which file it came from.
    """
    vba_modules = extract_vba_modules(data, filename, warnings)
    vba_procs: list[VbaProc] = analyze_vba(vba_modules) if vba_modules else []
    if refs_dir is not None:
        for addin in macro_files(Path(refs_dir)):
            extra = extract_vba_modules(addin.read_bytes(), addin.name, warnings)
            if not extra:
                continue
            origin = {f"{addin.name}:{name}": code for name, code in extra.items()}
            vba_modules.update(origin)
            vba_procs.extend(analyze_vba(origin))

    # Node ids keep the declared spelling, but both lookups are keyed on the
    # lowercased name: VBA is case-insensitive, so Module1.Taux and
    # module1.TAUX designate the same procedure. proc_ids resolves a qualified
    # name, procs_by_name the unqualified ones _find_calls reports.
    proc_ids: dict[str, str] = {}
    procs_by_name: dict[str, list[str]] = defaultdict(list)
    for proc in vba_procs:
        qualified = f"{proc.module}.{proc.name}"
        pid = f"vp:{qualified}"
        proc_ids[qualified.lower()] = pid
        procs_by_name[proc.name.lower()].append(qualified.lower())
        builder.nodes[pid] = {
            "id": pid,
            "kind": "vba",
            "label": f"{proc.module}.{proc.name}",
            "sheet": None,
            "module": proc.module,
            "proc": proc.name,
            "procKind": proc.kind,
            "lines": [proc.line_start, proc.line_end],
            "code": proc.code[:MAX_VBA_CODE_CHARS],
        }
    for proc in vba_procs:
        pid = proc_ids[f"{proc.module}.{proc.name}".lower()]
        for callee in proc.calls:
            target = _resolve_call(callee, proc.module, proc_ids, procs_by_name)
            if target is not None:
                builder.add_edge(pid, target, "call")
        for ref in proc.refs:
            detail = parse_ref_detailed(ref.ref, default_sheet=ref.sheet)
            if detail is None or detail.rect.sheet is None:
                opaque_id = builder.ensure_input_node(
                    Rect(None, 1, 1, 1, 1),
                    opaque_label=f"VBA:{ref.sheet or '?'}!{ref.ref}",
                )
                if ref.access == "write":
                    builder.add_edge(pid, opaque_id, "vba-write")
                else:
                    builder.add_edge(opaque_id, pid, "vba-read")
                continue
            if ref.access == "write":
                _resolve_vba_write(
                    detail.rect,
                    pid,
                    builder.sheet_dims,
                    builder.cell_owner,
                    builder.add_edge,
                    builder.ensure_input_node,
                )
            else:
                builder.resolve_rect_edges(detail.rect, pid, kind="vba-read")
    return vba_modules, vba_procs


def _process_power_query(
    data: bytes,
    table_index: dict[str, list[dict[str, Any]]],
    name_nodes: dict[str, str],
    warnings: list[str],
    builder: _GraphBuilder,
) -> list[Query]:
    """Extract Power Query queries and wire their source and load edges.

    A range filled by a query has no formula above it, so without this the
    graph shows where the data landed and nothing about where it came from.
    """
    queries = read_queries(data)
    # Keyed on the exact name: M is case-sensitive, so folding here would let
    # two queries that differ only in case collapse onto one node.
    query_ids = {q.name: f"q:{q.name}" for q in queries}
    tables_by_name = {
        table["name"].casefold(): (sheet, table["ref"])
        for sheet, entries in table_index.items()
        for table in entries
        if table.get("name") and table.get("ref")
    }

    for query in queries:
        qid = query_ids[query.name]
        builder.nodes[qid] = {
            "id": qid,
            "kind": "query",
            "label": query.name,
            "sheet": query.loaded_to[0].sheet if query.loaded_to else None,
            "code": query.source[:MAX_QUERY_CODE_CHARS],
            "loadedTo": [d.as_dict() for d in query.loaded_to],
            "sources": [s.as_dict() for s in query.sources],
        }
    for query in queries:
        qid = query_ids[query.name]
        for source in query.sources:
            if source.kind == "query":
                upstream = query_ids.get(source.target)
                if upstream is not None:
                    builder.add_edge(upstream, qid, "query")
                continue
            if source.kind == "table":
                # ``Excel.CurrentWorkbook`` reads a table or a defined name of
                # this very file: that end of the link is in the graph already.
                placed = tables_by_name.get(source.target.casefold())
                if placed is not None:
                    rect = parse_ref(placed[1], default_sheet=placed[0])
                    if rect is not None:
                        builder.resolve_rect_edges(rect, qid, kind="query")
                        continue
                named = name_nodes.get(source.target.upper())
                if named is not None:
                    builder.add_edge(named, qid, "query")
                    continue
            builder.add_edge(builder.query_source_node(source), qid, "query")
        for destination in query.loaded_to:
            rect = (
                parse_ref(destination.ref, default_sheet=destination.sheet)
                if destination.ref
                else None
            )
            if rect is not None:
                builder.add_edge(qid, builder.ensure_input_node(rect), "query-load")

    query_warning = _query_warning(queries)
    if query_warning:
        warnings.append(query_warning)

    return queries


def _process_formula_nodes(
    kept_groups: list[tuple[str, FormulaGroup]],
    name_nodes: dict[str, str],
    defined_names: dict[str, list[Rect]],
    resolver: _ValueResolver,
    table_index: dict[str, list[dict[str, Any]]],
    builder: _GraphBuilder,
) -> None:
    """Builds a node (+ precedent edges) for every kept formula group.

    Mutates ``builder.nodes``/``builder.edges`` in place via the builder's
    own methods; nothing is returned. ``ast_cache`` is local — parsed ASTs
    are reused across cells sharing the same canonical formula, but nothing
    outside this loop needs them afterwards.
    """
    ast_cache: dict[str, Any] = {}
    for node_id, grp in kept_groups:
        rep_r, rep_c = grp.rep
        formula = grp.formulas.get((rep_r, rep_c)) or next(iter(grp.formulas.values()))
        sheet = grp.sheet
        is_group = len(grp.cells) > 1
        try:
            ast = ast_cache.get(formula)
            if ast is None:
                ast = ast_cache[formula] = fz.parse(
                    formula if formula.startswith("=") else "=" + formula
                )
            ast_dict = ast.to_dict()
        except Exception:
            ast, ast_dict = None, None

        refs = _collect_ref_strings(ast_dict) if ast_dict else []
        rmin, cmin, rmax, cmax = grp.bbox
        agg_rects: list[Rect] = []
        for ref in refs:
            detail = parse_ref_detailed(ref, default_sheet=sheet)
            if detail is None:
                up = ref.upper()
                if up in name_nodes:
                    builder.add_edge(name_nodes[up], node_id, "name")
                else:
                    builder.add_edge(builder.ensure_opaque_node(ref), node_id, "dep")
                continue
            rect = (
                stretch_ref(detail, rep_r, rep_c, (rmin, rmax), (cmin, cmax))
                if is_group
                else detail.rect
            )
            agg_rects.append(rect)

        for rect in _merge_rects(agg_rects):
            builder.resolve_rect_edges(rect, node_id)

        value_fields = resolver.describe(sheet, rep_r, rep_c, formula)
        samples = None
        if is_group:
            samples = [
                {
                    "addr": a1(r, c),
                    **resolver.describe(sheet, r, c, grp.formulas.get((r, c))),
                }
                for r, c in _spread_cells(grp.cells, MAX_VALUE_SAMPLE)
            ]

        steps = None
        # A volatile cell is shown as *not* recalculated, so decomposing it
        # would contradict its own card: every step under `=TODAY()+7` would
        # carry a figure computed from today's clock.
        if ast_dict is not None and value_fields.get("valueSource") != "volatile":
            # The root step is the formula itself: when the engine computed the
            # cell, its value is that step's value and needs no scratch pass.
            root_value = (
                value_fields["value"]
                if value_fields.get("valueSource") == "engine"
                else None
            )
            step_exprs = _collect_step_exprs(ast_dict, skip_root=root_value is not None)
            if step_exprs:
                resolver.preload_steps(step_exprs, sheet)
            steps = _decompose(
                ast_dict, sheet, resolver, defined_names, root_value=root_value
            )

        node: dict[str, Any] = {
            "id": node_id,
            "kind": "group" if is_group else "cell",
            "sheet": sheet,
            "addr": a1(rep_r, rep_c),
            "label": (
                f"{sheet}!{a1(rep_r, rep_c)}"
                + (f" x{len(grp.cells)}" if is_group else "")
            ),
            "formula": formula if formula.startswith("=") else "=" + formula,
            "r1c1": grp.r1c1,
            "count": len(grp.cells),
            "bbox": _bbox_a1(grp),
            **value_fields,
            "samples": samples,
            "steps": steps,
        }
        books = resolver.external_books(formula)
        if books:
            node["externalBooks"] = books
        _enrich_with_table(node, table_index, sheet, rep_r, rep_c)
        builder.nodes[node_id] = node


def _process_defined_names(
    defined_names: dict[str, list[Rect]],
    resolver: _ValueResolver,
    builder: _GraphBuilder,
) -> dict[str, str]:
    """Creates a 'name' node per defined name and wires its precedent edges.

    Returns the name -> node id map (uppercased keys) that the formula loop
    uses to resolve an unqualified reference back to its defined name.
    """
    name_nodes: dict[str, str] = {}
    for name, targets in defined_names.items():
        node_id = f"n:{name}"
        name_nodes[name.upper()] = node_id
        value_fields: dict[str, Any] = {"value": None}
        if targets:
            first = targets[0]
            if (
                first.sheet is not None
                and first.r1 == first.r2
                and first.c1 == first.c2
            ):
                value_fields = resolver.describe(first.sheet, first.r1, first.c1)
            else:
                val_samples = _sample_range_values(resolver, first)
                if val_samples:
                    value_fields = {"value": val_samples[0]["value"]}
        builder.nodes[node_id] = {
            "id": node_id,
            "kind": "name",
            "label": name,
            "sheet": targets[0].sheet if targets else None,
            "targets": [t.to_a1() for t in targets],
            **value_fields,
        }
        for rect in targets:
            builder.resolve_rect_edges(rect, node_id, kind="name")
    return name_nodes


def _assemble_graph(
    *,
    filename: str,
    warnings: list[str],
    engine_alive: bool,
    resolver: _ValueResolver,
    refs_dir: str | Path | None,
    sheet_dims: dict[str, tuple[int, int]],
    sheet_stats: list[dict[str, Any]],
    formula_count: int,
    nodes: dict[str, dict[str, Any]],
    edges: dict[tuple[str, str, str], dict[str, Any]],
    kept_groups: list[tuple[str, FormulaGroup]],
    vba_modules: dict[str, str],
    vba_procs: list[VbaProc],
    defined_names: dict[str, list[Rect]],
    table_index: dict[str, list[dict[str, Any]]],
    queries: list[Query],
) -> dict[str, Any]:
    """Appends the recovery/uncomputed/external warnings and builds the graph dict.

    The three warnings depend on the resolver's final tally (cells recovered
    cell by cell, cells left uncomputed, external workbooks not fully read),
    which is only known once every node has been resolved — so this runs
    last, after every other phase of ``analyze_workbook``.
    """
    if not engine_alive and resolver.n_recovered + resolver.n_unrecovered:
        warnings.append(
            f"Values recovered cell by cell: {resolver.n_recovered} recomputed, "
            f"{resolver.n_unrecovered} left to the value stored in the file"
        )
    uncomputed = resolver.uncomputed_warning()
    if uncomputed:
        warnings.append(uncomputed)
    external_warning = _external_warning(resolver.external_workbooks(), refs_dir)
    if external_warning:
        warnings.append(external_warning)

    return {
        "meta": {
            "filename": filename,
            "analyzedAt": datetime.datetime.now(datetime.UTC).isoformat(),
            "engine": "formualizer (Rust)",
            "warnings": warnings,
            "stats": {
                "sheets": sheet_stats,
                "totalFormulas": formula_count,
                "totalNodes": len(nodes),
                "totalEdges": len(edges),
                "groupedPatterns": sum(1 for _, g in kept_groups if len(g.cells) > 1),
                "vbaModules": len(vba_modules),
                "vbaProcs": len(vba_procs),
                "definedNames": len(defined_names),
                "tables": sum(len(t) for t in table_index.values()),
                "externalWorkbooks": len(
                    {b.name for b in resolver.external_workbooks() if b.name}
                ),
                "externalWorkbooksRead": len(
                    {b.name for b in resolver.external_workbooks() if b.resolved}
                ),
                "queries": len(queries),
                "queriesLoaded": sum(1 for q in queries if q.loaded),
            },
        },
        "sheets": list(sheet_dims.keys()),
        "nodes": list(nodes.values()),
        "edges": list(edges.values()),
    }


def analyze_workbook(
    data: bytes,
    filename: str = "workbook.xlsx",
    *,
    verbose: bool = False,
    refs_dir: str | Path | None = None,
) -> dict[str, Any]:
    """Full analysis: returns the JSON-serializable graph and the engine.

    ``refs_dir`` is a folder holding the workbooks this one links to. Without
    it, a cell reading ``'[1]Annual'!B4`` is named and left unresolved — the
    engine has nothing to follow. With it, the referenced file is read and the
    reference is evaluated as the value it stands for; the report says, per
    workbook, whether it was read from that folder, from the cache Excel left
    in the file, or not at all.
    """
    warnings: list[str] = []
    _t0 = time.perf_counter()
    reporter = Reporter(verbose)

    def _v(label: str, t: float) -> None:
        reporter.note(f"{label}: {time.perf_counter() - t:.1f}s")

    # --- 1. structure -----------------------------------------------------
    _t = time.perf_counter()
    sheet_dims, defined_names, externals, refs_files = _load_workbook_structure(
        data, refs_dir, warnings
    )
    _v("structure", _t)

    # values the file itself carries: last resort, and the only source of
    # dates and of what the user actually saw on screen
    _t = time.perf_counter()
    cached = load_cached_values(data, warnings, reporter)
    _v("cached_values", _t)

    # --- 2. computation engine -------------------------------------------
    _t = time.perf_counter()
    engine, engine_sheets, engine_alive, quarantined, scratch_ready = _init_engine(
        data, sheet_dims, warnings
    )
    _v("engine_init+evaluate_all", _t)

    # Tables: declared ones from the package parts, static ones from a small
    # window the engine already holds. A per-cell lookup enriching the nodes.
    _t = time.perf_counter()
    table_index = _build_table_index(data, engine, sheet_dims, engine_sheets)
    _v("tables", _t)

    budget = _Budget(MAX_SCRATCH_EVALS)
    resolver = _ValueResolver(
        engine,
        engine_sheets,
        cached,
        warnings,
        budget,
        scratch_ready,
        engine_alive=engine_alive,
        sheet_dims=sheet_dims,
        externals=externals,
        refs_files=refs_files,
    )

    # --- 3. extraction + grouping ----------------------------------------
    _t = time.perf_counter()
    groups, cell_owner, formula_count, sheet_stats = _extract_formula_groups(
        engine, sheet_dims, engine_sheets, quarantined, warnings, reporter
    )

    # --- 4. formula nodes -------------------------------------------------
    _v("extraction+grouping", _t)
    _t = time.perf_counter()
    nodes: dict[str, dict[str, Any]] = {}
    edges: dict[tuple[str, str, str], dict[str, Any]] = {}

    kept_groups = _finalize_formula_groups(groups, cell_owner, nodes, warnings)

    builder = _GraphBuilder(
        sheet_dims=sheet_dims,
        cell_owner=cell_owner,
        resolver=resolver,
        table_index=table_index,
        kept_groups=kept_groups,
        nodes=nodes,
        edges=edges,
    )

    # defined names -----------------------------------------------------------
    name_nodes = _process_defined_names(defined_names, resolver, builder)

    # formula nodes + edges -------------------------------------------------
    _process_formula_nodes(
        kept_groups, name_nodes, defined_names, resolver, table_index, builder
    )

    # --- 6. VBA --------------------------------------------------------------
    vba_modules, vba_procs = _process_vba(data, filename, refs_dir, warnings, builder)

    # --- 7. Power Query ------------------------------------------------------
    queries = _process_power_query(data, table_index, name_nodes, warnings, builder)

    graph = _assemble_graph(
        filename=filename,
        warnings=warnings,
        engine_alive=engine_alive,
        resolver=resolver,
        refs_dir=refs_dir,
        sheet_dims=sheet_dims,
        sheet_stats=sheet_stats,
        formula_count=formula_count,
        nodes=nodes,
        edges=edges,
        kept_groups=kept_groups,
        vba_modules=vba_modules,
        vba_procs=vba_procs,
        defined_names=defined_names,
        table_index=table_index,
        queries=queries,
    )
    _v("nodes+edges+graph", _t)
    if verbose:
        print(
            f"[linexcel] total: {time.perf_counter() - _t0:.1f}s | "
            f"{len(nodes)} nodes | {len(edges)} edges | "
            f"{formula_count:,} formulas",
            file=sys.stderr,
        )
    return {"graph": graph, "engine": engine, "analysisId": uuid.uuid4().hex[:16]}


# ---------------------------------------------------------------------------
# Composite function decomposition
# ---------------------------------------------------------------------------

_STEP_KINDS = {"Function", "BinaryOp", "UnaryOp"}
#: Excel operator precedence, loosest first. An unknown operator scores 0 and
#: is therefore always parenthesized, which is the safe way to be wrong.
_PRECEDENCE = {
    "=": 1,
    "<": 1,
    ">": 1,
    "<=": 1,
    ">=": 1,
    "<>": 1,
    "&": 2,
    "+": 3,
    "-": 3,
    "*": 4,
    "/": 4,
    "^": 5,
}
#: Unary minus binds tighter than any binary operator: ``-2^2`` is 4 in Excel,
#: and the parser agrees — it reads ``-A1^2`` as ``(-A1)^2``.
_UNARY_PRECEDENCE = 6


def _collect_step_exprs(ast_dict: dict, *, skip_root: bool = False) -> list[str]:
    """First pass: collect every step expression ``_decompose`` will evaluate.

    The counter and ``_STEP_KINDS`` filter mirror ``_decompose`` exactly so
    the expressions batch-evaluated here are the same ones looked up later
    in ``_StepEvaluator.eval_raw``.  Order is pre-order (parent before
    children) — it does not match the post-order evaluation in
    ``_decompose`` but that is fine:
    the batch evaluates all at once and the cache is order-independent.

    ``skip_root`` leaves out the whole-formula expression, which the caller
    passes to ``_decompose`` as ``root_value`` instead. The counter still
    counts it, so both passes keep budgeting the same steps.
    """
    exprs: list[str] = []
    counter = itertools.count()

    def walk(node: dict, is_root: bool) -> None:
        ntype = node.get("node_type")
        if ntype not in _STEP_KINDS:
            return
        if next(counter) >= MAX_STEPS_PER_FORMULA:
            return
        if not (is_root and skip_root):
            exprs.append(_render_expr(node))
        if ntype == "Function":
            children = node.get("args", [])
        elif ntype == "BinaryOp":
            children = [node.get("left"), node.get("right")]
        else:
            children = [node.get("operand") or node.get("expr")]
        for c in children:
            if c:
                walk(c, False)

    walk(ast_dict, True)
    return exprs


def _decompose(
    ast_dict: dict,
    sheet: str,
    resolver: _ValueResolver,
    defined_names: dict[str, list[Rect]] | None = None,
    root_value: Any = None,
) -> dict | None:
    """Step tree: each function / operator becomes an evaluated step.

    ``root_value`` is the value the engine already computed for the cell
    itself. The root step *is* the whole formula, so re-evaluating it in the
    scratch sheet recomputes a value the workbook already holds — and a root
    such as ``SUM(Calculs!H2:H200001)`` makes the engine walk 200,000 formula
    cells again, which measured 29 s on one node. Pass it only when it came
    from the engine: a value read from the file, or an error-guarded fallback,
    is not what evaluating this expression yields.
    """
    counter = itertools.count()

    def expr_of(node: dict) -> str:
        return _render_expr(node)

    def walk(node: dict, depth: int) -> dict | None:
        ntype = node.get("node_type")
        if ntype not in _STEP_KINDS:
            return None
        if next(counter) >= MAX_STEPS_PER_FORMULA:
            return None
        expr = expr_of(node)
        if ntype == "Function":
            label = node.get("name", "?")
            children_ast = node.get("args", [])
        elif ntype == "BinaryOp":
            label = node.get("operator", "?")
            children_ast = [node.get("left"), node.get("right")]
        else:
            label = node.get("operator", "?")
            children_ast = [node.get("operand") or node.get("expr")]
        children_ast = [c for c in children_ast if c]

        inputs = []
        children = []
        for child in children_ast:
            sub = walk(child, depth + 1)
            if sub is not None:
                children.append(sub)
            else:
                ctype = child.get("node_type")
                if ctype == "Reference":
                    ref = child.get("reference", "?")
                    preview, date_text = _ref_preview(
                        resolver, ref, sheet, defined_names
                    )
                    entry: dict[str, Any] = {"ref": ref, "value": preview}
                    if date_text is not None:
                        entry["date"] = date_text
                    inputs.append(entry)
                elif ctype == "Literal":
                    inputs.append({"literal": child.get("value")})

        if depth == 0 and root_value is not None:
            value, evaluated = root_value, True
        else:
            value, evaluated = resolver.eval_expr(expr, sheet)
        return {
            "kind": ntype,
            "label": label,
            "expr": expr,
            "value": value,
            "evaluated": evaluated,
            "inputs": inputs,
            "children": children,
        }

    return walk(ast_dict, 0)


def _render_expr(node: dict) -> str:
    """Reconstruct the expression of an AST subtree (readable form)."""
    ntype = node.get("node_type")
    if ntype == "Function":
        args = ", ".join(_render_expr(a) for a in node.get("args", []))
        return f"{node.get('name', '?')}({args})"
    if ntype == "BinaryOp":
        operator = node.get("operator", "?")
        precedence = _PRECEDENCE.get(operator, 0)
        left_node, right_node = node.get("left", {}), node.get("right", {})
        # Rendered inline rather than through a helper: one Python frame per
        # AST level, and the stress workbook carries a 700-term chain of
        # additions — two frames per level exhausts the interpreter on it.
        left = _render_expr(left_node)
        if _needs_parens(left_node, precedence, right_side=False):
            left = f"({left})"
        right = _render_expr(right_node)
        if _needs_parens(right_node, precedence, right_side=True):
            right = f"({right})"
        return f"{left} {operator} {right}"
    if ntype == "UnaryOp":
        operator = node.get("operator", "?")
        operand = node.get("operand") or node.get("expr") or {}
        rendered = _render_expr(operand)
        if _needs_parens(operand, _UNARY_PRECEDENCE, right_side=True):
            rendered = f"({rendered})"
        # Percent is the one postfix operator: 25%, not %25.
        return f"{rendered}%" if operator == "%" else f"{operator}{rendered}"
    if ntype == "Reference":
        return str(node.get("reference", "?"))
    if ntype == "Literal":
        v = node.get("value")
        if isinstance(v, str):
            return '"' + v.replace('"', '""') + '"'
        if isinstance(v, bool):
            return "TRUE" if v else "FALSE"
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v)
    if ntype == "Array":
        return "{...}"
    if ntype == "Paren":
        inner = node.get("expr") or node.get("inner") or {}
        return f"({_render_expr(inner)})"
    return "?"


def _needs_parens(node: dict, parent_precedence: int, right_side: bool) -> bool:
    """Whether a child operand has to be parenthesized under its parent.

    The parser keeps grouping in the *shape* of the tree and drops the
    parentheses themselves, so rendering a subtree flat changes what the text
    means: ``=D2*(1-Rate)`` came back as ``D2 * 1 - Rate``. That is not only
    misread by a human — each step is evaluated by re-parsing its own rendered
    text, so the step reported 2470.06 for a cell holding 1976.208.

    A right operand of *equal* precedence is parenthesized too: ``A - (B - C)``
    and ``A - B - C`` are different sums, and spelling the grouping out keeps
    the text re-parsing to the very tree it was rendered from, whichever way
    the parser happens to associate.
    """
    if node.get("node_type") != "BinaryOp":
        return False
    precedence = _PRECEDENCE.get(node.get("operator", ""), 0)
    return precedence < parent_precedence or (
        right_side and precedence == parent_precedence
    )


def _scratch_eval(engine, expr: str, sheet: str) -> tuple[Any, bool]:
    """Evaluate an expression on its own in the scratch sheet.

    The cell is primed with a sentinel first: when the engine cannot compute
    an expression it leaves the previous value in place instead of reporting
    an error, and an unchanged sentinel is the only way to tell. The very
    first evaluation on a workbook holding a broken reference raises — the
    engine walks the whole dirty graph — while later ones stay isolated,
    hence the single retry.
    """
    try:
        qualified = qualify_sheet(expr, sheet)
    except Exception:
        return None, False
    for _ in range(2):
        try:
            engine.set_formula(SCRATCH_SHEET, 1, 1, f'="{SCRATCH_SENTINEL}"')
            if engine.evaluate_cell(SCRATCH_SHEET, 1, 1) != SCRATCH_SENTINEL:
                continue
            engine.set_formula(SCRATCH_SHEET, 1, 1, qualified)
            value = engine.evaluate_cell(SCRATCH_SHEET, 1, 1)
        except Exception:
            continue
        if value == SCRATCH_SENTINEL:
            return None, False
        return value, True
    return None, False


def _guard_fallback_expr(expr: str) -> str | None:
    """Fallback branch of a top-level IFERROR/IFNA, as Excel would show it."""
    # ponytail: only a top-level guard is recovered. With a nested one —
    # =IFERROR(SUM(IFERROR(NOSHEET!A1,0)),1) — the whole expression fails to
    # evaluate, so the outer fallback branch is taken (1) where Excel lets the
    # inner guard absorb the broken reference and returns 0. That gap is the
    # accepted ceiling of this recovery.
    try:
        ast_dict = fz.parse(expr).to_dict()
    except Exception:
        return None
    if not isinstance(ast_dict, dict) or ast_dict.get("node_type") != "Function":
        return None
    if str(ast_dict.get("name", "")).upper() not in GUARD_FUNCTIONS:
        return None
    args = ast_dict.get("args") or []
    if len(args) < 2:
        return None
    return "=" + _render_expr(args[1])


#: A bracketed group: an external workbook (``[Budget.xlsx]Sheet1!A1``) or a
#: structured table reference (``SalesTable[Revenue]``). Neither resolves to a
#: cell the engine holds.
_BRACKETED_RE = re.compile(r"\[[^\]]*\]")
#: A 3-D sheet span — ``'First:Last'!A1`` or ``First:Last!A1``. The engine reads
#: the span as one sheet name and reports it missing.
_SPAN_RE = re.compile(r"'[^']*:[^']*'!|(?<![\w$)])[A-Za-z_][\w.]*:[A-Za-z_][\w.]*!")
#: Error guards. A formula using one may well have a defined value despite a
#: reference that cannot be resolved, so it is never isolated.
_GUARD_RE = re.compile(r"\b(?:IFERROR|IFNA|ISERROR|ISERR|ISNA)\s*\(", re.IGNORECASE)
#: A sheet qualifier, quoted or bare. The lookbehind keeps ``#REF!`` out of the
#: bare form: that is an error literal, not a sheet called REF.
_SHEET_QUALIFIER_RE = re.compile(r"'((?:[^']|'')+)'!|(?<![#\w.$])([A-Za-z_][\w.]*)!")


def _quarantine_unresolvable(
    engine, sheet_dims: dict[str, tuple[int, int]], engine_sheets: set[str]
) -> dict[tuple[str, int, int], str]:
    """Blank the formulas that stop ``evaluate_all``, returning what was removed.

    Only ever called after a global evaluation has already failed, so both ways
    of being wrong are safe: quarantining a formula that would in fact have
    evaluated costs it the same per-cell recovery every cell was getting anyway,
    and missing one simply leaves the retry failing as before.

    The formula text is returned rather than restored — ``set_formula`` does not
    put it back once the cell holds a value — and the scan re-injects it so the
    graph still shows what the cell really contains.
    """
    suspects: dict[tuple[str, int, int], str] = {}
    for sheet, (max_row, max_col) in sheet_dims.items():
        if sheet not in engine_sheets:
            continue
        try:
            fsheet = engine.sheet(sheet)
        except Exception:
            continue
        for r0 in range(1, max_row + 1, SCAN_CHUNK_ROWS):
            r1 = min(r0 + SCAN_CHUNK_ROWS - 1, max_row)
            try:
                rows = fsheet.get_formulas(fz.RangeAddress(sheet, r0, 1, r1, max_col))
            except Exception:
                break
            for i, row_vals in enumerate(rows):
                for j, formula in enumerate(row_vals):
                    if formula and _is_unresolvable(formula, engine_sheets):
                        suspects[(sheet, r0 + i, j + 1)] = formula
    for (sheet, row, col), _formula in suspects.items():
        try:
            engine.set_value(sheet, row, col, None)
        except Exception:
            continue
    return suspects


def _is_unresolvable(formula: str, engine_sheets: set[str]) -> bool:
    """Whether ``formula`` names something the engine cannot resolve.

    A guarded formula is never isolated, however broken its reference looks.
    ``IFERROR(NOSHEET!A1, 456)`` *has* a correct value — 456 — and blanking it
    does not merely cost that cell its own value: every range that spans it
    silently loses a term, so a `SUM` over the column returns a different
    number. Quarantine is only safe where there was no value to lose.
    """
    if _GUARD_RE.search(formula):
        return False
    if _BRACKETED_RE.search(formula) or _SPAN_RE.search(formula):
        return True
    # A sheet qualifier naming a sheet the engine never loaded.
    for match in _SHEET_QUALIFIER_RE.finditer(formula):
        name = match.group(1)
        name = name.replace("''", "'") if name else match.group(2)
        if name and name not in engine_sheets:
            return True
    return False


def _ensure_scratch(engine) -> bool:
    try:
        engine.add_sheet(SCRATCH_SHEET)
        return True
    except Exception:
        return SCRATCH_SHEET in set(engine.sheet_names)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _build_table_index(
    data: bytes,
    engine,
    sheet_dims: dict[str, tuple[int, int]],
    engine_sheets: set[str],
) -> dict[str, list[dict[str, Any]]]:
    """Per-sheet table list, for cell→table enrichment of lineage nodes.

    Two sources, both cheap. Declared tables come from the package's own
    ``xl/tables/*.xml`` parts, which carry the range and the column names that
    structured references (``Sales[Amount]``) resolve against. Static ones stay
    a heuristic, over a window of at most 30 × 50 cells read from the engine.

    Neither needs openpyxl. Reaching ``ws.tables`` did — read-only mode drops
    them — and that second, eager parse of the whole workbook cost 45 s of a
    246 s run on a 2.6M-formula file, to read a handful of definitions.
    """
    from linexcel.insights import (
        MAX_TABLES_PER_SHEET,
        STATIC_TABLE_SCAN_COLS,
        STATIC_TABLE_SCAN_ROWS,
        static_tables_from_rows,
    )

    index: dict[str, list[dict[str, Any]]] = {}
    declared = _declared_tables(data)
    for sheet, (max_row, max_col) in sheet_dims.items():
        tables = declared.get(sheet, [])[:MAX_TABLES_PER_SHEET]
        covered = [
            (t["header_row"], t["first_col"], t["last_row"], t["last_col"])
            for t in tables
        ]
        if sheet in engine_sheets:
            rows = _sheet_window(
                engine,
                sheet,
                min(max_row, STATIC_TABLE_SCAN_ROWS),
                min(max_col, STATIC_TABLE_SCAN_COLS),
            )
            try:
                tables = tables + static_tables_from_rows(rows, covered)
            except Exception:
                pass
        if tables:
            index[sheet] = tables
    return index


def _sheet_window(engine, sheet: str, rows: int, cols: int) -> list[list[Any]]:
    """Top-left window of a sheet, as ``data_only=False`` would hand it over.

    A formula cell reads back as its ``=…`` text rather than as its result:
    the static-table heuristic tells a text header from a computed one that
    way, so the window has to keep the distinction openpyxl gave it.
    """
    if rows < 1 or cols < 1:
        return []
    try:
        fsheet = engine.sheet(sheet)
        address = fz.RangeAddress(sheet, 1, 1, rows, cols)
        values = fsheet.get_values(address)
        formulas = fsheet.get_formulas(address)
    except Exception:
        return []
    window: list[list[Any]] = []
    for r, row in enumerate(values):
        frow = formulas[r] if r < len(formulas) else ()
        line: list[Any] = []
        for c, value in enumerate(row):
            formula = frow[c] if c < len(frow) else ""
            if formula:
                line.append(formula if formula.startswith("=") else "=" + formula)
            else:
                line.append(value if value != "" else None)
        window.append(line)
    return window


def _declared_tables(data: bytes) -> dict[str, list[dict[str, Any]]]:
    """Table definitions per sheet, read straight from the package parts.

    Follows the relationships the format itself uses: workbook → worksheet
    parts → the table parts each sheet owns. Anything unreadable is skipped —
    a missing table only costs a node its ``table_name`` tag.
    """
    out: dict[str, list[dict[str, Any]]] = defaultdict(list)
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            book = _xml_part(zf, "xl/workbook.xml")
            if book is None:
                return {}
            rels = _part_rels(zf, "xl/workbook.xml")
            for element in book.iter():
                if _local_name(element.tag) != "sheet":
                    continue
                name = element.get("name")
                rid = next(
                    (v for k, v in element.attrib.items() if _local_name(k) == "id"),
                    None,
                )
                part = rels.get(rid or "")
                if not name or not part:
                    continue
                for table_part in _part_rels(zf, part, kind="table").values():
                    entry = _table_entry(_xml_part(zf, table_part))
                    if entry is not None:
                        out[name].append(entry)
    except Exception:
        return {}
    return dict(out)


def _table_entry(root) -> dict[str, Any] | None:
    """One ``xl/tables/tableN.xml`` part as the index's table dict."""
    if root is None:
        return None
    rect = parse_ref(root.get("ref") or "")
    if rect is None:
        return None
    # An absent headerRowCount means one header row; a table declared without
    # headers says so with a 0, and its first row is data.
    try:
        header_rows = int(root.get("headerRowCount", "1"))
    except ValueError:
        header_rows = 1
    headers = [
        column.get("name")
        for column in root.iter()
        if _local_name(column.tag) == "tableColumn"
    ]
    first_row = rect.r1 + max(header_rows, 0)
    return {
        "name": root.get("displayName") or root.get("name") or "Table",
        "kind": "dynamic",
        "ref": rect.to_a1(),
        "header_row": rect.r1,
        "first_row": first_row,
        "last_row": rect.r2,
        "first_col": rect.c1,
        "last_col": rect.c2,
        "headers": headers,
        "data_rows": max(0, rect.r2 - first_row + 1),
    }


def _local_name(tag: str) -> str:
    """``{namespace}sheet`` → ``sheet``."""
    return tag.rsplit("}", 1)[-1]


def _xml_part(zf: zipfile.ZipFile, part: str):
    try:
        return ElementTree.fromstring(zf.read(part))
    except (KeyError, ElementTree.ParseError):
        return None


def _part_rels(
    zf: zipfile.ZipFile, part: str, kind: str | None = None
) -> dict[str, str]:
    """``{relationship id: part path}`` for one part, optionally by type."""
    base, _, name = part.rpartition("/")
    root = _xml_part(zf, f"{base}/_rels/{name}.rels")
    if root is None:
        return {}
    out: dict[str, str] = {}
    for rel in root:
        rid, target = rel.get("Id"), rel.get("Target")
        if not rid or not target:
            continue
        if kind is not None and not rel.get("Type", "").endswith("/" + kind):
            continue
        out[rid] = _resolve_part(base, target)
    return out


def _resolve_part(base: str, target: str) -> str:
    """Resolve a relationship target against the part that declares it."""
    if target.startswith("/"):
        return target.lstrip("/")
    segments = base.split("/") if base else []
    for segment in target.split("/"):
        if segment == "..":
            if segments:
                segments.pop()
        elif segment not in ("", "."):
            segments.append(segment)
    return "/".join(segments)


def _enrich_with_table(
    node: dict[str, Any],
    table_index: dict[str, list[dict[str, Any]]],
    sheet: str,
    row: int,
    col: int,
) -> None:
    """Tag a node with ``table_name``/``table_column``/``table_row``.

    The representative cell (``row``, ``col``) is checked against every table
    on its sheet; the first match wins. ``table_column`` is the header text of
    the column the cell sits in, ``table_row`` the 0-based offset into the
    table's data area (header row excluded).
    """
    for tbl in table_index.get(sheet, ()):
        if (
            tbl["first_col"] <= col <= tbl["last_col"]
            and tbl["header_row"] <= row <= tbl["last_row"]
        ):
            node["table_name"] = tbl["name"]
            node["table_kind"] = tbl["kind"]
            idx = col - tbl["first_col"]
            headers = tbl["headers"]
            node["table_column"] = headers[idx] if 0 <= idx < len(headers) else None
            node["table_row"] = max(0, row - tbl["first_row"])
            return


def _force_dimensions(ws) -> tuple[int, int]:
    max_row = max_col = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)
        if max_row > 200_000:
            break
    return max_row or 1, max_col or 1


def _collect_defined_names(owb) -> dict[str, list[Rect]]:
    """Defined names, workbook-scoped first then sheet-scoped.

    A name can be declared on a single worksheet, in which case it lives on that
    sheet rather than on the workbook and is invisible to ``owb.defined_names``.
    Such names are common in real files — a per-sheet ``Total`` or ``Limit`` —
    and skipping them turned every formula using one into an unresolved
    reference.

    Scope is not modelled on the node itself: the graph keys names by their text,
    so a workbook-scoped name wins over a sheet-scoped one of the same name, and
    the first sheet to declare it wins among sheets. That matches how the rest of
    the graph resolves names (case-insensitively, by label) and keeps a shadowed
    name from silently replacing the one most formulas mean.
    """
    out: dict[str, list[Rect]] = {}
    for name, dn in _iter_defined_names(owb):
        if name.startswith("_xlnm.") or name in out:
            continue
        rects: list[Rect] = []
        try:
            for sheet, coord in dn.destinations:
                # openpyxl strips the surrounding quotes but leaves the doubled
                # apostrophes of the escaped form, so a sheet called O'Brien
                # arrives as O''Brien. Storing that verbatim makes to_a1() quote
                # it a second time, and the name never resolves to its sheet.
                sheet = sheet.replace("''", "'") if sheet else sheet
                detail = parse_ref_detailed(coord, default_sheet=sheet)
                if detail is not None:
                    rect = detail.rect
                    rects.append(
                        Rect(sheet or rect.sheet, rect.r1, rect.c1, rect.r2, rect.c2)
                    )
        except Exception:
            continue
        if rects:
            out[name] = rects
    return out


def _iter_defined_names(owb):
    """``(name, DefinedName)`` pairs, workbook scope before sheet scope."""
    try:
        yield from owb.defined_names.items()
    except Exception:
        pass
    for worksheet in getattr(owb, "worksheets", []):
        try:
            yield from worksheet.defined_names.items()
        except Exception:
            continue


def _collect_ref_strings(ast_dict: dict) -> list[str]:
    """Every reference a formula makes, excluding names it binds itself.

    The parser reports a ``LET`` binding — and a ``LAMBDA`` parameter — as a
    ``Reference`` node, because syntactically it looks like one. It is not: the
    name is local to the formula and points at no cell, so treating it as a
    reference put a node on the graph for every intermediate a modeller happened
    to name. A formula's own bindings are its business; the graph shows the
    cells, and the LET is visible in that cell's step decomposition.
    """
    bound = _bound_names(ast_dict)
    refs: list[str] = []

    def walk(node) -> None:
        if isinstance(node, dict):
            if node.get("node_type") == "Reference":
                ref = node.get("reference")
                if ref and str(ref).upper() not in bound:
                    refs.append(str(ref))
            for v in node.values():
                walk(v)
        elif isinstance(node, list):
            for v in node:
                walk(v)

    walk(ast_dict)
    # dedupe preserving order
    seen: set[str] = set()
    out = []
    for r in refs:
        if r not in seen:
            seen.add(r)
            out.append(r)
    return out


def _bound_names(ast_dict: dict) -> set[str]:
    """Upper-cased names bound by any ``LET`` or ``LAMBDA`` in the formula.

    ``LET(n1, v1, [n2, v2, ...], calc)`` binds the even-indexed arguments before
    the trailing calculation; ``LAMBDA(p1, ..., calc)`` binds every argument but
    the last. Only bare identifiers count — ``LET(A1, ...)`` is not legal Excel,
    so a name that parses as a real reference is left alone.
    """
    bound: set[str] = set()

    def walk(node) -> None:
        if isinstance(node, dict):
            if node.get("node_type") == "Function":
                name = str(node.get("name") or "").upper()
                args = node.get("args") or []
                if name == "LET":
                    indices = range(0, max(len(args) - 1, 0), 2)
                elif name == "LAMBDA":
                    indices = range(max(len(args) - 1, 0))
                else:
                    indices = range(0)
                for index in indices:
                    arg = args[index]
                    if not isinstance(arg, dict):
                        continue
                    if arg.get("node_type") != "Reference":
                        continue
                    text = str(arg.get("reference") or "")
                    if text and parse_ref_detailed(text, default_sheet="") is None:
                        bound.add(text.upper())
            for v in node.values():
                walk(v)
        elif isinstance(node, list):
            for v in node:
                walk(v)

    walk(ast_dict)
    return bound


def _merge_rects(rects: list[Rect]) -> list[Rect]:
    seen: set[tuple] = set()
    out = []
    for r in rects:
        key = (r.sheet, r.r1, r.c1, r.r2, r.c2)
        if key not in seen:
            seen.add(key)
            out.append(r)
    return out


def _bbox_a1(grp: FormulaGroup) -> str:
    r1, c1, r2, c2 = grp.bbox
    if (r1, c1) == (r2, c2):
        return a1(r1, c1)
    return f"{a1(r1, c1)}:{a1(r2, c2)}"


def _chunk_rows(max_col: int) -> int:
    """How many rows to read per ``get_formulas`` call on a sheet that wide.

    Rows alone are the wrong unit: the engine returns a dense grid of Python
    strings, so 20,000 rows of a 16,384-column sheet is 327 million of them in
    a single call. Never zero — one row at a time is the floor, however wide.
    """
    return max(1, min(SCAN_CHUNK_ROWS, SCAN_CHUNK_CELLS // max(max_col, 1)))


def _spread_cells(cells: list[tuple[int, int]], n: int) -> list[tuple[int, int]]:
    """``n`` cells spread evenly over a group, in reading order.

    The head of a stretch is made of near-identical neighbours: sampling
    ``B2, C2, B3`` out of 400,000 cells says nothing about the far end, which
    is exactly where a pattern that broke — a hard-coded cell dropped into the
    middle of a column — shows up. First and last are always included.
    """
    ordered = sorted(cells)
    if n < 2:
        return ordered[: max(n, 0)]
    if len(ordered) <= n:
        return ordered
    last = len(ordered) - 1
    picks = sorted({round(i * last / (n - 1)) for i in range(n)})
    return [ordered[i] for i in picks]


def _sample_range_values(resolver: _ValueResolver, rect: Rect) -> list:
    if rect.sheet is None:
        return []
    out = []
    for r in range(rect.r1, min(rect.r2, rect.r1 + MAX_VALUE_SAMPLE - 1) + 1):
        for c in range(rect.c1, min(rect.c2, rect.c1 + MAX_VALUE_SAMPLE - 1) + 1):
            if len(out) >= MAX_VALUE_SAMPLE:
                return out
            value, source, date_text = resolver.value(rect.sheet, r, c)
            out.append(
                {
                    "addr": a1(r, c),
                    "value": value,
                    "source": source,
                    "date": date_text,
                }
            )
    return out


def _ref_preview(
    resolver: _ValueResolver,
    ref: str,
    sheet: str,
    defined_names: dict[str, list[Rect]] | None = None,
) -> tuple[Any, str | None]:
    """Preview of a referenced cell or range: ``(value, date_text)``."""
    # A reference into another workbook parses as nothing local; it still has a
    # value whenever that workbook was read, and the step is unreadable without.
    external = parse_external_refs(ref)
    if external:
        value, source = resolver.external_value(external[0])
        if source is not None:
            return _jsonable(value), None
    detail = parse_ref_detailed(ref, default_sheet=sheet)
    if detail is None:
        # may be a defined name: show the value of its target
        if defined_names:
            for name, rects in defined_names.items():
                if name.upper() == ref.upper() and rects:
                    rect = rects[0]
                    if rect.ncells == 1:
                        value, _, date_text = resolver.value(
                            rect.sheet or sheet, rect.r1, rect.c1
                        )
                        return value, date_text
                    return {"range": rect.to_a1(), "n": rect.ncells}, None
        return None, None
    rect = detail.rect
    if rect.ncells == 1:
        value, _, date_text = resolver.value(rect.sheet or sheet, rect.r1, rect.c1)
        return value, date_text
    return {"range": rect.to_a1(), "n": rect.ncells}, None


def _resolve_call(
    callee: str,
    caller_module: str,
    proc_ids: dict[str, str],
    procs_by_name: dict[str, list[str]],
) -> str | None:
    """Resolve a VBA call to a procedure node id.

    A callee already qualified with its module (``Module1.Taux``) resolves
    directly. VBA looks an unqualified name up in the calling module first,
    then in the other modules; a name declared by several other modules stays
    unresolved rather than pointing at an arbitrary one. All lookups are
    case-insensitive, as the language is.
    """
    if "." in callee:
        return proc_ids.get(callee.lower())
    same_module = f"{caller_module}.{callee}".lower()
    if same_module in proc_ids:
        return proc_ids[same_module]
    candidates = procs_by_name.get(callee.lower(), [])
    if len(candidates) == 1:
        return proc_ids[candidates[0]]
    return None


def _resolve_vba_write(
    rect: Rect, pid: str, sheet_dims, cell_owner, add_edge, ensure_input_node
) -> None:
    """A VBA write feeds the target cells: edge proc → target."""
    sheet = rect.sheet
    if sheet not in sheet_dims:
        opaque = ensure_input_node(rect, opaque_label=rect.to_a1())
        add_edge(pid, opaque, "vba-write")
        return
    clipped = rect.clipped(*sheet_dims[sheet]) or rect
    owners = cell_owner.get(sheet, {})
    seen: set[str] = set()
    has_plain = False
    if clipped.ncells <= SMALL_RANGE_CELLS:
        for r in range(clipped.r1, clipped.r2 + 1):
            for c in range(clipped.c1, clipped.c2 + 1):
                owner = owners.get((r, c))
                if owner is None:
                    has_plain = True
                elif owner not in seen:
                    seen.add(owner)
                    add_edge(pid, owner, "vba-write")
    else:
        has_plain = True
    if has_plain:
        add_edge(pid, ensure_input_node(clipped), "vba-write")
