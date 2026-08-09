"""Workbook context extraction and optional screenshot rendering."""

from __future__ import annotations

import datetime
import io
import math
import os
import re
import shutil
import struct
import subprocess
import sys
import tempfile
from collections.abc import Iterator
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from linexcel.refs import num_to_col

PREVIEW_ROWS = 12
PREVIEW_COLUMNS = 8
MAX_COMMENTS_PER_SHEET = 20
MAX_COMMENT_SCAN_CELLS = 100_000
MAX_COMMENT_CHARS = 1_000
MAX_MERGED_RANGES = 30
#: Under this, in either dimension, a rendered page carries nothing to look at.
MIN_SCREENSHOT_PIXELS = 8

# ``soffice.com`` first on Windows: it is the console front-end and waits for
# the conversion, while ``soffice.exe`` detaches and returns to the caller
# before the PDF exists.
_LAUNCHER_NAMES = (
    ("soffice.com", "soffice.exe", "soffice", "libreoffice")
    if sys.platform == "win32"
    else ("libreoffice", "soffice")
)

_INSTALL_HINTS = {
    "win32": (
        "winget install TheDocumentFoundation.LibreOffice"
        " and winget install oschwartz10612.Poppler"
        " (open a new terminal afterwards so PATH is refreshed)"
    ),
    "darwin": "brew install --cask libreoffice and brew install poppler",
}
_DEFAULT_INSTALL_HINT = "sudo apt install libreoffice-calc poppler-utils"

#: LibreOffice PDF export asking for one page per sheet. It is what ties an
#: image to a sheet at all: under the workbook's own print layout a long sheet
#: spans several pages and short ones share one, so page number and sheet have
#: no relation left to read off.
_PDF_FILTER_SINGLE_PAGE = (
    'pdf:calc_pdf_Export:{"SinglePageSheets":{"type":"boolean","value":"true"}}'
)


class WorkbookRenderError(RuntimeError):
    """Raised when the optional workbook screenshot renderer is unavailable."""


def extract_workbook_context(
    data: bytes,
    filename: str = "workbook.xlsx",
    *,
    preview_rows: int = PREVIEW_ROWS,
    preview_columns: int = PREVIEW_COLUMNS,
) -> dict[str, Any]:
    """Extract bounded presentation context without launching Excel.

    The preview preserves the first cells as they are, without guessing which
    row is a header. Comments and sheet layout markers complement formula
    lineage with the cues users usually see when opening a workbook.
    """
    workbook = load_workbook(
        io.BytesIO(data),
        read_only=False,
        data_only=False,
        keep_vba=filename.lower().endswith((".xlsm", ".xltm")),
    )
    warnings: list[str] = []
    sheets: list[dict[str, Any]] = []
    total_comments = 0
    try:
        for worksheet in workbook.worksheets:
            max_row = max(worksheet.max_row or 1, 1)
            max_column = max(worksheet.max_column or 1, 1)
            row_limit = min(max_row, preview_rows)
            column_limit = min(max_column, preview_columns)
            preview = [
                {
                    "row": row[0].row,
                    "values": [_safe_value(cell.value) for cell in row],
                }
                for row in worksheet.iter_rows(
                    min_row=1,
                    max_row=row_limit,
                    min_col=1,
                    max_col=column_limit,
                )
            ]
            comments, comments_truncated = _extract_comments(
                worksheet, max_row, max_column
            )
            total_comments += len(comments)
            if comments_truncated:
                warnings.append(
                    f"Comments on '{worksheet.title}' were truncated for inspection"
                )
            sheets.append(
                {
                    "name": worksheet.title,
                    "visibility": worksheet.sheet_state,
                    "dimensions": {"rows": max_row, "columns": max_column},
                    "preview_range": f"A1:{num_to_col(column_limit)}{row_limit}",
                    "preview": preview,
                    "freeze_panes": str(worksheet.freeze_panes)
                    if worksheet.freeze_panes
                    else None,
                    "merged_ranges": [
                        str(cell_range)
                        for cell_range in list(worksheet.merged_cells.ranges)[
                            :MAX_MERGED_RANGES
                        ]
                    ],
                    "hidden_columns": _hidden_columns(worksheet, column_limit),
                    "comments": comments,
                }
            )
    finally:
        workbook.close()
    return {
        "filename": filename,
        "sheets": sheets,
        "stats": {"sheets": len(sheets), "comments": total_comments},
        "warnings": warnings,
    }


def find_libreoffice() -> str | None:
    """Locate the LibreOffice launcher, on ``PATH`` or in a standard install.

    Windows and macOS installers do not put LibreOffice on ``PATH``, so a
    ``PATH``-only lookup reports the renderer as missing on machines where it is
    installed. Well-known install directories are therefore searched as well.
    """
    for name in _LAUNCHER_NAMES:
        found = shutil.which(name)
        if found:
            return found
    for candidate in _launcher_install_paths():
        if candidate.is_file():
            return str(candidate)
    return None


def find_pdftoppm() -> str | None:
    """Locate Poppler's ``pdftoppm``, on ``PATH`` or in a standard install."""
    found = shutil.which("pdftoppm")
    if found:
        return found
    for candidate in _pdftoppm_install_paths():
        if candidate.is_file():
            return str(candidate)
    return None


def render_workbook_screenshots(
    data: bytes,
    filename: str,
    output_dir: str | Path,
    *,
    dpi: int = 144,
    timeout: int = 180,
    per_sheet: bool = True,
) -> dict[str, list[Path]] | list[Path]:
    """Render workbook sheets to PNG with LibreOffice and Poppler.

    Works on Linux, macOS and Windows. LibreOffice runs headlessly; no desktop
    Excel process is needed. It exports the workbook to PDF, then ``pdftoppm``
    creates one PNG per rendered page.

    With ``per_sheet`` — the default — LibreOffice is asked to put each sheet on
    a single page, so the result is a ``{sheet name: [png]}`` mapping keyed by
    the workbook's own sheet names, and the report shows each image under the
    sheet it belongs to. Sheets are not split across pages, so a long one comes
    out as one tall image rather than as print pages nobody can map back.

    Setting ``per_sheet=False`` returns the flat ``list[Path]`` of print pages
    instead, as the page setup of the workbook lays them out.

    The mapping is only returned when LibreOffice produced exactly one page per
    sheet. When it did not — an older build ignoring the option, a page setup
    that overrides it — the flat page list is returned rather than a guessed
    mapping, because a screenshot filed under the wrong sheet is worse than one
    filed under none.
    """
    office = find_libreoffice()
    converter = find_pdftoppm()
    if not office or not converter:
        raise WorkbookRenderError(_missing_renderer_message(office, converter))
    if dpi <= 0:
        raise ValueError("dpi must be positive")
    if timeout <= 0:
        raise ValueError("timeout must be positive")

    target = Path(output_dir)
    target.mkdir(parents=True, exist_ok=True)
    suffix = Path(filename).suffix.lower()
    if suffix not in {".xls", ".xlsx", ".xlsm", ".xlsb", ".xltx", ".xltm"}:
        suffix = ".xlsx"
    stem = re.sub(r"[^A-Za-z0-9_.-]+", "-", Path(filename).stem).strip(".-")
    stem = stem or "workbook"
    sheet_names = _sheet_names(data) if per_sheet else []

    with tempfile.TemporaryDirectory(prefix="linexcel-render-") as temp_dir:
        temp = Path(temp_dir)
        input_path = temp / f"workbook{suffix}"
        pdf_dir = temp / "pdf"
        profile_dir = temp / "profile"
        input_path.write_bytes(data)
        pdf_dir.mkdir()
        try:
            subprocess.run(
                [
                    office,
                    # A throwaway profile: a LibreOffice already open on the
                    # desktop otherwise owns the default one, and the headless
                    # process exits 0 without converting anything.
                    f"-env:UserInstallation={profile_dir.as_uri()}",
                    "--headless",
                    "--norestore",
                    "--convert-to",
                    _PDF_FILTER_SINGLE_PAGE if per_sheet else "pdf",
                    "--outdir",
                    str(pdf_dir),
                    str(input_path),
                ],
                check=True,
                capture_output=True,
                text=True,
                timeout=timeout,
            )
        except subprocess.TimeoutExpired as exc:
            raise WorkbookRenderError(
                f"LibreOffice did not finish within {timeout} seconds"
            ) from exc
        except subprocess.CalledProcessError as exc:
            details = (exc.stderr or exc.stdout or "unknown error").strip()
            raise WorkbookRenderError(
                f"LibreOffice could not render the workbook: {details}"
            ) from exc

        pdfs = list(pdf_dir.glob("*.pdf"))
        if not pdfs:
            raise WorkbookRenderError("LibreOffice did not produce a PDF")
        # Rendered aside, then moved in: pdftoppm pads the page number to the
        # width of the page count, so a directory reused across two workbooks
        # holds both "-1.png" and "-01.png" and a glob over it would return the
        # previous run's pages alongside this one's.
        png_dir = temp / "png"
        png_dir.mkdir()
        try:
            subprocess.run(
                [converter, "-png", "-r", str(dpi), str(pdfs[0]), str(png_dir / stem)],
                check=True,
                capture_output=True,
                text=True,
                timeout=timeout,
            )
        except subprocess.TimeoutExpired as exc:
            raise WorkbookRenderError(
                f"PDF conversion did not finish within {timeout} seconds"
            ) from exc
        except subprocess.CalledProcessError as exc:
            details = (exc.stderr or exc.stdout or "unknown error").strip()
            raise WorkbookRenderError(
                f"pdftoppm could not create screenshots: {details}"
            ) from exc
        pages = sorted(png_dir.glob(f"{stem}-*.png"))
        if not pages:
            raise WorkbookRenderError("pdftoppm did not produce PNG screenshots")
        if per_sheet and sheet_names and len(pages) == len(sheet_names):
            return _place_by_sheet(pages, sheet_names, target, stem)
        return _place_pages(pages, target, stem)


def _sheet_names(data: bytes) -> list[str]:
    """Sheet names in workbook order — the order LibreOffice paginates in.

    Hidden sheets included: LibreOffice gives them a page like any other, so
    dropping them here would shift every name onto the wrong image.
    """
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        return []
    try:
        return [worksheet.title for worksheet in workbook.worksheets]
    finally:
        workbook.close()


def _place_pages(pages: list[Path], target: Path, stem: str) -> list[Path]:
    """Move rendered pages into the output directory, numbered in page order."""
    width = len(str(len(pages)))
    return [
        _move(page, target / f"{stem}-{index:0{width}d}.png")
        for index, page in enumerate(pages, start=1)
    ]


def _place_by_sheet(
    pages: list[Path], sheet_names: list[str], target: Path, stem: str
) -> dict[str, list[Path]]:
    """Move each page to a file named after the sheet it renders.

    The mapping is keyed by the sheet's real name; only the file name is
    sanitized, and two sheets whose names sanitize alike are kept apart by a
    suffix so neither overwrites the other. A sheet with nothing on it renders
    as a page one pixel tall and is left out: shown at the report's width it
    reads as a broken image, and its card already says the sheet is empty.
    """
    by_sheet: dict[str, list[Path]] = {}
    taken: set[str] = set()
    for index, (page, name) in enumerate(zip(pages, sheet_names), start=1):
        width, height = _png_size(page)
        if min(width, height) < MIN_SCREENSHOT_PIXELS:
            continue
        slug = re.sub(r"[^A-Za-z0-9_.-]+", "-", name).strip(".-") or f"sheet{index}"
        unique, attempt = slug, 2
        while unique in taken:
            unique, attempt = f"{slug}-{attempt}", attempt + 1
        taken.add(unique)
        by_sheet[name] = [_move(page, target / f"{stem}-{unique}.png")]
    return by_sheet


def _move(source: Path, destination: Path) -> Path:
    destination.unlink(missing_ok=True)
    shutil.move(str(source), str(destination))
    return destination


def _png_size(path: Path) -> tuple[int, int]:
    """``(width, height)`` from a PNG's IHDR, ``(0, 0)`` if unreadable."""
    try:
        header = path.read_bytes()[:24]
    except OSError:
        return 0, 0
    if header[:8] != b"\x89PNG\r\n\x1a\n" or len(header) < 24:
        return 0, 0
    try:
        return struct.unpack(">II", header[16:24])
    except struct.error:
        return 0, 0


def _program_roots() -> Iterator[Path]:
    """Windows directories that hold per-machine and per-user installs."""
    seen: set[str] = set()
    for variable in ("ProgramW6432", "ProgramFiles", "ProgramFiles(x86)"):
        root = os.environ.get(variable)
        if root and root not in seen:
            seen.add(root)
            yield Path(root)
    local = os.environ.get("LOCALAPPDATA")
    if local:
        yield Path(local) / "Programs"


def _launcher_install_paths() -> Iterator[Path]:
    if sys.platform == "win32":
        for root in _program_roots():
            program = root / "LibreOffice" / "program"
            yield program / "soffice.com"
            yield program / "soffice.exe"
    elif sys.platform == "darwin":
        yield Path("/Applications/LibreOffice.app/Contents/MacOS/soffice")


def _pdftoppm_install_paths() -> Iterator[Path]:
    if sys.platform != "win32":
        return
    # Poppler ships as a zip rather than an installer, so it lands wherever the
    # package manager unpacked it. winget appends that directory to the user
    # PATH in the registry, which existing shells do not see until they are
    # restarted -- hence looking inside its package store directly.
    local = os.environ.get("LOCALAPPDATA")
    if local:
        winget = Path(local) / "Microsoft" / "WinGet"
        yield winget / "Links" / "pdftoppm.exe"
        yield from winget.glob("Packages/*Poppler*/*/Library/bin/pdftoppm.exe")
    for root in _program_roots():
        yield from root.glob("poppler*/Library/bin/pdftoppm.exe")
        yield from root.glob("poppler*/bin/pdftoppm.exe")


def _missing_renderer_message(office: str | None, converter: str | None) -> str:
    missing = ", ".join(
        name
        for name, found in (("LibreOffice", office), ("pdftoppm", converter))
        if not found
    )
    hint = _INSTALL_HINTS.get(sys.platform, _DEFAULT_INSTALL_HINT)
    return (
        f"Workbook screenshots require LibreOffice and pdftoppm; {missing} "
        f"could not be found. Install with: {hint}"
    )


def _safe_value(value: Any) -> Any:
    # A date cell reads back as a midnight datetime, and "2026-01-03T00:00:00"
    # is a timestamp nobody typed: the preview and the AI dossier both show what
    # the cell holds, which is a day.
    if isinstance(value, datetime.datetime):
        if value.time() == datetime.time.min:
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, (datetime.date, datetime.time)):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _extract_comments(
    worksheet, max_row: int, max_column: int
) -> tuple[list[dict], bool]:
    scan_rows = min(max_row, math.ceil(MAX_COMMENT_SCAN_CELLS / max_column))
    comments: list[dict[str, str | None]] = []
    for row in worksheet.iter_rows(
        min_row=1, max_row=scan_rows, min_col=1, max_col=max_column
    ):
        for cell in row:
            if cell.comment is None:
                continue
            comments.append(
                {
                    "cell": cell.coordinate,
                    "author": cell.comment.author,
                    "text": cell.comment.text[:MAX_COMMENT_CHARS],
                }
            )
            if len(comments) >= MAX_COMMENTS_PER_SHEET:
                return comments, True
    return comments, max_row > scan_rows


def _hidden_columns(worksheet, column_limit: int) -> list[str]:
    return [
        num_to_col(column)
        for column in range(1, column_limit + 1)
        if worksheet.column_dimensions[num_to_col(column)].hidden
    ]
