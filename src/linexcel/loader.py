"""Reading the values a workbook already carries, and how big it claims to be.

Every formula cell of a saved workbook holds two things: the formula, and the
result the spreadsheet application last computed for it. This module reads the
second — the value the user actually saw on screen — which is what the report
compares its own recalculation against, and the only source for a cell no
engine can compute.

Two readers, on purpose. python-calamine is the fast one and returns native
Python types, so a date is a date rather than a number wearing a format
string. openpyxl is the lazy one, and it takes over for a file calamine cannot
open or must not be handed: it builds a sheet as a dense rows × columns array
before returning anything, and a workbook declaring `A1:XFD1048576` would ask
the allocator for 512 GiB. That is why :func:`declared_cells` exists, and why
it is consulted before anything reads a cell.
"""

from __future__ import annotations

import datetime
import io
import re
import zipfile
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles.numbers import is_date_format

from linexcel.progress import Reporter
from linexcel.refs import col_to_num

#: Ceiling on the formula sweep of a single sheet. It is not there to keep the
#: analysis quick — sweeping costs about 0.7 µs per cell, so the 2.4M cells
#: this used to cut from the reference workbook bought 1.6 s of a 95 s run,
#: and cost the sheet its tail. It is there for the file that *declares* far
#: more than it holds: one stray cell at XFD1048576 makes the used range 17
#: billion cells, which would sweep for hours. At 64M the ceiling is past any
#: real used range (a full-height sheet 60 columns wide) and still bounded.
MAX_CELLS_PER_SHEET = 64_000_000
#: The largest declared rectangle python-calamine is allowed to attempt.
#: It builds a sheet as a dense rows × columns array — about 32 bytes a cell —
#: before returning anything to Python, so the limit is memory, not time, and
#: it is separate from the sweep budget above. 20M cells is roughly 640 MB:
#: past any honest sheet's dense form, and far below the 512 GiB that one
#: stray cell at XFD1048576 asks for. Beyond it the lazy reader takes over.
MAX_DENSE_CELLS = 20_000_000


class CachedValues:
    """Values the spreadsheet application cached in the file.

    They are what the user last saw on screen. Workbooks written by openpyxl
    carry no cache for formula cells, workbooks saved by Excel or LibreOffice
    do; either way constants and their number formats are always readable.
    """

    def __init__(
        self,
        values: dict[tuple[str, int, int], Any],
        date_cells: set[tuple[str, int, int]],
        epoch_1904: bool,
    ):
        self._values = values
        self._date_cells = date_cells
        self.epoch_1904 = epoch_1904

    def get(self, sheet: str, row: int, col: int) -> Any:
        return self._values.get((sheet, row, col))

    def is_date(self, sheet: str, row: int, col: int) -> bool:
        return (sheet, row, col) in self._date_cells

    def __len__(self) -> int:
        return len(self._values)


#: ``<dimension ref="A1:XFD1048576"/>``, the rectangle a sheet claims to use.
#: Only the bottom-right corner matters; the ``A1:`` half is absent on a sheet
#: holding a single cell.
_DIMENSION_RE = re.compile(
    rb'<dimension[^>]*\sref="(?:[A-Z]{1,3}\d+:)?([A-Z]{1,3})(\d+)"'
)
#: How much of a sheet part is read looking for it. ``<dimension>`` is the
#: first child of ``<worksheet>``, so it is within the first few hundred bytes
#: of any file Excel wrote.
_DIMENSION_WINDOW = 4096


def declared_cells(data: bytes) -> int:
    """The largest rectangle any sheet of the package *declares*, in cells.

    Not what it holds — what it says it uses. The two differ wildly: one stray
    cell at XFD1048576, and a sheet with three numbers in it declares 17
    billion. Read from the ``<dimension>`` element rather than from the cells,
    because the whole point is to know the size before reading anything.

    ``0`` when no sheet declares one, which is also what a package this cannot
    parse returns: the callers treat it as "no reason to worry", since a writer
    that omits ``<dimension>`` is not the one that writes a stray corner cell.
    """
    largest = 0
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            for part in zf.namelist():
                if not re.fullmatch(r"xl/worksheets/sheet\d+\.xml", part):
                    continue
                with zf.open(part) as handle:
                    match = _DIMENSION_RE.search(handle.read(_DIMENSION_WINDOW))
                if match is None:
                    continue
                columns = col_to_num(match.group(1).decode("ascii"))
                largest = max(largest, columns * int(match.group(2)))
    except Exception:
        return 0
    return largest


#: What to tell someone whose file declares far more than it holds. It is a
#: real spreadsheet, made by a real Excel: a cell was touched once in the far
#: corner, and the used range never shrank back.
STRAY_CORNER_ADVICE = (
    "Select the rows below and the columns to the right of your data, delete "
    "them, then save — the used range shrinks back and the analysis takes the "
    "fast path again."
)


def load_cached_values(
    data: bytes,
    warnings: list[str] | None = None,
    reporter: Reporter | None = None,
) -> CachedValues:
    """Read the file's cached values once, keyed by (sheet, row, col).

    python-calamine (the Rust engine) is the hot path: it returns native Python
    types directly, so dates are detected by type instead of by number_format
    string, and it is roughly an order of magnitude faster than openpyxl on
    large files. openpyxl remains the fallback for the rare file calamine cannot
    open; there its number_format-based date detection keeps the edge case (a
    number formatted as a date but stored as a float) covered.

    A sheet that declares more than :data:`MAX_DENSE_CELLS` never reaches
    calamine at all. It builds a sheet as a dense rows × columns array before
    returning anything to Python, so ``A1:XFD1048576`` asks the allocator for
    512 GiB — and an allocation failure in Rust *aborts the process*. That is
    not an exception, and no ``try`` around this call would see it. openpyxl's
    read-only reader is lazy and already bounded, so it takes the file instead.
    """
    declared = declared_cells(data)
    if declared > MAX_DENSE_CELLS:
        if warnings is not None:
            warnings.append(
                f"A sheet declares a used range of {declared:,} cells. Values "
                f"were read the slow way, and only the first "
                f"{MAX_CELLS_PER_SHEET:,} cells of each sheet were kept, so "
                f"some may be missing from the report. If the sheet does not "
                f"really hold that much: {STRAY_CORNER_ADVICE}"
            )
        return _load_cached_values_openpyxl(data, reporter)
    try:
        return _load_cached_values_calamine(data, reporter)
    except Exception as exc:
        # Not silent: the slow reader detects dates from the number format
        # rather than from the type, which is a different answer on an edge
        # case, and someone comparing two runs deserves to know which read it.
        if warnings is not None:
            warnings.append(
                f"Values were read with openpyxl rather than the fast reader "
                f"({type(exc).__name__}: {exc}). The lineage is unaffected; a "
                f"cell whose date is stored as a plain number may read "
                f"differently."
            )
        return _load_cached_values_openpyxl(data, reporter)


def _load_cached_values_calamine(
    data: bytes, reporter: Reporter | None = None
) -> CachedValues:
    """Fast path: read cached values via python-calamine.

    ``to_python(skip_empty_area=False)`` preserves the cell coordinates —
    linexcel keys values by ``(sheet, row, col)`` 1-indexed, and the default
    ``skip_empty_area=True`` would shift everything toward the top-left.
    Empty cells come back as ``""`` (calamine) or ``None``; both are skipped,
    matching openpyxl which reads an empty cell back as ``None``.

    Dates are detected by type: calamine returns ``datetime.date`` for
    date-only cells and ``datetime.datetime`` for timestamps. openpyxl always
    returns ``datetime.datetime`` (even for a date-only cell), so a bare
    ``datetime.date`` is normalized to a midnight ``datetime.datetime`` to keep
    downstream equality checks and the resolver seeing one consistent type.
    """
    from python_calamine import CalamineWorkbook

    values: dict[tuple[str, int, int], Any] = {}
    date_cells: set[tuple[str, int, int]] = set()
    epoch_1904 = _detect_epoch_1904(data)
    wb = CalamineWorkbook.from_object(io.BytesIO(data))
    bar = (reporter or Reporter()).phase("cached values", total=len(wb.sheet_names))
    with bar as progress:
        for name in wb.sheet_names:
            progress.step(f"reading {name}")
            sheet = wb.get_sheet_by_name(name)
            rows = sheet.to_python(skip_empty_area=False)
            scanned = 0
            for r_idx, row in enumerate(rows):
                scanned += len(row)
                if scanned > MAX_CELLS_PER_SHEET:
                    break
                for c_idx, v in enumerate(row):
                    if v is None or v == "":
                        continue
                    key = (name, r_idx + 1, c_idx + 1)
                    if isinstance(v, datetime.datetime):
                        date_cells.add(key)
                    elif isinstance(v, datetime.date):
                        # openpyxl reads a date-only cell as a midnight datetime;
                        # normalize so the two readers are interchangeable.
                        v = datetime.datetime(v.year, v.month, v.day)
                        date_cells.add(key)
                    values[key] = v
    return CachedValues(values, date_cells, epoch_1904)


def _stepped(phase_cm, items, verb: str):
    """Iterate ``items`` inside a reporter phase, one step each.

    A generator rather than a ``with`` block around the loop: the loop bodies
    it wraps are long, and re-indenting them to gain a bar is a diff nobody
    can review against the logic it contains.
    """
    with phase_cm as progress:
        for item in items:
            progress.step(f"{verb} {_label_of(item)}")
            yield item


def _label_of(item) -> str:
    """What to call one item on the progress line."""
    title = getattr(item, "title", None)
    if title is not None:
        return str(title)
    # (node id, group) pairs: the id is what a reader would recognise
    if isinstance(item, tuple) and item:
        return str(item[0])
    return str(item)


def _load_cached_values_openpyxl(
    data: bytes, reporter: Reporter | None = None
) -> CachedValues:
    """Fallback: read cached values via openpyxl with number_format date detection."""
    values: dict[tuple[str, int, int], Any] = {}
    date_cells: set[tuple[str, int, int]] = set()
    epoch_1904 = False
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        return CachedValues(values, date_cells, epoch_1904)
    try:
        epoch_1904 = getattr(wb.epoch, "year", 1899) == 1904
        bar = (reporter or Reporter()).phase("cached values", total=len(wb.worksheets))
        for ws in _stepped(bar, wb.worksheets, "reading"):
            scanned = 0
            for row in ws.iter_rows():
                scanned += len(row)
                if scanned > MAX_CELLS_PER_SHEET:
                    break
                for cell in row:
                    # read-only sheets pad gaps with EmptyCell (no coordinates)
                    r = getattr(cell, "row", None)
                    c = getattr(cell, "column", None)
                    if r is None or c is None:
                        continue
                    key = (ws.title, r, c)
                    if _is_date_format(getattr(cell, "number_format", None)):
                        date_cells.add(key)
                    if cell.value is not None:
                        values[key] = cell.value
    except Exception:
        pass
    finally:
        wb.close()
    return CachedValues(values, date_cells, epoch_1904)


def _detect_epoch_1904(data: bytes) -> bool:
    """True if the workbook declares the 1904 date system.

    calamine converts 1904-epoch dates to correct values itself, but linexcel
    also needs the flag to interpret *engine* serials (from formualizer, not
    calamine) via :func:`serial_to_date_text`. The flag lives on
    ``<workbookPr date1904="1" />`` in ``xl/workbook.xml``; reading it from the
    zip avoids opening openpyxl a second time just for this one attribute.
    """
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            xml = zf.read("xl/workbook.xml").decode("utf-8", "ignore")
    except Exception:
        return False
    return bool(re.search(r"""date1904=["']1""", xml))


def _is_date_format(number_format: Any) -> bool:
    if not number_format:
        return False
    try:
        return bool(is_date_format(number_format))
    except Exception:
        return False
