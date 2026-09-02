"""What the file says about itself, before anything analyses it.

Extracted mechanically from analyzer.py: sheet dimensions, defined names and
the lightweight pre-analysis (``inspect_workbook``) that reads only package
headers, never formulas or values.
"""

from __future__ import annotations

import io
import re
import zipfile
from dataclasses import dataclass
from typing import Any

from openpyxl import load_workbook

from linexcel.external import read_external_links
from linexcel.loader import MAX_CELLS_PER_SHEET, MAX_DENSE_CELLS, declared_cells
from linexcel.tables import _collect_defined_names, _force_dimensions

MAX_NODES_PER_SHEET = 400

#: Seconds per megabyte of uncompressed sheet XML. Measured across workbooks
#: from a thousand cells to two hundred thousand formulas, where the cost per
#: byte stays near enough constant to be worth quoting: what an analysis
#: really costs tracks how much formula there is to read, and that is what the
#: sheet parts weigh. A file of mostly values is faster than this says, which
#: is the right direction for a warning to be wrong in.
SECONDS_PER_SHEET_MB = 0.5
#: Below this the estimate is noise and nobody was going to wait anyway.
WORTH_MENTIONING_SECONDS = 5.0

_SHEET_PART_RE = re.compile(r"xl/worksheets/sheet\d+\.xml")


@dataclass
class Structure:
    sheet_dims: dict[str, tuple[int, int]]
    defined_names: dict[str, list]


def read_structure(data: bytes) -> Structure:
    owb = load_workbook(io.BytesIO(data), read_only=True, data_only=False)
    try:
        sheet_dims: dict[str, tuple[int, int]] = {}
        for ws in owb.worksheets:
            max_row, max_col = ws.max_row, ws.max_column
            if not max_row or not max_col:
                max_row, max_col = _force_dimensions(ws)
            sheet_dims[ws.title] = (max_row or 1, max_col or 1)
        defined_names = _collect_defined_names(owb)
    finally:
        owb.close()
    return Structure(sheet_dims, defined_names)


def sheet_bytes(data: bytes) -> int:
    """Uncompressed weight of the sheet parts, without unpacking one.

    The zip central directory carries each entry's real size, so this costs a
    read of the index — microseconds on a file that takes minutes to analyse.
    That is the whole point: an estimate nobody waits for.
    """
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            return sum(
                entry.file_size
                for entry in zf.infolist()
                if _SHEET_PART_RE.fullmatch(entry.filename)
            )
    except Exception:
        return 0


def inspect_workbook(data: bytes) -> dict[str, Any]:
    """What the file says about itself, before anything analyses it.

    Everything here is read from the package headers — sheet dimensions and
    external link declarations — so it costs milliseconds on a file that would
    take minutes to analyse. That is the point: it answers "is this going to
    be long, and will anything be left out?" *before* someone commits to
    finding out the slow way.

    Declared sizes, not real ones. A sheet claiming 17 billion cells holds
    nothing of the sort, and saying so is exactly the warning worth having.
    """
    owb = load_workbook(io.BytesIO(data), read_only=True, data_only=False)
    try:
        sheets = []
        for ws in owb.worksheets:
            max_row, max_col = ws.max_row, ws.max_column
            if not max_row or not max_col:
                max_row, max_col = _force_dimensions(ws)
            rows, cols = max_row or 1, max_col or 1
            sheets.append(
                {
                    "name": ws.title,
                    "rows": rows,
                    "cols": cols,
                    "cells": rows * cols,
                    "state": ws.sheet_state,
                    "truncated": rows * cols > MAX_CELLS_PER_SHEET,
                }
            )
    finally:
        owb.close()
    books = read_external_links(data)
    weight = sheet_bytes(data)
    return {
        "bytes": len(data),
        "sheetBytes": weight,
        "estimatedSeconds": round(weight / 1_048_576 * SECONDS_PER_SHEET_MB, 1),
        "sheets": sheets,
        "declaredCells": sum(s["cells"] for s in sheets),
        "externalWorkbooks": [b.name for b in books.values()],
        "densePathRefused": declared_cells(data) > MAX_DENSE_CELLS,
        "ceilings": {
            "cellsPerSheet": MAX_CELLS_PER_SHEET,
            "nodesPerSheet": MAX_NODES_PER_SHEET,
            "denseCells": MAX_DENSE_CELLS,
        },
    }
