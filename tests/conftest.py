import io

import pytest


def build_lineage_workbook() -> bytes:
    """Rich workbook for lineage tests: stretched formulas,
    cross-sheet references, defined name, composite functions."""
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.workbook.defined_name import DefinedName

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventes"
    ws["A1"], ws["B1"], ws["C1"], ws["D1"] = "Produit", "Qté", "Prix", "CA"
    for r in range(2, 102):
        ws.cell(row=r, column=1, value=f"P{r - 1}")
        ws.cell(row=r, column=2, value=r % 7 + 1)
        ws.cell(row=r, column=3, value=10.5 + (r % 13))
        ws.cell(row=r, column=4, value=f"=B{r}*C{r}")  # stretched formula ×100
    ws["A1"].comment = Comment("Exported product category", "Data team")
    ws.freeze_panes = "A2"
    ws.column_dimensions["C"].hidden = True
    ws.merge_cells("F1:G1")
    ws["F1"] = "Presentation context"

    syn = wb.create_sheet("Synthese")
    syn["B1"] = "=SUM(Ventes!D2:D101)"
    syn["B2"] = "=ROUND(AVERAGE(Ventes!D2:D101), 2)"
    syn["B3"] = (
        "=IF(SUM(Ventes!D2:D101)>TauxCible, "
        'CONCATENATE("OK: ", ROUND(B1/1000,1), "k"), "KO")'
    )

    params = wb.create_sheet("Params")
    params["A1"] = 5000
    wb.defined_names.add(DefinedName("TauxCible", attr_text="Params!$A$1"))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture()
def lineage_excel() -> bytes:
    return build_lineage_workbook()
