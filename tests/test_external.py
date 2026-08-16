"""Tests for workbooks a file depends on, and how far linexcel follows them.

A workbook that reads another file is the common case in practice, and the one
where a lineage tool most easily lies: the engine cannot follow the link, so
every value above it is missing or stale. These check the three answers —
named, cached, resolved — and that the report says which one it is giving.
"""

import io
from pathlib import Path

import pytest
from openpyxl import Workbook

from linexcel import analyze
from linexcel.analyzer import analyze_workbook
from linexcel.external import (
    find_workbooks,
    macro_files,
    parse_external_refs,
    read_external_links,
)


def workbook(cells: dict[str, object], sheet: str = "S") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for address, value in cells.items():
        ws[address] = value
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def referenced(tmp_path: Path) -> Path:
    """A folder holding ``Ref.xlsx``, with values worth reading."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["B2"] = 21
    ws["B3"] = "Contoso"
    wb.save(tmp_path / "Ref.xlsx")
    return tmp_path


def node_of(graph: dict, node_id: str) -> dict:
    return next(n for n in graph["nodes"] if n["id"] == node_id)


class TestParsing:
    @pytest.mark.parametrize(
        ("formula", "expected"),
        [
            ("='[1]Annual'!B4", ("1", "Annual", "B4", "")),
            ("=[1]Annual!B4", ("1", "Annual", "B4", "")),
            ("='[Budget FY26.xlsx]Q1'!$A$1", ("Budget FY26.xlsx", "Q1", "A1", "")),
            (
                "='C:\\Reports\\[Old.xlsx]Sheet1'!A1",
                ("Old.xlsx", "Sheet1", "A1", "C:\\Reports\\"),
            ),
        ],
    )
    def test_a_reference_is_read_apart(self, formula, expected):
        (ref,) = parse_external_refs(formula)
        assert (ref.book, ref.sheet, ref.cell, ref.directory) == expected

    def test_a_local_formula_names_no_workbook(self):
        assert parse_external_refs("=SUM(A1:A3) + Sheet2!B1") == []

    def test_a_structured_reference_is_not_an_external_one(self):
        """``SalesTable[Revenue]`` brackets a column, not a file."""
        assert parse_external_refs("=SUM(SalesTable[Revenue])") == []


class TestDeclaredLinks:
    def test_the_declared_path_is_read_from_the_package(self):
        books = read_external_links(Path("validation_stress.xlsx").read_bytes())
        assert {b.name for b in books.values()} == {"Budget FY26.xlsx", "Old.xlsx"}
        assert books["1"].target.endswith("Budget FY26.xlsx")

    def test_a_workbook_without_links_declares_none(self):
        assert read_external_links(workbook({"A1": 1})) == {}


class TestNamingWithoutResolving:
    """Without the folder, the answer is *which* file — not silence."""

    def test_the_node_names_the_workbook_it_depends_on(self):
        graph = analyze_workbook(
            workbook({"A1": 2, "B1": "='[Ref.xlsx]Data'!B2 * A1"}), "main.xlsx"
        )["graph"]
        (book,) = node_of(graph, "c:S!B1")["externalBooks"]
        assert book["name"] == "Ref.xlsx"
        assert book["read"] == "none"

    def test_the_grey_node_is_labelled_by_file_rather_than_by_index(self):
        graph = analyze_workbook(Path("validation_stress.xlsx").read_bytes(), "s.xlsx")[
            "graph"
        ]
        labels = {n["label"] for n in graph["nodes"] if n["kind"] == "opaque"}
        assert "[Budget FY26.xlsx]Annual!B4" in labels
        assert "'[1]Annual'!B4" not in labels

    def test_the_warning_names_the_files_and_the_way_out(self):
        graph = analyze_workbook(
            workbook({"A1": 2, "B1": "='[Ref.xlsx]Data'!B2 * A1"}), "main.xlsx"
        )["graph"]
        (warning,) = [w for w in graph["meta"]["warnings"] if "external workbook" in w]
        assert "Ref.xlsx" in warning
        assert "--refs-dir" in warning

    def test_the_stats_count_what_is_read_and_what_is_not(self):
        graph = analyze_workbook(
            workbook({"A1": 2, "B1": "='[Ref.xlsx]Data'!B2 * A1"}), "main.xlsx"
        )["graph"]
        assert graph["meta"]["stats"]["externalWorkbooks"] == 1
        assert graph["meta"]["stats"]["externalWorkbooksRead"] == 0


class TestResolvingFromAFolder:
    """With the folder, the reference is evaluated as the value it stands for."""

    def test_a_referenced_value_reaches_the_formula_above_it(self, referenced):
        graph = analyze_workbook(
            workbook({"A1": 2, "B1": "='[Ref.xlsx]Data'!B2 * A1"}),
            "main.xlsx",
            refs_dir=referenced,
        )["graph"]
        node = node_of(graph, "c:S!B1")
        assert node["value"] == 42
        assert node["valueSource"] == "external"

    def test_text_crosses_the_link_as_well_as_numbers(self, referenced):
        graph = analyze_workbook(
            workbook({"B1": "='[Ref.xlsx]Data'!B3"}), "main.xlsx", refs_dir=referenced
        )["graph"]
        assert node_of(graph, "c:S!B1")["value"] == "Contoso"

    def test_the_same_cell_without_the_folder_has_no_value(self):
        graph = analyze_workbook(
            workbook({"A1": 2, "B1": "='[Ref.xlsx]Data'!B2 * A1"}), "main.xlsx"
        )["graph"]
        assert node_of(graph, "c:S!B1")["value"] is None

    def test_the_node_says_where_the_file_was_read_from(self, referenced):
        graph = analyze_workbook(
            workbook({"B1": "='[Ref.xlsx]Data'!B2"}), "main.xlsx", refs_dir=referenced
        )["graph"]
        (book,) = node_of(graph, "c:S!B1")["externalBooks"]
        assert book["read"] == "folder"
        assert book["file"].endswith("Ref.xlsx")

    def test_the_grey_node_carries_the_value_it_stands_for(self, referenced):
        graph = analyze_workbook(
            workbook({"B1": "='[Ref.xlsx]Data'!B2"}), "main.xlsx", refs_dir=referenced
        )["graph"]
        node = node_of(graph, "x:[Ref.xlsx]Data!B2")
        assert node["value"] == 21
        assert node["valueSource"] == "external"

    def test_a_missing_file_is_reported_rather_than_guessed(self, tmp_path):
        graph = analyze_workbook(
            workbook({"B1": "='[Nowhere.xlsx]Data'!B2"}),
            "main.xlsx",
            refs_dir=tmp_path,
        )["graph"]
        assert node_of(graph, "c:S!B1")["value"] is None
        (book,) = node_of(graph, "c:S!B1")["externalBooks"]
        assert book["read"] == "none"

    def test_the_public_api_takes_the_folder(self, referenced):
        result = analyze(
            workbook({"B1": "='[Ref.xlsx]Data'!B2"}),
            filename="main.xlsx",
            refs_dir=referenced,
        )
        assert result.stats["externalWorkbooksRead"] == 1


class TestFolderScanning:
    def test_only_workbooks_are_picked_up(self, tmp_path):
        (tmp_path / "notes.txt").write_text("not a workbook", encoding="utf-8")
        (tmp_path / "Book.xlsx").write_bytes(workbook({"A1": 1}))
        assert set(find_workbooks(tmp_path)) == {"book.xlsx"}

    def test_a_subfolder_is_searched_too(self, tmp_path):
        nested = tmp_path / "add-ins"
        nested.mkdir()
        (nested / "Deep.xlsx").write_bytes(workbook({"A1": 1}))
        assert "deep.xlsx" in find_workbooks(tmp_path)

    def test_macro_files_are_singled_out_for_their_code(self, tmp_path):
        (tmp_path / "Plain.xlsx").write_bytes(workbook({"A1": 1}))
        (tmp_path / "Tools.xlam").write_bytes(
            (Path("tests/fixtures/macros.xlsm")).read_bytes()
        )
        assert [p.name for p in macro_files(tmp_path)] == ["Tools.xlam"]


class TestAddInCode:
    def test_vba_in_the_folder_becomes_part_of_the_lineage(self, tmp_path):
        """Code a workbook calls often lives in an add-in, not in the file."""
        (tmp_path / "Tools.xlam").write_bytes(
            Path("tests/fixtures/macros.xlsm").read_bytes()
        )
        graph = analyze_workbook(workbook({"A1": 1}), "main.xlsx", refs_dir=tmp_path)[
            "graph"
        ]
        vba = [n for n in graph["nodes"] if n["kind"] == "vba"]
        assert vba, "the add-in's procedures should be nodes"
        assert all("Tools.xlam:" in n["module"] for n in vba)

    def test_nothing_is_read_from_disk_without_the_folder(self, tmp_path):
        (tmp_path / "Tools.xlam").write_bytes(
            Path("tests/fixtures/macros.xlsm").read_bytes()
        )
        graph = analyze_workbook(workbook({"A1": 1}), "main.xlsx")["graph"]
        assert not [n for n in graph["nodes"] if n["kind"] == "vba"]
