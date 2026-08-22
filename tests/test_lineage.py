"""Tests for the lineage module: references, grouping, graph, VBA, API."""

import io
import json
import re
import struct
from pathlib import Path
from typing import Any, cast

import pytest
from openpyxl.comments import Comment

from linexcel import LineageResult, analyze
from linexcel import viewer as viewer_module
from linexcel.analyzer import analyze_workbook
from linexcel.refs import (
    Rect,
    col_to_num,
    num_to_col,
    parse_ref,
    parse_ref_detailed,
    ref_to_r1c1,
    stretch_ref,
)
from linexcel.rewrite import canonical_r1c1, qualify_sheet
from linexcel.vba import analyze_vba

#: Everything :func:`linexcel.aidoc._resolve_provider` reads. A developer with
#: one of these exported would otherwise silently configure a provider for the
#: tests that assert nothing is configured.
_AI_ENV_VARS = (
    "LINEXCEL_AI_BASE_URL",
    "LINEXCEL_AI_MODEL",
    "LINEXCEL_AI_API_KEY",
    "OPENAI_BASE_URL",
    "OPENAI_MODEL",
    "OPENAI_API_KEY",
)


def _clear_ai_env(monkeypatch) -> None:
    for var in _AI_ENV_VARS:
        monkeypatch.delenv(var, raising=False)


def _strip_comments_and_docstrings(source: str) -> str:
    """Return ``source`` with comments and docstrings removed.

    Prose is free to name endpoints as examples; executable code is not. Other
    string literals are kept on purpose — a vendor name smuggled in as an
    environment variable or a default model is exactly what must be caught.
    """
    import ast
    import io
    import tokenize

    docstrings = set()
    for node in ast.walk(ast.parse(source)):
        body = getattr(node, "body", None)
        if not isinstance(node, ast.Module | ast.ClassDef | ast.FunctionDef):
            continue
        if not body or not isinstance(body[0], ast.Expr):
            continue
        first = body[0].value
        if isinstance(first, ast.Constant) and isinstance(first.value, str):
            docstrings.add((first.lineno, first.col_offset))

    kept = []
    for token in tokenize.generate_tokens(io.StringIO(source).readline):
        if token.type == tokenize.COMMENT:
            continue
        if token.type == tokenize.STRING and token.start in docstrings:
            continue
        kept.append(token.string)
    return "\n".join(kept)


class TestRefs:
    def test_col_roundtrip(self):
        for col in ("A", "Z", "AA", "XFD"):
            assert num_to_col(col_to_num(col)) == col

    def test_parse_cell(self):
        rect = parse_ref("B2", "S")
        assert rect is not None
        assert rect == Rect("S", 2, 2, 2, 2)

    def test_parse_range_with_sheet(self):
        rect = parse_ref("'Ma Feuille'!A1:C10")
        assert rect is not None
        assert rect.sheet == "Ma Feuille"
        assert (rect.r1, rect.c1, rect.r2, rect.c2) == (1, 1, 10, 3)

    def test_parse_whole_column(self):
        rect = parse_ref("A:B")
        assert rect is not None
        assert rect.c1 == 1 and rect.c2 == 2
        assert rect.r1 == 1 and rect.r2 == 1_048_576

    def test_structured_ref_rejected(self):
        assert parse_ref("Table1[Col]") is None
        assert parse_ref("MyName") is None

    def test_r1c1_relative(self):
        assert ref_to_r1c1("A2", 2, 2) == "RC[-1]"
        assert ref_to_r1c1("$A$1:A10", 5, 3) == "R1C1:R[5]C[-2]"

    def test_stretch_relative_follows_group(self):
        detail = parse_ref_detailed("A2", "S")
        assert detail is not None
        rect = stretch_ref(detail, 2, 4, (2, 101), (4, 4))
        assert rect is not None
        assert (rect.r1, rect.r2) == (2, 101)
        assert rect.c1 == rect.c2 == 1

    def test_stretch_anchored_stays_fixed(self):
        detail = parse_ref_detailed("$A$2", "S")
        assert detail is not None
        rect = stretch_ref(detail, 2, 4, (2, 101), (4, 4))
        assert rect is not None
        assert (rect.r1, rect.r2) == (2, 2)


class TestRewrite:
    def test_stretched_formulas_share_canonical_form(self):
        assert canonical_r1c1("A2*2+1", 2, 2) == canonical_r1c1("A9*2+1", 9, 2)

    def test_different_logic_differs(self):
        assert canonical_r1c1("A1", 1, 2) != canonical_r1c1("A1", 2, 2)

    def test_qualify_leaves_names_and_strings(self):
        out = qualify_sheet('SUM(A1:A3)+MonNom&"B2"', "Feuil 1")
        assert "'Feuil 1'!A1:A3" in out
        assert "MonNom" in out and "'Feuil 1'!MonNom" not in out
        assert '"B2"' in out

    def test_qualify_keeps_existing_sheet(self):
        out = qualify_sheet("Data!B2*C3", "S1")
        assert "Data!B2" in out and "S1!C3" in out


class TestAnalyze:
    def test_graph_structure(self, lineage_excel):
        result = analyze_workbook(lineage_excel, "test.xlsx")
        graph = result["graph"]
        stats = graph["meta"]["stats"]
        assert stats["totalFormulas"] == 103
        kinds = {n["kind"] for n in graph["nodes"]}
        assert {"group", "cell", "input", "name"} <= kinds

    def test_stretched_column_becomes_one_group(self, lineage_excel):
        graph = analyze_workbook(lineage_excel, "test.xlsx")["graph"]
        groups = [n for n in graph["nodes"] if n["kind"] == "group"]
        assert len(groups) == 1
        assert groups[0]["count"] == 100
        assert groups[0]["bbox"] == "D2:D101"

    def test_group_inputs_are_aggregated(self, lineage_excel):
        graph = analyze_workbook(lineage_excel, "test.xlsx")["graph"]
        input_labels = {n["label"] for n in graph["nodes"] if n["kind"] == "input"}
        assert "Ventes!B2:B101" in input_labels
        assert "Ventes!C2:C101" in input_labels

    def test_defined_name_resolved(self, lineage_excel):
        graph = analyze_workbook(lineage_excel, "test.xlsx")["graph"]
        names = [n for n in graph["nodes"] if n["kind"] == "name"]
        assert names and names[0]["label"] == "TauxCible"
        # name is fed by Params!A1 and feeds Synthese!B3
        edges = graph["edges"]
        assert any(
            e["target"] == names[0]["id"] and "Params" in e["source"] for e in edges
        )
        assert any(
            e["source"] == names[0]["id"] and e["target"].endswith("Synthese!B3")
            for e in edges
        )

    def test_composed_formula_steps_evaluated(self, lineage_excel):
        graph = analyze_workbook(lineage_excel, "test.xlsx")["graph"]
        node = next(n for n in graph["nodes"] if n["id"].endswith("Synthese!B3"))
        steps = node["steps"]
        assert steps["label"] == "IF"
        assert steps["evaluated"] and steps["value"] == node["value"]
        # the comparison and inner SUM are evaluated individually
        flat = _flatten(steps)
        by_label = {s["label"]: s for s in flat}
        assert by_label[">"]["value"] is True
        assert isinstance(by_label["SUM"]["value"], float)
        assert by_label["ROUND"]["evaluated"]

    def test_defined_name_on_an_apostrophe_sheet_resolves(self):
        """openpyxl hands back the *escaped* sheet name of a defined name.

        `destinations` strips the surrounding quotes but leaves the doubled
        apostrophes, so a sheet called ``O'Brien`` arrives as ``O''Brien``.
        Storing that verbatim made ``to_a1()`` quote it a second time, and the
        name resolved to a sheet nobody has: it became an opaque node instead of
        an edge to the real cell.
        """
        from openpyxl import Workbook
        from openpyxl.workbook.defined_name import DefinedName

        wb = Workbook()
        ws = wb.active
        ws.title = "O'Brien's Café"
        ws["B4"] = 7
        other = wb.create_sheet("Report")
        other["A1"] = "=Threshold*2"
        wb.defined_names.add(
            DefinedName("Threshold", attr_text="'O''Brien''s Café'!$B$4")
        )
        buf = io.BytesIO()
        wb.save(buf)

        graph = analyze_workbook(buf.getvalue(), "quoted.xlsx")["graph"]
        name = next(n for n in graph["nodes"] if n["kind"] == "name")
        assert name["targets"] == ["'O''Brien''s Café'!B4"]
        assert name["sheet"] == "O'Brien's Café"
        assert not [n for n in graph["nodes"] if n["kind"] == "opaque"]

    def test_sheet_scoped_defined_names_are_collected(self):
        """A name declared on one sheet is invisible to `owb.defined_names`.

        Per-sheet names are ordinary in real files — a `Total` or `Limit` local
        to a tab — and skipping them turned every formula using one into an
        unresolved reference.
        """
        from openpyxl import Workbook
        from openpyxl.workbook.defined_name import DefinedName

        wb = Workbook()
        config = wb.active
        config.title = "Config"
        config["B3"] = 250_000
        report = wb.create_sheet("Report")
        report["A1"] = "=LocalLimit*2"
        report.defined_names.add(DefinedName("LocalLimit", attr_text="Config!$B$3"))
        buf = io.BytesIO()
        wb.save(buf)

        graph = analyze_workbook(buf.getvalue(), "scoped.xlsx")["graph"]
        names = {n["label"]: n for n in graph["nodes"] if n["kind"] == "name"}
        assert "LocalLimit" in names
        assert names["LocalLimit"]["targets"] == ["Config!B3"]
        assert not [n for n in graph["nodes"] if n["kind"] == "opaque"]

    def test_workbook_scope_wins_over_sheet_scope(self):
        """A shadowed name must not replace the one most formulas mean."""
        from openpyxl import Workbook
        from openpyxl.workbook.defined_name import DefinedName

        wb = Workbook()
        data = wb.active
        data.title = "Data"
        data["A1"] = 1
        data["A2"] = 2
        other = wb.create_sheet("Other")
        other["A1"] = "=Limit"
        wb.defined_names.add(DefinedName("Limit", attr_text="Data!$A$1"))
        data.defined_names.add(DefinedName("Limit", attr_text="Data!$A$2"))
        buf = io.BytesIO()
        wb.save(buf)

        graph = analyze_workbook(buf.getvalue(), "shadow.xlsx")["graph"]
        name = next(n for n in graph["nodes"] if n["kind"] == "name")
        assert name["targets"] == ["Data!A1"]

    def test_let_bindings_are_not_graph_nodes(self):
        """`LET` names parse as references but point at no cell.

        They used to become one 'external reference' node per intermediate the
        modeller happened to name.
        """
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        for row in range(1, 6):
            ws.cell(row=row, column=1, value=row)
        ws["C1"] = "=LET(total, SUM(A1:A5), n, COUNT(A1:A5), total/n)"
        buf = io.BytesIO()
        wb.save(buf)

        graph = analyze_workbook(buf.getvalue(), "let.xlsx")["graph"]
        labels = {n["label"] for n in graph["nodes"]}
        assert "total" not in labels and "n" not in labels
        assert not [n for n in graph["nodes"] if n["kind"] == "opaque"]
        # The real range it reads is still an edge, so the cell keeps its lineage.
        assert any("A1:A5" in label for label in labels)

    def test_a_let_binding_does_not_hide_a_real_reference(self):
        """Only bare identifiers are bindings; `A1` inside LET stays a reference."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        ws["A1"] = 10
        ws["B1"] = 3
        ws["C1"] = "=LET(scale, B1, A1*scale)"
        buf = io.BytesIO()
        wb.save(buf)

        graph = analyze_workbook(buf.getvalue(), "let2.xlsx")["graph"]
        sources = {e["source"] for e in graph["edges"] if e["target"].endswith("S!C1")}
        assert any(src.endswith("S!A1") for src in sources)
        assert any(src.endswith("S!B1") for src in sources)

    def test_values_computed_by_engine(self, lineage_excel):
        graph = analyze_workbook(lineage_excel, "test.xlsx")["graph"]
        b1 = next(n for n in graph["nodes"] if n["id"].endswith("Synthese!B1"))
        assert isinstance(b1["value"], float) and b1["value"] > 0


def node_value(graph: dict, suffix: str):
    """Computed value of the formula node whose id ends with ``suffix``."""
    return next(n for n in graph["nodes"] if n["id"].endswith(suffix))["value"]


def _flatten(step):
    out = [step]
    for child in step.get("children", []):
        out.extend(_flatten(child))
    return out


class TestStretchRobustness:
    """A copied run must collapse into one group wherever it sits on the sheet.

    Each case builds a minimal workbook and asserts on the *number of group
    nodes*: a run that fails to group does not disappear, it explodes into one
    ``cell`` node per member, which is the symptom these tests guard against.
    """

    @staticmethod
    def _graph(cells: dict[str, Any], sheet: str = "S") -> dict:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = sheet
        for addr, value in cells.items():
            ws[addr] = value
        buf = io.BytesIO()
        wb.save(buf)
        return analyze_workbook(buf.getvalue(), "stretch.xlsx")["graph"]

    @staticmethod
    def _only_group(graph: dict) -> dict:
        groups = [n for n in graph["nodes"] if n["kind"] == "group"]
        assert len(groups) == 1, f"expected one group, got {graph['nodes']}"
        assert not [n for n in graph["nodes"] if n["kind"] == "cell"]
        return groups[0]

    def test_a_run_starting_at_b2_is_one_group(self):
        """Header row above, first column empty: the run still groups."""
        graph = self._graph(
            {
                "A1": "Qté",
                **{f"A{r}": r for r in range(2, 7)},
                **{f"B{r}": f"=A{r}*2" for r in range(2, 7)},
            }
        )
        group = self._only_group(graph)
        assert group["id"] == "g:S!B2#5"
        assert group["count"] == 5
        assert group["bbox"] == "B2:B6"

    def test_a_run_starting_at_a5_is_one_group(self):
        """The run starts below the used range's first rows."""
        graph = self._graph(
            {
                **{f"B{r}": r for r in range(5, 10)},
                **{f"A{r}": f"=B{r}*2" for r in range(5, 10)},
            }
        )
        group = self._only_group(graph)
        assert group["id"] == "g:S!A5#5"
        assert group["bbox"] == "A5:A9"

    def test_a_run_with_absolute_refs_is_one_group(self):
        """$A$1 is identical in every member, so the key must stay identical."""
        graph = self._graph(
            {
                "A1": 3,
                **{f"B{r}": r for r in range(2, 7)},
                **{f"C{r}": f"=B{r}*$A$1" for r in range(2, 7)},
            }
        )
        assert self._only_group(graph)["bbox"] == "C2:C6"

    def test_a_horizontal_run_is_one_group(self):
        """A run copied across columns, not down rows."""
        graph = self._graph(
            {
                **{f"{c}2": 1 for c in "CDEFG"},
                **{f"{c}3": f"={c}2*2" for c in "CDEFG"},
            }
        )
        group = self._only_group(graph)
        assert group["id"] == "g:S!C3#5"
        assert group["bbox"] == "C3:G3"

    def test_a_horizontal_run_survives_a_let_parameter(self):
        """Regression: the tokenizer reports the LET variable ``x`` as a Range
        operand. Converted to R1C1 it became a column offset that moved with
        the host cell (C[21], C[20], ...), splitting the run into five nodes."""
        graph = self._graph(
            {
                **{f"{c}2": 1 for c in "CDEFG"},
                **{f"{c}3": f"=LET(x,{c}2,x*2)" for c in "CDEFG"},
            }
        )
        assert self._only_group(graph)["bbox"] == "C3:G3"

    def test_a_short_defined_name_is_not_read_as_a_column(self):
        """Same defect, seen from the canonical form: a defined name of three
        letters or fewer looks like a column to :func:`ref_to_r1c1`."""
        keys = [canonical_r1c1(f"TVA*{c}2", 3, 3 + i) for i, c in enumerate("CDE")]
        assert len(set(keys)) == 1
        assert "TVA" in keys[0]

    def test_the_r1c1_conversion_still_applies_to_real_refs(self):
        """The guard must not disarm the happy path it protects."""
        assert canonical_r1c1("A2*2", 2, 4) == "RC[-3]*2"
        assert canonical_r1c1("SUM(A:A)", 2, 4) == "SUM(C[-3]:C[-3])"
        assert canonical_r1c1("SUM(S1:S3!A2)", 2, 4) == "SUM('S1:S3'!RC[-3])"

    def test_an_untokenizable_formula_groups_on_its_structure(self):
        """Fallback contract: no formula found in a real workbook makes the
        tokenizer throw (only malformed input does, which Excel would not
        store), so this exercises :func:`canonical_r1c1` directly rather than
        through a workbook. Two members of one run must share a key."""
        assert canonical_r1c1("SUM(B2:C2", 2, 4) == canonical_r1c1("SUM(B3:C3", 3, 4)

    def test_the_fallback_still_separates_different_logic(self):
        assert canonical_r1c1("SUM(B2:C2", 2, 4) != canonical_r1c1(
            "AVERAGE(B2:C2", 2, 4
        )

    def test_the_fallback_leaves_function_names_alone(self):
        """The ref regex must not bite into LOG10 or _xlfn.FOO2."""
        key = canonical_r1c1('IF(LOG10(A2)>0,"x",Sheet1!$B$4', 2, 4)
        assert key == 'if(log10(@)>0,"x",@'


class TestPackageApi:
    """The tool must be usable as a library, without FastAPI or AI."""

    def test_analyze_from_bytes(self, lineage_excel):
        result = analyze(lineage_excel, filename="demo.xlsx")
        assert isinstance(result, LineageResult)
        assert result.stats["totalFormulas"] == 103
        assert "Ventes" in result.sheets

    def test_analyze_from_path(self, tmp_path, lineage_excel):
        path = tmp_path / "workbook.xlsx"
        path.write_bytes(lineage_excel)
        result = analyze(path)
        assert result.stats["totalFormulas"] == 103

    def test_analyze_from_filelike(self, lineage_excel):
        result = analyze(io.BytesIO(lineage_excel), filename="stream.xlsx")
        assert result.stats["totalFormulas"] == 103

    def test_navigation_helpers(self, lineage_excel):
        result = analyze(lineage_excel)
        b3 = result.find("Synthese!B3")
        assert b3 and b3[0]["id"].endswith("Synthese!B3")
        node_id = b3[0]["id"]
        prec_labels = {n["label"] for n in result.precedents(node_id)}
        assert "TauxCible" in prec_labels
        node = result.node(node_id)
        assert node and node["formula"].startswith("=IF")

    def test_to_json_roundtrip(self, lineage_excel):
        import json

        result = analyze(lineage_excel)
        data = json.loads(result.to_json())
        assert data["meta"]["stats"]["totalFormulas"] == 103

    def test_workbook_context_preserves_preview_and_comments(self, lineage_excel):
        result = analyze(lineage_excel, filename="context.xlsx")
        context = result.workbook_context
        ventes = next(sheet for sheet in context["sheets"] if sheet["name"] == "Ventes")
        assert ventes["preview"][0]["values"][:4] == [
            "Produit",
            "Qté",
            "Prix",
            "CA",
        ]
        assert ventes["comments"] == [
            {
                "cell": "A1",
                "author": "Data team",
                "text": "Exported product category",
            }
        ]
        assert ventes["freeze_panes"] == "A2"
        assert ventes["hidden_columns"] == ["C"]
        assert "F1:G1" in ventes["merged_ranges"]

    def test_screenshots_report_a_missing_renderer(
        self, lineage_excel, monkeypatch, tmp_path
    ):
        from linexcel import WorkbookRenderError

        monkeypatch.setattr("linexcel.insights.shutil.which", lambda _: None)
        monkeypatch.setattr("linexcel.insights._launcher_install_paths", lambda: ())
        monkeypatch.setattr("linexcel.insights._pdftoppm_install_paths", lambda: ())
        with pytest.raises(WorkbookRenderError, match="LibreOffice, pdftoppm"):
            analyze(lineage_excel).save_screenshots(tmp_path)

    def test_renderer_is_found_in_a_standard_install_outside_path(
        self, monkeypatch, tmp_path
    ):
        """Windows and macOS installers do not extend PATH."""
        from linexcel.insights import find_libreoffice, find_pdftoppm

        installed_office = tmp_path / "soffice.com"
        installed_office.write_text("", encoding="utf-8")
        installed_converter = tmp_path / "pdftoppm.exe"
        installed_converter.write_text("", encoding="utf-8")

        monkeypatch.setattr("linexcel.insights.shutil.which", lambda _: None)
        monkeypatch.setattr(
            "linexcel.insights._launcher_install_paths",
            lambda: (tmp_path / "absent.com", installed_office),
        )
        monkeypatch.setattr(
            "linexcel.insights._pdftoppm_install_paths",
            lambda: (installed_converter,),
        )
        assert find_libreoffice() == str(installed_office)
        assert find_pdftoppm() == str(installed_converter)

    def test_screenshots_run_headless_conversion_pipeline(
        self, lineage_excel, monkeypatch, tmp_path
    ):
        calls = []

        def fake_run(command, **_kwargs):
            calls.append(command)
            if command[0] == "libreoffice":
                pdf_dir = Path(command[command.index("--outdir") + 1])
                (pdf_dir / "workbook.pdf").write_bytes(b"%PDF-1.4")
            else:
                Path(f"{command[-1]}-1.png").write_bytes(b"png")

        commands = {"libreoffice": "libreoffice", "pdftoppm": "pdftoppm"}
        monkeypatch.setattr("linexcel.insights.shutil.which", commands.get)
        monkeypatch.setattr("linexcel.insights.subprocess.run", fake_run)
        screenshots = analyze(lineage_excel).save_screenshots(tmp_path)
        assert isinstance(screenshots, list)
        assert [path.name for path in screenshots] == ["workbook-1.png"]
        assert calls[0][2:5] == ["--headless", "--norestore", "--convert-to"]
        assert calls[1][1:3] == ["-png", "-r"]

    def test_screenshots_use_a_throwaway_libreoffice_profile(
        self, lineage_excel, monkeypatch, tmp_path
    ):
        """A desktop LibreOffice owns the default profile; sharing it makes the
        headless run exit 0 having converted nothing."""
        calls = []

        def fake_run(command, **_kwargs):
            calls.append(command)
            if command[0] == "libreoffice":
                pdf_dir = Path(command[command.index("--outdir") + 1])
                (pdf_dir / "workbook.pdf").write_bytes(b"%PDF-1.4")
            else:
                Path(f"{command[-1]}-1.png").write_bytes(b"png")

        commands = {"libreoffice": "libreoffice", "pdftoppm": "pdftoppm"}
        monkeypatch.setattr("linexcel.insights.shutil.which", commands.get)
        monkeypatch.setattr("linexcel.insights.subprocess.run", fake_run)
        analyze(lineage_excel).save_screenshots(tmp_path)
        profile = calls[0][1]
        assert profile.startswith("-env:UserInstallation=file://")
        assert "linexcel-render-" in profile

    def test_to_html_is_offline_and_self_contained(self, lineage_excel):
        result = analyze(lineage_excel)
        html = result.to_html()
        assert html.startswith("<!doctype html>")
        # Cytoscape embedded, no network dependency
        assert "cdn.jsdelivr" not in html
        assert "cytoscape" in html
        # the composite formula and its decomposition are in the injected data
        assert "Synthese!B3" in html

    def test_workbook_doc_has_a_separate_html_tab(self, lineage_excel):
        result = analyze(lineage_excel)
        html = result.to_html(workbook_doc="# Workbook role\n\nA test overview.")
        assert "Workbook overview" in html
        assert "workbookDoc" in html
        assert "A test overview." in html

    def test_build_workbook_dossier(self, lineage_excel):
        from linexcel.aidoc import build_workbook_dossier

        dossier = build_workbook_dossier(analyze(lineage_excel).graph)
        sheets = {sheet["name"]: sheet for sheet in dossier["sheets"]}
        assert sheets["Ventes"]["formula_cells"] == 100
        assert sheets["Ventes"]["dimensions"]["columns"] == 7
        assert dossier["defined_names"] == [
            {"name": "TauxCible", "targets": ["Params!A1"]}
        ]
        assert dossier["formula_patterns"][0]["cells"] == 100

    def test_screenshot_tab_is_translated_like_the_rest_of_the_viewer(
        self, lineage_excel
    ):
        """The screenshot pane used to hard-code its French heading and the
        'Page N' labels, so an English report was partly French."""
        png = "data:image/png;base64,aGVsbG8="
        html = analyze(lineage_excel).to_html(language="en", screenshots=[png])
        assert "Aperçu" not in html
        assert '"page": "Page {n}"' in html
        assert "_t('page'" in html

    def test_every_language_defines_the_screenshot_page_label(self):
        from linexcel.i18n import LANGUAGES, UI_STRINGS

        for language in LANGUAGES:
            assert "{n}" in UI_STRINGS[language]["page"]

    def test_repr_html_wraps_in_data_iframe(self, lineage_excel):
        result = analyze(lineage_excel)
        frame = result._repr_html_()
        assert frame.startswith('<iframe src="data:text/html;base64,')

    def test_save_html(self, tmp_path, lineage_excel):
        result = analyze(lineage_excel)
        out = result.save_html(tmp_path / "graph.html")
        assert out.exists() and out.stat().st_size > 100_000

    def test_version_is_exposed(self):
        import linexcel

        assert isinstance(linexcel.__version__, str) and linexcel.__version__

    def test_unsupported_language_is_rejected(self, lineage_excel):
        result = analyze(lineage_excel)
        with pytest.raises(ValueError, match="Unsupported language"):
            result.to_html(language='"; alert(1); var x="')

    def test_to_html_without_source_bytes_drops_the_sheet_tab(self, lineage_excel):
        """A result built without the workbook bytes still renders."""
        payload = analyze_workbook(lineage_excel, "test.xlsx")
        result = LineageResult(graph=payload["graph"], engine=payload["engine"])
        html = result.to_html()
        assert html.startswith("<!doctype html>")
        assert '"workbookContext": null' in html or '"workbookContext":null' in html

    def test_title_does_not_overwrite_workbook_data(self, lineage_excel):
        """Placeholders must not be substituted inside the injected graph."""
        from linexcel.viewer import render_html

        graph = {
            "nodes": [{"id": "n1", "kind": "cell", "label": "__TITLE__"}],
            "edges": [],
            "sheets": [],
            "meta": {"stats": {}, "warnings": []},
        }
        html = render_html(graph, title="Report", language="fr")
        assert '"label": "__TITLE__"' in html
        # Matched loosely: this test guards placeholder substitution, not the
        # heading's markup, and pinning the exact tag makes any top-bar change
        # look like a data-integrity failure.
        assert ">Report</h1>" in html

    def test_title_cannot_break_the_embedded_json(self, lineage_excel):
        """A backslash-terminated title used to truncate the GRAPH literal."""
        import json
        import re

        result = analyze(lineage_excel)
        html = result.to_html(title="report\\")
        embedded = re.search(r"var GRAPH = (.*?);\n", html, re.S)
        assert embedded
        assert json.loads(embedded.group(1))["meta"]["stats"]["totalFormulas"] == 103

    def test_document_without_provider_raises_aidocerror(
        self, lineage_excel, monkeypatch
    ):
        """Bare calls no longer fall back to any vendor: no implicit default."""
        from linexcel.aidoc import AiDocError

        _clear_ai_env(monkeypatch)
        result = analyze(lineage_excel)
        try:
            result.document()
        except AiDocError as exc:
            assert "No AI provider selected" in str(exc)
        else:  # pragma: no cover
            raise AssertionError("AiDocError expected when no provider is configured")


class TestLanguages:
    """The language set is a closed allowlist, and every registry must cover it.

    A language present in one registry but not another would surface as raw
    interface keys in the report or as a KeyError when prompting.
    """

    def test_registries_cover_every_language(self):
        from linexcel.aidoc import _SYSTEM, _VISION_SYSTEM, _WORKBOOK_SYSTEM
        from linexcel.i18n import LANGUAGES, UI_STRINGS

        assert set(UI_STRINGS) == set(LANGUAGES)
        assert set(_SYSTEM) == set(LANGUAGES)
        assert set(_WORKBOOK_SYSTEM) == set(LANGUAGES)
        assert set(_VISION_SYSTEM) == set(LANGUAGES)

    def test_every_language_defines_every_ui_key(self):
        from linexcel.i18n import DEFAULT_LANGUAGE, UI_STRINGS

        expected = set(UI_STRINGS[DEFAULT_LANGUAGE])
        for language, strings in UI_STRINGS.items():
            assert set(strings) == expected, f"{language} key set differs"

    def test_ui_keys_match_what_the_viewer_asks_for(self):
        """Guards against a key used by the template but defined nowhere."""
        import re as _re

        from linexcel.i18n import DEFAULT_LANGUAGE, UI_STRINGS

        # The template itself, not the module that loads it: it moved out to
        # assets/viewer.html once, and reading it through _TEMPLATE means the
        # check follows it wherever it goes next.
        source = viewer_module._TEMPLATE
        used = set(_re.findall(r"_t\('([a-z_]+)'", source))
        used |= set(_re.findall(r"labelKey: '([a-z_]+)'", source))
        assert used, "no i18n key found in the viewer template"
        assert used <= set(UI_STRINGS[DEFAULT_LANGUAGE]), (
            f"keys used but not defined: {used - set(UI_STRINGS[DEFAULT_LANGUAGE])}"
        )

    def test_placeholders_are_preserved_in_every_translation(self):
        import re as _re

        from linexcel.i18n import DEFAULT_LANGUAGE, UI_STRINGS

        reference = UI_STRINGS[DEFAULT_LANGUAGE]
        for language, strings in UI_STRINGS.items():
            for key, text in strings.items():
                assert set(_re.findall(r"\{(\w+)\}", text)) == set(
                    _re.findall(r"\{(\w+)\}", reference[key])
                ), f"{language}/{key} lost or invented a placeholder"

    @pytest.mark.parametrize("language", ["es", "de", "it", "pt", "nl", "ja", "zh"])
    def test_each_language_renders(self, lineage_excel, language):
        html = analyze(lineage_excel).to_html(language=language)
        assert f"<html lang='{language}'>" in html
        assert f'"{language}"' in html

    def test_only_the_requested_locale_and_english_are_embedded(self, lineage_excel):
        """Shipping all nine locales in every report would be dead weight."""
        from linexcel.i18n import UI_STRINGS

        html = analyze(lineage_excel).to_html(language="ja")
        assert UI_STRINGS["ja"]["sheets_tab"] in html
        assert UI_STRINGS["en"]["sheets_tab"] in html
        assert UI_STRINGS["zh"]["placeholder_title"] not in html

    def test_unsupported_language_is_rejected_by_both_entry_points(self, lineage_excel):
        result = analyze(lineage_excel)
        with pytest.raises(ValueError, match="Unsupported language"):
            result.to_html(language="klingon")
        with pytest.raises(ValueError, match="Unsupported language"):
            result.document_workbook(
                provider=lambda s, u, *, temperature=0.2: "x", language="klingon"
            )


class TestAiProviders:
    """Provider resolution and failure handling — no network call involved."""

    @staticmethod
    def _calc_ids(result: LineageResult) -> list[str]:
        return [n["id"] for n in result.nodes if n["kind"] in ("cell", "group")]

    def test_no_provider_selected_raises_with_guidance(
        self, lineage_excel, monkeypatch
    ):
        """A bare call is an explicit error, never a vendor picked for you."""
        from linexcel.aidoc import AiDocError

        _clear_ai_env(monkeypatch)
        result = analyze(lineage_excel)
        with pytest.raises(AiDocError, match="No AI provider selected"):
            result.document_workbook()

    def test_model_alone_selects_nothing(self, lineage_excel, monkeypatch):
        """A model name identifies no endpoint: there is no vendor to infer it."""
        from linexcel.aidoc import AiDocError

        _clear_ai_env(monkeypatch)
        result = analyze(lineage_excel)
        with pytest.raises(AiDocError, match="No AI provider selected"):
            result.document(model="some-model-name")

    def test_api_key_alone_selects_nothing(self, lineage_excel, monkeypatch):
        """api_key= names no endpoint either, so it cannot select a provider."""
        from linexcel.aidoc import AiDocError

        _clear_ai_env(monkeypatch)
        result = analyze(lineage_excel)
        with pytest.raises(AiDocError, match="No AI provider selected"):
            result.document_workbook(api_key="some-key")

    def test_base_url_without_model_says_so(self, lineage_excel, monkeypatch):
        """Endpoints disagree on a default model, so one must be named."""
        from linexcel.aidoc import AiDocError

        _clear_ai_env(monkeypatch)
        result = analyze(lineage_excel)
        with pytest.raises(AiDocError, match="No model named for the endpoint"):
            result.document_workbook(base_url="http://localhost:11434/v1")

    def test_base_url_and_model_reach_the_openai_compatible_client(
        self, lineage_excel, monkeypatch
    ):
        """The only built-in client: whatever answers at base_url."""
        from linexcel import aidoc

        captured = {}

        class StubClient:
            def __init__(self, *, base_url, api_key, model):
                captured.update(base_url=base_url, api_key=api_key, model=model)

            def generate(
                self, system_prompt, user_prompt, *, temperature=0.2, max_tokens=None
            ):
                return "# overview"

        _clear_ai_env(monkeypatch)
        monkeypatch.setattr(aidoc, "_OpenAICompatProvider", StubClient)
        result = analyze(lineage_excel)
        assert (
            result.document_workbook(
                base_url="http://localhost:11434/v1", model="qwen3.8"
            )
            == "# overview"
        )
        assert captured["base_url"] == "http://localhost:11434/v1"
        assert captured["model"] == "qwen3.8"
        assert captured["api_key"] is None  # left to the client to resolve

    def test_env_vars_are_the_equivalent_of_base_url_and_model(
        self, lineage_excel, monkeypatch
    ):
        """LINEXCEL_AI_BASE_URL + LINEXCEL_AI_MODEL configure the same client."""
        from linexcel import aidoc

        captured = {}

        class StubClient:
            def __init__(self, *, base_url, api_key, model):
                captured.update(base_url=base_url, model=model)

            def generate(
                self, system_prompt, user_prompt, *, temperature=0.2, max_tokens=None
            ):
                return "# card"

        _clear_ai_env(monkeypatch)
        monkeypatch.setattr(aidoc, "_OpenAICompatProvider", StubClient)
        monkeypatch.setenv("LINEXCEL_AI_BASE_URL", "https://openrouter.ai/api/v1")
        monkeypatch.setenv("LINEXCEL_AI_MODEL", "some-org/some-model")
        result = analyze(lineage_excel)
        node_id = self._calc_ids(result)[0]
        assert result.document([node_id]) == {node_id: "# card"}
        assert captured["base_url"] == "https://openrouter.ai/api/v1"
        assert captured["model"] == "some-org/some-model"

    def test_no_vendor_is_named_in_the_resolution_path(self):
        """Regression guard: providers are configuration, never source code.

        A vendor name reintroduced as a constant, an env var or a default model
        would quietly make that vendor the privileged one again. Prose may name
        endpoints as examples, so only code lines are scanned.
        """
        from pathlib import Path

        from linexcel import aidoc

        source = Path(aidoc.__file__).read_text(encoding="utf-8")
        code = _strip_comments_and_docstrings(source)
        for vendor in ("gemini", "google", "genai", "anthropic", "claude", "gpt-"):
            assert vendor not in code.lower(), (
                f"{vendor!r} is hard-coded in aidoc.py; providers are chosen by "
                "the caller, not named in the source"
            )

    def test_plain_callable_is_accepted(self, lineage_excel):
        seen = []

        def my_llm(system_prompt, user_prompt, *, temperature=0.2):
            seen.append((system_prompt, user_prompt, temperature))
            return "# card"

        result = analyze(lineage_excel)
        node_id = self._calc_ids(result)[0]
        assert result.document([node_id], provider=my_llm) == {node_id: "# card"}
        assert seen[0][2] == 0.2
        assert node_id in seen[0][1]

    def test_object_exposing_generate_is_accepted(self, lineage_excel):
        class Provider:
            def generate(self, system_prompt, user_prompt, *, temperature=0.2, **kw):
                return "# from object"

        result = analyze(lineage_excel)
        node_id = self._calc_ids(result)[0]
        assert result.document([node_id], provider=Provider()) == {
            node_id: "# from object"
        }

    def test_workbook_overview_accepts_a_callable(self, lineage_excel):
        result = analyze(lineage_excel)
        assert (
            result.document_workbook(
                provider=lambda system, user, *, temperature=0.2, **kw: "# overview"
            )
            == "# overview"
        )

    def test_provider_without_generate_is_rejected(self, lineage_excel):
        from linexcel.aidoc import AiDocError

        result = analyze(lineage_excel)
        with pytest.raises(AiDocError, match="must expose a generate"):
            # cast: the point of the test is the runtime rejection
            result.document_workbook(provider=cast(Any, object()))

    def test_partial_failure_keeps_the_successful_cards(self, lineage_excel):
        result = analyze(lineage_excel)
        ids = self._calc_ids(result)
        failing = ids[0]

        def flaky(system_prompt, user_prompt, *, temperature=0.2):
            if f'"node_id": "{failing}"' in user_prompt:
                raise RuntimeError("rate limited")
            return "# card"

        with pytest.warns(UserWarning, match=f"failed for 1 of {len(ids)} nodes"):
            docs = result.document(ids, provider=flaky, max_workers=1)
        assert set(docs) == set(ids) - {failing}

    def test_total_failure_raises(self, lineage_excel):
        from linexcel.aidoc import AiDocError

        result = analyze(lineage_excel)
        ids = self._calc_ids(result)[:2]

        def broken(system_prompt, user_prompt, *, temperature=0.2):
            raise RuntimeError("no backend")

        with pytest.raises(AiDocError, match="failed for all 2 nodes"):
            result.document(ids, provider=broken, max_workers=1)


class TestTokenUsage:
    """Token accounting: real counts when the provider reports them, else
    an estimate that is flagged as such."""

    @staticmethod
    def _calc_ids(result: LineageResult) -> list[str]:
        return [n["id"] for n in result.nodes if n["kind"] in ("cell", "group")]

    def test_estimate_counts_latin_words(self):
        from linexcel.aidoc import estimate_tokens

        assert estimate_tokens("") == 0
        assert estimate_tokens("the quick brown fox jumps") == 6  # 5 words * 4/3

    def test_estimate_counts_cjk_per_character(self):
        """A spaceless Japanese sentence is one `\\w+` match, not one token."""
        import re

        from linexcel.aidoc import estimate_tokens

        text = "日本語のテキストです"
        assert len(re.findall(r"\w+", text)) == 1  # what a naive counter sees
        assert estimate_tokens(text) == len(text) == 10

    def test_mixed_script_counts_both_parts(self):
        from linexcel.aidoc import estimate_tokens

        assert estimate_tokens("SUM 合計 total") == 2 + (2 * 4 // 3)

    def test_callable_provider_usage_is_estimated_and_flagged(self, lineage_excel):
        result = analyze(lineage_excel)
        node_id = self._calc_ids(result)[0]
        result.document([node_id], provider=lambda s, u, *, temperature=0.2: "# card")

        usage = result.token_usage
        assert usage.requests == 1
        assert usage.estimated is True
        assert usage.input_tokens > 0 and usage.output_tokens > 0
        assert usage.total == usage.input_tokens + usage.output_tokens

    def test_provider_reported_counts_are_used_verbatim(self, lineage_excel):
        from linexcel.aidoc import TokenUsage

        class Reporting:
            def generate(self, system_prompt, user_prompt, *, temperature=0.2, **kw):
                return "# card"

            def generate_with_usage(
                self, system_prompt, user_prompt, *, temperature=0.2, **kw
            ):
                return "# card", TokenUsage(
                    input_tokens=1234,
                    output_tokens=56,
                    requests=1,
                    model="m",
                    provider="p",
                )

        result = analyze(lineage_excel)
        result.document([self._calc_ids(result)[0]], provider=Reporting())

        usage = result.token_usage
        assert (usage.input_tokens, usage.output_tokens) == (1234, 56)
        assert usage.estimated is False
        assert usage.model == "m" and usage.provider == "p"

    def test_usage_accumulates_across_calls(self, lineage_excel):
        result = analyze(lineage_excel)
        ids = self._calc_ids(result)
        provider = lambda s, u, *, temperature=0.2: "# card"  # noqa: E731

        result.document(ids, provider=provider, max_workers=1)
        after_nodes = result.token_usage.requests
        result.document_workbook(provider=provider)

        assert after_nodes == len(ids)
        assert result.token_usage.requests == len(ids) + 1

    def test_tokens_spent_before_a_failure_are_still_counted(self, lineage_excel):
        """Tokens already spent are billed even if a later node fails."""
        result = analyze(lineage_excel)
        ids = self._calc_ids(result)
        failing = ids[0]

        def flaky(system_prompt, user_prompt, *, temperature=0.2):
            if f'"node_id": "{failing}"' in user_prompt:
                raise RuntimeError("rate limited")
            return "# card"

        with pytest.warns(UserWarning):
            result.document(ids, provider=flaky, max_workers=1)
        assert result.token_usage.requests == len(ids) - 1

    def test_usage_starts_empty_and_reads_cleanly(self, lineage_excel):
        usage = analyze(lineage_excel).token_usage
        assert (usage.total, usage.requests) == (0, 0)
        assert "0 tokens" in str(usage)

    def test_reported_usage_is_preferred_over_estimation(self):
        """``_usage_from`` reads whatever field names a provider uses."""
        from linexcel.aidoc import _usage_from

        class Meta:
            prompt_tokens = 900
            completion_tokens = 100

        usage = _usage_from(
            Meta(),
            ("prompt_tokens", "completion_tokens"),
            "prompt",
            "text",
            model="some-model",
            provider="openai-compatible",
        )
        assert (usage.input_tokens, usage.output_tokens, usage.total) == (
            900,
            100,
            1000,
        )
        assert usage.estimated is False

    def test_missing_usage_block_falls_back_to_estimation(self):
        """Local OpenAI-compatible runtimes do not always report usage."""
        from linexcel.aidoc import _usage_from

        usage = _usage_from(
            None,
            ("prompt_tokens", "completion_tokens"),
            "one two three four",
            "five six",
            model="local",
            provider="openai-compatible",
        )
        assert usage.estimated is True
        assert usage.input_tokens > 0 and usage.output_tokens > 0


class TestUnresolvableCellsAreIsolated:
    """One bad reference used to cost every other cell its computed value.

    `evaluate_all` is all-or-nothing and gives up on the *first* reference it
    cannot resolve, so a single formula pointing at another workbook dropped the
    whole file back to slow per-cell recovery. Those few cells are now set aside
    so the global pass completes.
    """

    @staticmethod
    def _workbook() -> bytes:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        for row in range(1, 11):
            ws.cell(row=row, column=1, value=row)
            ws.cell(row=row, column=2, value=f"=A{row}*3")
        ws["D1"] = "=SUM(B1:B10)"
        # The blocker: a link to a workbook that is not there.
        ws["D2"] = "='[Missing Book.xlsx]Sheet1'!$A$1"
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_every_other_cell_keeps_its_recomputed_value(self):
        graph = analyze_workbook(self._workbook(), "blocked.xlsx")["graph"]
        total = next(n for n in graph["nodes"] if n["id"].endswith("Data!D1"))
        assert total["value"] == 165  # 3 * (1+..+10)
        assert total["valueSource"] == "engine"

    def test_the_isolated_formula_stays_in_the_graph(self):
        """Blanking the cell in the engine must not delete it from the report."""
        graph = analyze_workbook(self._workbook(), "blocked.xlsx")["graph"]
        blocked = next(n for n in graph["nodes"] if n["id"].endswith("Data!D2"))
        assert blocked["formula"] == "='[Missing Book.xlsx]Sheet1'!$A$1"

    def test_the_warning_says_the_pass_completed(self):
        graph = analyze_workbook(self._workbook(), "blocked.xlsx")["graph"]
        joined = " ".join(graph["meta"]["warnings"])
        assert "Global evaluation completed after isolating 1 cell" in joined

    def test_a_healthy_workbook_is_untouched(self, lineage_excel):
        """The isolation path only runs once evaluation has already failed."""
        graph = analyze_workbook(lineage_excel, "test.xlsx")["graph"]
        assert not [w for w in graph["meta"]["warnings"] if "isolating" in w]

    def test_an_error_literal_is_not_mistaken_for_a_sheet(self):
        """`#REF!` looks like a sheet qualifier but guards evaluate around it."""
        from linexcel.analyzer import _is_unresolvable

        assert not _is_unresolvable('=IFERROR(#REF!, "handled")', {"Data"})
        assert not _is_unresolvable("=Data!A1", {"Data"})
        assert _is_unresolvable("=Gone!A1", {"Data"})
        assert _is_unresolvable("='[Other.xlsx]S'!A1", {"Data"})

    def test_a_guarded_formula_is_never_isolated(self):
        """Isolation must not cost a range one of its terms.

        `IFERROR(NOSHEET!A1, 456)` has a correct value despite an unresolvable
        reference. Blanking it does not merely lose that cell — every SUM
        spanning it quietly returns a smaller number.
        """
        from openpyxl import Workbook

        from linexcel.analyzer import _is_unresolvable

        assert not _is_unresolvable("=IFERROR(NOSHEET!A1, 456)", {"S"})
        assert not _is_unresolvable("=IFNA(Gone!A1, 0)", {"S"})

        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        ws["A1"] = 10
        ws["A2"] = "=IFERROR(NOSHEET!A1, 456)"
        ws["A3"] = "=SUM(A1:A2)"
        buf = io.BytesIO()
        wb.save(buf)

        graph = analyze_workbook(buf.getvalue(), "guarded.xlsx")["graph"]
        assert node_value(graph, "S!A2") == 456
        assert node_value(graph, "S!A3") == 466


class TestHostileCellText:
    """Workbook text is attacker-controlled: it must reach the page inert.

    Each payload below has broken some report generator — the first two close
    the script tag the graph is embedded in, the next two collide with the
    viewer's own template placeholders, `$&` is a regex replacement reference,
    and U+2028 terminates a JavaScript string literal though not a JSON one.
    """

    PAYLOADS = (
        "</script><script>alert(1)</script>",
        '"><img src=x onerror=alert(1)>',
        "__TITLE__",
        "__GRAPH_JSON__",
        "value is $& and $1",
        "line\u2028separator",
    )

    @staticmethod
    def _workbook(payloads) -> bytes:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Hostile"
        for row, payload in enumerate(payloads, start=1):
            ws.cell(row=row, column=1, value=payload)
            ws.cell(row=row, column=2, value=f'=A{row} & " (copied)"')
        ws["A1"].comment = Comment("</script><b>bold</b>", "</script>")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @pytest.fixture
    def hostile_html(self) -> str:
        return analyze(self._workbook(self.PAYLOADS), "hostile.xlsx").to_html()

    def test_no_payload_can_close_the_script_tag(self, hostile_html):
        """`<` is escaped everywhere, so no cell can end the graph's script."""
        assert "</script><script>alert" not in hostile_html
        assert "<img src=x onerror" not in hostile_html
        assert "\\u003c/script>" in hostile_html

    def test_the_embedded_graph_still_parses(self, hostile_html):
        embedded = re.search(r"var GRAPH = (.*?);\n", hostile_html, re.S)
        assert embedded
        graph = json.loads(embedded.group(1))
        assert graph["meta"]["stats"]["totalFormulas"] == len(self.PAYLOADS)

    def test_every_payload_survives_verbatim(self, hostile_html):
        """Escaping must be reversible: the reader has to see the real text."""
        embedded = re.search(r"var GRAPH = (.*?);\n", hostile_html, re.S)
        assert embedded
        graph = json.loads(embedded.group(1))
        sheet = graph["meta"]["workbookContext"]["sheets"][0]
        stored = {
            value
            for row in sheet["preview"]
            for value in row["values"]
            if isinstance(value, str)
        }
        for payload in self.PAYLOADS:
            assert payload in stored, payload

    def test_line_separators_cannot_break_the_literal(self, hostile_html):
        """U+2028 is a line terminator to JavaScript but legal inside JSON."""
        embedded = re.search(r"var GRAPH = (.*?);\n", hostile_html, re.S)
        assert embedded
        assert "\u2028" not in embedded.group(1)
        assert "\\u2028" in embedded.group(1)

    def test_a_payload_cannot_impersonate_a_template_placeholder(self):
        """`__TITLE__` in a cell must not be substituted with the real title."""
        html = analyze(self._workbook(["__TITLE__"]), "hostile.xlsx").to_html(
            title="Real Title"
        )
        embedded = re.search(r"var GRAPH = (.*?);\n", html, re.S)
        assert embedded
        assert "__TITLE__" in embedded.group(1)
        assert "Real Title" not in embedded.group(1)

    def test_a_hostile_comment_author_is_escaped(self, hostile_html):
        assert "<b>bold</b>" not in hostile_html


class TestTokenBudget:
    """A ceiling on total spend, so a workbook cannot quietly run up a bill."""

    @staticmethod
    def _provider(counter: list[int]):
        def provider(system_prompt, user_prompt, *, temperature=0.2):
            counter.append(1)
            return "card"

        return provider

    def test_budget_stops_the_run_and_reports_what_was_skipped(self, lineage_excel):
        calls: list[int] = []
        result = analyze(lineage_excel)
        ids = [n["id"] for n in result.nodes if n["kind"] in ("cell", "group")]
        assert len(ids) > 2, "fixture must offer several nodes to skip"

        with pytest.warns(UserWarning, match="never sent"):
            docs = result.document(
                ids, provider=self._provider(calls), max_workers=1, token_budget=1
            )

        # One request establishes the cost; the budget then stops the rest.
        assert len(calls) == 1
        assert len(docs) == 1 and len(docs) < len(ids)
        assert result.token_usage.requests == 1

    def test_a_budget_that_covers_the_run_changes_nothing(self, lineage_excel):
        calls: list[int] = []
        result = analyze(lineage_excel)
        ids = [n["id"] for n in result.nodes if n["kind"] in ("cell", "group")]

        docs = result.document(
            ids, provider=self._provider(calls), token_budget=10_000_000
        )

        assert len(calls) == len(ids) and len(docs) == len(ids)

    def test_budget_is_cumulative_over_the_result(self, lineage_excel):
        """One ceiling covers every call: the user pays per workbook, not per call."""
        from linexcel.aidoc import AiDocError

        result = analyze(lineage_excel)
        ids = [n["id"] for n in result.nodes if n["kind"] in ("cell", "group")][:1]
        result.document(ids, provider=self._provider([]))
        spent = result.token_usage.total
        assert spent > 0

        with pytest.raises(AiDocError, match="already spent"):
            result.document(ids, provider=self._provider([]), token_budget=spent)

    def test_the_budget_holds_without_a_usage_accumulator(self, lineage_excel):
        """aidoc's own entry point takes no accumulator unless asked."""
        from linexcel.aidoc import document_nodes

        calls: list[int] = []
        result = analyze(lineage_excel)
        ids = [n["id"] for n in result.nodes if n["kind"] in ("cell", "group")]

        with pytest.warns(UserWarning, match="never sent"):
            docs = document_nodes(
                result.graph,
                ids,
                provider=self._provider(calls),
                max_workers=1,
                token_budget=1,
            )

        assert len(calls) == 1 and len(docs) == 1

    def test_a_non_positive_budget_is_rejected(self, lineage_excel):
        result = analyze(lineage_excel)
        with pytest.raises(ValueError, match="token_budget must be > 0"):
            result.document(provider=self._provider([]), token_budget=0)

    def test_workbook_overview_refuses_to_start_over_budget(self, lineage_excel):
        from linexcel.aidoc import AiDocError

        result = analyze(lineage_excel)
        result.document_workbook(provider=self._provider([]))
        with pytest.raises(AiDocError, match="already spent"):
            result.document_workbook(
                provider=self._provider([]), token_budget=result.token_usage.total
            )


class TestWorkbookPresentationContext:
    """The overview dossier carries what the sheet screenshots show."""

    def test_context_reaches_the_dossier(self, lineage_excel):
        from linexcel.aidoc import build_workbook_dossier

        result = analyze(lineage_excel)
        dossier = build_workbook_dossier(result.graph, context=result.workbook_context)
        ventes = next(s for s in dossier["sheets"] if s["name"] == "Ventes")
        assert ventes["preview"], "the first rows a reader sees must be included"
        assert ventes["formula_cells"] == 100, "lineage facts must survive the merge"

    def test_comments_and_layout_reach_the_dossier(self, lineage_excel):
        import json

        from linexcel.aidoc import build_workbook_dossier

        result = analyze(lineage_excel)
        dossier = build_workbook_dossier(result.graph, context=result.workbook_context)
        blob = json.dumps(dossier, ensure_ascii=False, default=str)
        comments = [c for s in dossier["sheets"] for c in s.get("comments", [])]
        assert comments, "cell comments are context no formula can express"
        assert "freeze_panes" in blob or "merged_ranges" in blob

    def test_empty_padding_is_not_sent(self, lineage_excel):
        """A preview is a fixed rectangle; its blank cells cost tokens for nothing."""
        from linexcel.aidoc import build_workbook_dossier

        result = analyze(lineage_excel)
        dossier = build_workbook_dossier(result.graph, context=result.workbook_context)
        for sheet in dossier["sheets"]:
            for row in sheet.get("preview", []):
                assert row["values"][-1] not in (None, "")

    def test_without_context_the_dossier_is_unchanged(self, lineage_excel):
        from linexcel.aidoc import build_workbook_dossier

        dossier = build_workbook_dossier(analyze(lineage_excel).graph)
        assert all("preview" not in sheet for sheet in dossier["sheets"])

    def test_a_date_cell_previews_as_the_day_it_holds(self):
        """openpyxl reads a date back as a midnight datetime, and
        "2026-01-03T00:00:00" is a timestamp nobody typed into the sheet."""
        import datetime
        import io as _io

        from openpyxl import Workbook

        from linexcel.aidoc import build_workbook_dossier

        workbook = Workbook()
        workbook.active["A1"] = datetime.date(2026, 1, 3)
        workbook.active["A1"].number_format = "yyyy-mm-dd"
        workbook.active["A2"] = datetime.datetime(2026, 1, 3, 14, 30)
        workbook.active["A2"].number_format = "yyyy-mm-dd hh:mm:ss"
        workbook.active["A3"] = datetime.datetime(2026, 1, 3, 0, 0)
        workbook.active["A3"].number_format = "yyyy-mm-dd hh:mm:ss"
        buffer = _io.BytesIO()
        workbook.save(buffer)

        result = analyze(buffer.getvalue(), filename="d.xlsx")
        sheet = result.workbook_context["sheets"][0]
        assert [row["values"][0] for row in sheet["preview"]] == [
            "2026-01-03",
            "2026-01-03 14:30:00",
            "2026-01-03 00:00:00",
        ]
        dossier = build_workbook_dossier(result.graph, context=result.workbook_context)
        assert [row["values"][0] for row in dossier["sheets"][0]["preview"]] == [
            "2026-01-03",
            "2026-01-03 14:30:00",
            "2026-01-03 00:00:00",
        ]

    def test_document_workbook_sends_the_context_by_default(self, lineage_excel):
        seen: list[str] = []

        def provider(system_prompt, user_prompt, *, temperature=0.2):
            seen.append(user_prompt)
            return "overview"

        analyze(lineage_excel).document_workbook(provider=provider)
        assert "preview" in seen[0]

    def test_include_context_false_keeps_cell_contents_local(self, lineage_excel):
        seen: list[str] = []

        def provider(system_prompt, user_prompt, *, temperature=0.2):
            seen.append(user_prompt)
            return "overview"

        analyze(lineage_excel).document_workbook(
            provider=provider, include_context=False
        )
        assert "preview" not in seen[0]

    def test_a_result_without_workbook_bytes_still_documents(self, lineage_excel):
        """Context needs the file; a result rebuilt from a graph has none.

        An overview without the presentation cues beats raising on a result the
        caller assembled themselves.
        """
        payload = analyze_workbook(lineage_excel, "test.xlsx")
        result = LineageResult(graph=payload["graph"], engine=payload["engine"])

        overview = result.document_workbook(
            provider=lambda system, user, *, temperature=0.2: "overview"
        )
        assert overview == "overview"

    def test_an_oversized_dossier_sheds_the_preview_last(self):
        """Shrinking must cost the least useful part first, not the newest one."""
        from linexcel import aidoc

        dossier = {
            "sheets": [{"name": "S", "preview": [{"row": 1, "values": ["x"]}] * 40}],
            "formula_patterns": [
                {"formula": "=SUMIFS(Data!A:A, Data!B:B, $B4, C:C, C$3)", "cells": 1}
            ]
            * 400,
            "vba_procedures": [{"procedure": "Proc", "module": "Module1"}] * 50,
        }
        blob = aidoc._fit_workbook_dossier(dossier)

        assert len(blob) <= aidoc.MAX_WORKBOOK_DOSSIER_CHARS
        assert len(dossier["formula_patterns"]) == 5, "the pattern tail goes first"
        assert dossier["sheets"][0]["preview"], "preview goes only as a last resort"


class TestVba:
    MODULES = {
        "Module1": (
            """Public Sub MAJ()
        total = WorksheetFunction.Sum(Worksheets("Ventes").Range("D2:D101"))
        Worksheets("Synthese").Range("B10").Value = total * Taux()
        Cells(3, 2) = "ok"
    End Sub
    Private Function Taux() As Double
        Taux = Sheets("Params").Range("A1").Value
    End Function
"""
        )
    }

    def test_procedures_and_calls(self):
        procs = analyze_vba(self.MODULES)
        names = {p.name: p for p in procs}
        assert set(names) == {"MAJ", "Taux"}
        assert names["MAJ"].calls == ["Taux"]
        assert names["MAJ"].kind == "Sub"
        assert names["Taux"].kind == "Function"

    def test_read_write_detection(self):
        procs = analyze_vba(self.MODULES)
        maj = next(p for p in procs if p.name == "MAJ")
        accesses = {(r.sheet, r.ref): r.access for r in maj.refs}
        assert accesses[("Ventes", "D2:D101")] == "read"
        assert accesses[("Synthese", "B10")] == "write"
        assert accesses[(None, "B3")] == "write"

    def test_comments_ignored(self):
        procs = analyze_vba(
            {"M": 'Sub S()\n    \' Range("Z9") = 1 in comment\nEnd Sub\n'}
        )
        assert procs[0].refs == []

    def test_bracket_shortcut_references_are_found(self):
        """`[A1] = 1` is the shortcut form; the docstring advertises it."""
        procs = analyze_vba({"M": "Sub S()\n    [B7] = 1\n    x = [D2:E4]\nEnd Sub\n"})
        accesses = {(r.sheet, r.ref): r.access for r in procs[0].refs}
        assert accesses == {(None, "B7"): "write", (None, "D2:E4"): "read"}

    def test_a_procedure_missing_its_end_still_yields_its_body(self):
        """A truncated module is what a partial extraction leaves behind."""
        procs = analyze_vba({"M": 'Sub S()\n    Range("A1") = 1\n'})
        assert [p.name for p in procs] == ["S"]
        assert procs[0].line_end == 2
        assert [(r.ref, r.access) for r in procs[0].refs] == [("A1", "write")]

    def test_calls_are_case_insensitive(self):
        """VBA is case-insensitive: `helper` must reach `Helper`."""
        procs = analyze_vba(
            {
                "M": (
                    "Sub A()\n    Call helper\nEnd Sub\n"
                    "Sub Helper()\n    x = 1\nEnd Sub\n"
                )
            }
        )
        assert next(p for p in procs if p.name == "A").calls == ["Helper"]


class TestVbaGraph:
    """VBA branch of the analyzer.

    openpyxl cannot author a ``vbaProject.bin``, so extraction is stubbed and
    the graph-building code below it runs for real.
    """

    @staticmethod
    def _graph(workbook: bytes, monkeypatch, modules: dict[str, str]) -> dict:
        monkeypatch.setattr(
            "linexcel.analyzer.extract_vba_modules",
            lambda data, filename, warnings=None: dict(modules),
        )
        return analyze_workbook(workbook, "macro.xlsm")["graph"]

    def test_call_edge_links_the_two_procedures(self, lineage_excel, monkeypatch):
        graph = self._graph(
            lineage_excel,
            monkeypatch,
            {
                "Module1": (
                    "Public Sub Refresh()\n"
                    '    Worksheets("Synthese").Range("B10").Value = Rate()\n'
                    "End Sub\n"
                    "Private Function Rate() As Double\n"
                    '    Rate = Sheets("Params").Range("A1").Value\n'
                    "End Function\n"
                )
            },
        )
        assert {n["id"] for n in graph["nodes"] if n["kind"] == "vba"} == {
            "vp:Module1.Refresh",
            "vp:Module1.Rate",
        }
        calls = [
            (e["source"], e["target"]) for e in graph["edges"] if e["kind"] == "call"
        ]
        assert calls == [("vp:Module1.Refresh", "vp:Module1.Rate")]

    def test_read_and_write_edges_reach_the_sheets(self, lineage_excel, monkeypatch):
        graph = self._graph(
            lineage_excel,
            monkeypatch,
            {
                "Module1": (
                    "Sub Refresh()\n"
                    '    Worksheets("Synthese").Range("B10").Value = 1\n'
                    '    x = Sheets("Params").Range("A1").Value\n'
                    "End Sub\n"
                )
            },
        )
        by_id = {n["id"]: n for n in graph["nodes"]}
        writes = [e for e in graph["edges"] if e["kind"] == "vba-write"]
        reads = [e for e in graph["edges"] if e["kind"] == "vba-read"]
        assert [by_id[e["target"]]["label"] for e in writes] == ["Synthese!B10"]
        assert [by_id[e["source"]]["label"] for e in reads] == ["Params!A1"]

    def test_call_resolves_in_the_calling_module_first(
        self, lineage_excel, monkeypatch
    ):
        graph = self._graph(
            lineage_excel,
            monkeypatch,
            {
                "ModA": (
                    "Sub Run()\n    Helper\nEnd Sub\nSub Helper()\n    x = 1\nEnd Sub\n"
                ),
                "ModB": "Sub Helper()\n    y = 2\nEnd Sub\n",
            },
        )
        calls = {
            (e["source"], e["target"]) for e in graph["edges"] if e["kind"] == "call"
        }
        assert ("vp:ModA.Run", "vp:ModA.Helper") in calls
        assert ("vp:ModA.Run", "vp:ModB.Helper") not in calls

    def test_ambiguous_call_stays_unresolved(self, lineage_excel, monkeypatch):
        """A name owned by two other modules must not pick one arbitrarily.

        ``Solo`` is the positive control: without it an empty edge list would
        also pass if call edges stopped being emitted altogether.
        """
        graph = self._graph(
            lineage_excel,
            monkeypatch,
            {
                "ModA": "Sub Helper()\n    x = 1\nEnd Sub\n",
                "ModB": "Sub Helper()\n    y = 2\nEnd Sub\n",
                "ModC": (
                    "Sub Run()\n    Helper\n    Solo\nEnd Sub\n"
                    "Sub Solo()\n    z = 3\nEnd Sub\n"
                ),
            },
        )
        calls = {
            (e["source"], e["target"]) for e in graph["edges"] if e["kind"] == "call"
        }
        assert calls == {("vp:ModC.Run", "vp:ModC.Solo")}

    def test_call_resolution_ignores_casing(self, lineage_excel, monkeypatch):
        """VBA is case-insensitive, so ModA.Taux owns the local `Taux` call.

        The lowercase twin in ModB must not capture it, and the function's own
        `Taux = 1` return assignment is not a call to anything.
        """
        graph = self._graph(
            lineage_excel,
            monkeypatch,
            {
                "ModA": (
                    "Sub Run()\n    Taux\nEnd Sub\n"
                    "Function Taux()\n    Taux = 1\nEnd Function\n"
                ),
                "ModB": "Function taux()\n    taux = 2\nEnd Function\n",
            },
        )
        calls = sorted(
            (e["source"], e["target"]) for e in graph["edges"] if e["kind"] == "call"
        )
        assert calls == [("vp:ModA.Run", "vp:ModA.Taux")]

    def test_member_access_is_not_a_call(self, lineage_excel, monkeypatch):
        """`.Value` is member access even when a procedure is named Value."""
        graph = self._graph(
            lineage_excel,
            monkeypatch,
            {
                "Helpers": (
                    "Public Function Value() As Double\n    Value = 0\nEnd Function\n"
                ),
                "Main": (
                    "Sub Refresh()\n"
                    '    total = Worksheets("S").Range("A1").Value\n'
                    "End Sub\n"
                ),
            },
        )
        assert [e for e in graph["edges"] if e["kind"] == "call"] == []

    def test_module_qualified_call_resolves_exactly(self, lineage_excel, monkeypatch):
        graph = self._graph(
            lineage_excel,
            monkeypatch,
            {
                "ModA": "Sub Run()\n    ModB.Helper\nEnd Sub\n",
                "ModB": "Sub Helper()\n    x = 1\nEnd Sub\n",
            },
        )
        calls = [
            (e["source"], e["target"]) for e in graph["edges"] if e["kind"] == "call"
        ]
        assert calls == [("vp:ModA.Run", "vp:ModB.Helper")]


def png_bytes(width: int = 40, height: int = 30) -> bytes:
    """Enough of a PNG for the renderer to read its dimensions back."""
    return (
        b"\x89PNG\r\n\x1a\n" + b"\x00\x00\x00\rIHDR" + struct.pack(">II", width, height)
    )


class TestScreenshotsPerSheet:
    """A page can only be shown under a sheet if it *is* that sheet.

    LibreOffice is asked to put each sheet on one page, which is what makes the
    two line up at all: under the workbook's own print layout a long sheet spans
    several pages and short ones share one, and no page number maps back.
    """

    @staticmethod
    def _renderer(monkeypatch, pages: list[bytes]):
        """Stand in for LibreOffice and pdftoppm, producing ``pages``."""
        calls: list[list[str]] = []

        def fake_run(command, **_kwargs):
            calls.append(command)
            if command[0] == "libreoffice":
                pdf_dir = Path(command[command.index("--outdir") + 1])
                (pdf_dir / "workbook.pdf").write_bytes(b"%PDF-1.4")
            else:
                width = len(str(len(pages)))
                for index, page in enumerate(pages, start=1):
                    Path(f"{command[-1]}-{index:0{width}d}.png").write_bytes(page)

        monkeypatch.setattr(
            "linexcel.insights.shutil.which",
            {"libreoffice": "libreoffice", "pdftoppm": "pdftoppm"}.get,
        )
        monkeypatch.setattr("linexcel.insights.subprocess.run", fake_run)
        return calls

    def test_one_page_per_sheet_is_keyed_by_sheet_name(
        self, lineage_excel, monkeypatch, tmp_path
    ):
        self._renderer(monkeypatch, [png_bytes()] * 3)
        shots = analyze(lineage_excel).save_screenshots(tmp_path)
        assert isinstance(shots, dict)
        assert list(shots) == ["Ventes", "Synthese", "Params"]
        assert [p.name for pages in shots.values() for p in pages] == [
            "workbook-Ventes.png",
            "workbook-Synthese.png",
            "workbook-Params.png",
        ]

    def test_the_single_page_filter_is_what_is_asked_for(
        self, lineage_excel, monkeypatch, tmp_path
    ):
        calls = self._renderer(monkeypatch, [png_bytes()] * 3)
        analyze(lineage_excel).save_screenshots(tmp_path)
        assert "SinglePageSheets" in calls[0][calls[0].index("--convert-to") + 1]

    def test_print_pages_are_a_flat_list_under_the_workbook_layout(
        self, lineage_excel, monkeypatch, tmp_path
    ):
        calls = self._renderer(monkeypatch, [png_bytes()] * 2)
        shots = analyze(lineage_excel).save_screenshots(tmp_path, per_sheet=False)
        assert isinstance(shots, list)
        assert calls[0][calls[0].index("--convert-to") + 1] == "pdf"
        assert [p.name for p in shots] == ["workbook-1.png", "workbook-2.png"]

    def test_a_page_count_that_does_not_match_is_not_mapped_by_guesswork(
        self, lineage_excel, monkeypatch, tmp_path
    ):
        """An older LibreOffice ignores the option and paginates as it prints.

        Naming those pages after sheets would file each image under a sheet it
        may not show, so the flat list is returned instead.
        """
        self._renderer(monkeypatch, [png_bytes()] * 5)  # 5 pages, 3 sheets
        shots = analyze(lineage_excel).save_screenshots(tmp_path)
        assert isinstance(shots, list)
        assert [p.name for p in shots] == [f"workbook-{n}.png" for n in range(1, 6)]

    def test_a_sheet_with_nothing_on_it_gets_no_image(
        self, lineage_excel, monkeypatch, tmp_path
    ):
        """An empty sheet renders one pixel tall and would read as a broken
        image; the sheet is simply left without one."""
        self._renderer(monkeypatch, [png_bytes(), png_bytes(97, 1), png_bytes()])
        shots = analyze(lineage_excel).save_screenshots(tmp_path)
        assert isinstance(shots, dict)
        assert list(shots) == ["Ventes", "Params"]

    def test_pages_of_an_earlier_run_are_not_returned_as_this_one(
        self, lineage_excel, monkeypatch, tmp_path
    ):
        """pdftoppm pads the page number to the width of the page count, so a
        reused directory holds both "-1.png" and "-01.png"."""
        self._renderer(monkeypatch, [png_bytes()] * 12)
        analyze(lineage_excel).save_screenshots(tmp_path, per_sheet=False)
        self._renderer(monkeypatch, [png_bytes()] * 2)
        shots = analyze(lineage_excel).save_screenshots(tmp_path, per_sheet=False)
        assert isinstance(shots, list)
        assert [p.name for p in shots] == ["workbook-1.png", "workbook-2.png"]

    def test_a_sheet_name_no_filesystem_accepts_still_gets_a_file(
        self, monkeypatch, tmp_path
    ):
        from openpyxl import Workbook

        workbook = Workbook()
        workbook.active.title = "O'Brien & Café"
        workbook.create_sheet("O Brien  Caf")  # sanitizes to the same name
        buffer = io.BytesIO()
        workbook.save(buffer)

        self._renderer(monkeypatch, [png_bytes()] * 2)
        shots = analyze(buffer.getvalue(), filename="w.xlsx").save_screenshots(tmp_path)
        assert isinstance(shots, dict)
        assert list(shots) == ["O'Brien & Café", "O Brien  Caf"]
        assert [p.name for pages in shots.values() for p in pages] == [
            "w-O-Brien-Caf.png",
            "w-O-Brien-Caf-2.png",
        ]


class _StubVbaParser:
    """Stand-in for ``olevba.VBA_Parser``.

    The real one needs a ``vbaProject.bin``, which is an OLE compound file that
    ``openpyxl`` cannot author — so no generated fixture can carry macros and
    the extraction loop had never run in a test. This drives it with the tuples
    olevba yields, including the shapes that only turn up on real files: a
    module name and a body coming back as undecoded bytes.
    """

    #: Set per test: (subfile, stream_path, vba_filename, code) tuples.
    macros: tuple = ()
    #: Set per test to raise from the matching call.
    fail_on_open = False
    fail_on_extract = False
    detects = True
    closed = False

    def __init__(self, filename, data=None, **_kwargs):
        if type(self).fail_on_open:
            raise OSError("not an OLE file")
        self.filename = filename

    def detect_vba_macros(self):
        return type(self).detects

    def extract_macros(self):
        for index, macro in enumerate(type(self).macros):
            if type(self).fail_on_extract and index:
                raise ValueError("stream 2 is corrupt")
            yield macro

    def close(self):
        type(self).closed = True


class _VisionStub:
    """A provider that accepts an image, and records what it was handed."""

    def __init__(self, reply: str = "A grid of blue input cells."):
        self.reply = reply
        self.calls: list[dict] = []

    def generate(self, system_prompt, user_prompt, *, temperature=0.2, max_tokens=None):
        raise AssertionError("a screenshot must go through generate_with_image")

    def generate_with_image(
        self,
        system_prompt,
        user_prompt,
        image,
        *,
        media_type="image/png",
        temperature=0.2,
        max_tokens=None,
    ):
        from linexcel.aidoc import TokenUsage

        self.calls.append(
            {
                "system": system_prompt,
                "user": user_prompt,
                "bytes": image,
                "media_type": media_type,
            }
        )
        if isinstance(self.reply, Exception):
            raise self.reply
        return self.reply, TokenUsage(input_tokens=800, output_tokens=40, requests=1)


class TestScreenshotDescriptions:
    """The one thing linexcel documents from a picture rather than the graph.

    Colour conventions, conditional formatting and layout never reach a text
    dossier, so a model is shown the screenshot itself. What matters is that it
    is *shown* one — a text-only provider must fail loudly rather than have the
    image dropped from the request.
    """

    def test_each_sheet_comes_back_described(self, lineage_excel, tmp_path):
        shot = tmp_path / "Ventes.png"
        shot.write_bytes(png_bytes())
        vision = _VisionStub()
        result = analyze(lineage_excel)
        docs = result.describe_screenshots({"Ventes": [shot]}, provider=vision)
        assert docs == {"Ventes": "A grid of blue input cells."}
        assert vision.calls[0]["user"] == "Sheet: Ventes"
        assert vision.calls[0]["bytes"] == png_bytes()

    def test_the_sheet_name_reaches_the_model_but_the_dossier_does_not(
        self, lineage_excel, tmp_path
    ):
        """The evidence is the image; sending the graph too would blur that."""
        shot = tmp_path / "Ventes.png"
        shot.write_bytes(png_bytes())
        vision = _VisionStub()
        analyze(lineage_excel).describe_screenshots({"Ventes": [shot]}, provider=vision)
        (call,) = vision.calls
        assert "formula" not in call["user"].lower()
        assert "screenshot" in call["system"].lower()

    def test_the_media_type_follows_the_file(self, lineage_excel, tmp_path):
        shot = tmp_path / "Ventes.jpg"
        shot.write_bytes(png_bytes())
        vision = _VisionStub()
        analyze(lineage_excel).describe_screenshots({"Ventes": [shot]}, provider=vision)
        assert vision.calls[0]["media_type"] == "image/jpeg"

    def test_one_image_per_sheet_may_be_given_without_a_list(
        self, lineage_excel, tmp_path
    ):
        """A lone path is a path: indexing it would send the letter ``C``."""
        shot = tmp_path / "Ventes.png"
        shot.write_bytes(png_bytes())
        vision = _VisionStub()
        docs = analyze(lineage_excel).describe_screenshots(
            {"Ventes": str(shot)}, provider=vision
        )
        assert set(docs) == {"Ventes"}
        assert vision.calls[0]["bytes"] == png_bytes()

    def test_a_flat_list_of_pages_is_keyed_by_file_name(self, lineage_excel, tmp_path):
        """No page belongs to one sheet, so the page is what gets named."""
        pages = []
        for index in (1, 2):
            page = tmp_path / f"workbook-{index}.png"
            page.write_bytes(png_bytes())
            pages.append(page)
        vision = _VisionStub()
        docs = analyze(lineage_excel).describe_screenshots(pages, provider=vision)
        assert set(docs) == {"workbook-1", "workbook-2"}

    def test_a_text_only_provider_is_refused_by_name(self, lineage_excel, tmp_path):
        from linexcel.aidoc import AiDocError

        shot = tmp_path / "Ventes.png"
        shot.write_bytes(png_bytes())

        class TextOnly:
            def generate(self, system, user, *, temperature=0.2, max_tokens=None):
                return "text"

        result = analyze(lineage_excel)
        with pytest.raises(AiDocError, match="generate_with_image"):
            result.describe_screenshots({"Ventes": [shot]}, provider=TextOnly())

    def test_an_oversized_image_is_named_rather_than_posted(
        self, lineage_excel, tmp_path, monkeypatch
    ):
        from linexcel import aidoc

        monkeypatch.setattr(aidoc, "MAX_IMAGE_BYTES", 8)
        shot = tmp_path / "Ventes.png"
        shot.write_bytes(png_bytes())
        vision = _VisionStub()
        with pytest.raises(aidoc.AiDocError, match="Ventes"):
            analyze(lineage_excel).describe_screenshots(
                {"Ventes": [shot]}, provider=vision
            )
        assert vision.calls == []

    def test_one_failure_does_not_discard_the_others(self, lineage_excel, tmp_path):
        from linexcel.aidoc import AiDocError, describe_images

        good = tmp_path / "Ventes.png"
        good.write_bytes(png_bytes())

        class Choosy(_VisionStub):
            def generate_with_image(self, system, user, image, **kwargs):
                if "Params" in user:
                    raise AiDocError("model refused")
                return super().generate_with_image(system, user, image, **kwargs)

        with pytest.warns(UserWarning, match="not described"):
            docs = describe_images({"Ventes": good, "Params": good}, provider=Choosy())
        assert set(docs) == {"Ventes"}

    def test_the_tokens_are_counted_on_the_result(self, lineage_excel, tmp_path):
        shot = tmp_path / "Ventes.png"
        shot.write_bytes(png_bytes())
        result = analyze(lineage_excel)
        result.describe_screenshots({"Ventes": [shot]}, provider=_VisionStub())
        assert result.token_usage.total == 840

    def test_the_budget_stops_the_next_image(self, lineage_excel, tmp_path):
        shot = tmp_path / "Ventes.png"
        shot.write_bytes(png_bytes())
        vision = _VisionStub()
        result = analyze(lineage_excel)
        with pytest.warns(UserWarning, match="budget"):
            result.describe_screenshots(
                {"A": shot, "B": shot, "C": shot},
                provider=vision,
                token_budget=1000,
            )
        assert len(vision.calls) == 2  # the third is never sent

    def test_the_description_travels_with_the_report(self, lineage_excel, tmp_path):
        shot = tmp_path / "Ventes.png"
        shot.write_bytes(png_bytes())
        html = analyze(lineage_excel).to_html(
            screenshots={"Ventes": [shot]},
            screenshot_docs={"Ventes": "Blue inputs, black formulas."},
        )
        assert "Blue inputs, black formulas." in html
        assert '"screenshotDocs"' in html


@pytest.fixture()
def stub_olevba(monkeypatch):
    """Install the stub and reset its per-test state."""
    import oletools.olevba

    for attribute, value in (
        ("macros", ()),
        ("fail_on_open", False),
        ("fail_on_extract", False),
        ("detects", True),
        ("closed", False),
    ):
        setattr(_StubVbaParser, attribute, value)
    monkeypatch.setattr(oletools.olevba, "VBA_Parser", _StubVbaParser)
    return _StubVbaParser


class TestVbaExtraction:
    """The olevba boundary: everything below it was tested, this was not."""

    def test_modules_are_read_and_named_after_their_stream(self, stub_olevba):
        from linexcel.vba import extract_vba_modules

        stub_olevba.macros = (
            ("x.xlsm", "VBA/Module1", "VBA/Module1.bas", "Sub A()\nEnd Sub\n"),
            ("x.xlsm", "VBA/Sheet1", "VBA/Sheet1.cls", "Sub B()\nEnd Sub\n"),
        )
        modules = extract_vba_modules(b"", "x.xlsm")
        assert list(modules) == ["Module1", "Sheet1"]
        assert modules["Module1"].startswith("Sub A()")
        assert stub_olevba.closed, "the parser must be closed even on success"

    def test_an_undecodable_stream_is_decoded_rather_than_dropped(self, stub_olevba):
        from linexcel.vba import extract_vba_modules

        stub_olevba.macros = (
            ("x", "p", b"VBA/M\xff.bas", b"Sub A()\n  x = \xff\nEnd Sub\n"),
        )
        modules = extract_vba_modules(b"", "x.xlsm")
        assert list(modules) == ["M�"]
        assert "Sub A()" in modules["M�"]

    def test_two_streams_of_one_module_are_joined(self, stub_olevba):
        from linexcel.vba import extract_vba_modules

        stub_olevba.macros = (
            ("x", "p", "VBA/M.bas", "Sub A()\nEnd Sub\n"),
            ("x", "p", "VBA/M.cls", "Sub B()\nEnd Sub\n"),
        )
        modules = extract_vba_modules(b"", "x.xlsm")
        assert list(modules) == ["M"]
        assert "Sub A()" in modules["M"] and "Sub B()" in modules["M"]

    def test_an_empty_stream_contributes_no_module(self, stub_olevba):
        from linexcel.vba import extract_vba_modules

        stub_olevba.macros = (("x", "p", "VBA/M.bas", "   \n"),)
        warnings: list[str] = []
        assert extract_vba_modules(b"", "x.xlsm", warnings) == {}
        assert warnings == [], "an empty module stream is not a failure"

    def test_macros_announced_but_no_stream_handed_back_is_reported(self, stub_olevba):
        """olevba saying yes and then yielding nothing is the anomaly."""
        from linexcel.vba import extract_vba_modules

        stub_olevba.macros = ()
        warnings: list[str] = []
        assert extract_vba_modules(b"", "x.xlsm", warnings) == {}
        assert warnings == [
            "the workbook declares VBA macros but no module stream could be read"
        ]

    def test_a_workbook_without_macros_says_nothing(self, stub_olevba):
        from linexcel.vba import extract_vba_modules

        stub_olevba.detects = False
        warnings: list[str] = []
        assert extract_vba_modules(b"", "plain.xlsx", warnings) == {}
        assert warnings == [], "having no macros is not a defect to report"

    def test_an_unreadable_project_is_reported(self, stub_olevba):
        from linexcel.vba import extract_vba_modules

        stub_olevba.fail_on_open = True
        warnings: list[str] = []
        assert extract_vba_modules(b"", "x.xlsm", warnings) == {}
        assert warnings == ["the VBA project could not be opened: not an OLE file"]

    def test_a_failure_part_way_keeps_what_was_read(self, stub_olevba):
        """Dropping the modules already read trades a partial answer for none."""
        from linexcel.vba import extract_vba_modules

        stub_olevba.macros = (
            ("x", "p", "VBA/M1.bas", "Sub A()\nEnd Sub\n"),
            ("x", "p", "VBA/M2.bas", "Sub B()\nEnd Sub\n"),
        )
        stub_olevba.fail_on_extract = True
        warnings: list[str] = []
        modules = extract_vba_modules(b"", "x.xlsm", warnings)
        assert list(modules) == ["M1"]
        assert warnings == [
            "VBA extraction stopped after 1 module(s) read: stream 2 is corrupt"
        ]
        assert stub_olevba.closed, "the parser must be closed even on failure"

    def test_the_analyzer_surfaces_the_reason_in_its_warnings(
        self, lineage_excel, stub_olevba
    ):
        """A macro workbook showing no VBA must not do so silently."""
        stub_olevba.fail_on_open = True
        graph = analyze_workbook(lineage_excel, "macro.xlsm")["graph"]
        assert graph["meta"]["stats"]["vbaProcs"] == 0
        assert any(
            "VBA project could not be opened" in w for w in graph["meta"]["warnings"]
        )

    def test_a_real_macro_workbook_reaches_the_graph(self, lineage_excel, stub_olevba):
        """End to end through the real extraction code, only olevba stubbed."""
        stub_olevba.macros = (
            (
                "x",
                "p",
                "VBA/Module1.bas",
                'Sub Refresh()\n    Worksheets("Synthese").Range("B10").Value = 1\n'
                "End Sub\n",
            ),
        )
        graph = analyze_workbook(lineage_excel, "macro.xlsm")["graph"]
        assert graph["meta"]["stats"]["vbaModules"] == 1
        assert graph["meta"]["stats"]["vbaProcs"] == 1
        assert [n["label"] for n in graph["nodes"] if n["kind"] == "vba"] == [
            "Module1.Refresh"
        ]


class TestSheetClassModules:
    """Excel writes a class module per worksheet whether or not anybody uses it.

    Found on the real `macros.xlsm`: a workbook with a single macro module was
    reported as having five, because the four Excel wrote for its sheets and
    ThisWorkbook are not empty — they hold `Attribute VB_*` declarations.
    """

    SHEET_STUB = (
        'Attribute VB_Name = "Feuil1"\n'
        'Attribute VB_Base = "0{00020820-0000-0000-C000-000000000046}"\n'
        "Attribute VB_GlobalNameSpace = False\n"
        "Attribute VB_Exposed = True\n"
    )

    def test_an_attribute_only_module_is_not_a_module(self, stub_olevba):
        from linexcel.vba import extract_vba_modules

        stub_olevba.macros = (
            ("x", "p", "VBA/Feuil1.cls", self.SHEET_STUB),
            ("x", "p", "VBA/ThisWorkbook.cls", self.SHEET_STUB),
            (
                "x",
                "p",
                "VBA/Module1.bas",
                'Attribute VB_Name = "Module1"\nSub Go()\nEnd Sub\n',
            ),
        )
        warnings: list[str] = []
        assert list(extract_vba_modules(b"", "m.xlsm", warnings)) == ["Module1"]
        assert warnings == []

    def test_a_sheet_module_holding_an_event_handler_is_kept(self, stub_olevba):
        """The filter must drop empty shells, not sheet code someone wrote."""
        from linexcel.vba import extract_vba_modules

        stub_olevba.macros = (
            (
                "x",
                "p",
                "VBA/Feuil1.cls",
                self.SHEET_STUB
                + "Private Sub Worksheet_Change(ByVal Target As Range)\n"
                '    Range("Z1") = 1\n'
                "End Sub\n",
            ),
        )
        modules = extract_vba_modules(b"", "m.xlsm")
        assert list(modules) == ["Feuil1"]
        procs = analyze_vba(modules)
        assert [p.name for p in procs] == ["Worksheet_Change"]

    def test_a_workbook_of_nothing_but_shells_stays_quiet(self, stub_olevba):
        """A macro-enabled workbook nobody wrote code in is not a defect."""
        from linexcel.vba import extract_vba_modules

        stub_olevba.macros = (("x", "p", "VBA/Feuil1.cls", self.SHEET_STUB),)
        warnings: list[str] = []
        assert extract_vba_modules(b"", "m.xlsm", warnings) == {}
        assert warnings == []


FIXTURES = Path(__file__).parent / "fixtures"


class TestVbaOnARealWorkbook:
    """The one thing a stub cannot prove: olevba on a real vbaProject.bin.

    See ``tests/fixtures/README.md`` for how the workbook is built. The test
    skips itself while the file is absent so the suite stays green without it.
    """

    WORKBOOK = FIXTURES / "macros.xlsm"

    @pytest.fixture()
    def graph(self):
        if not self.WORKBOOK.exists():
            pytest.skip(f"{self.WORKBOOK.name} absent — see fixtures/README.md")
        return analyze(self.WORKBOOK).graph

    def test_the_modules_come_out_of_the_real_container(self, graph):
        # One, not five: Excel also writes a class module per worksheet and one
        # for ThisWorkbook. They come out of olevba as a handful of Attribute
        # declarations and no statement, and are not modules anybody wrote.
        assert graph["meta"]["stats"]["vbaModules"] == 1
        assert graph["meta"]["stats"]["vbaProcs"] == 2
        assert not [
            w for w in graph["meta"]["warnings"] if "VBA" in w or "oletools" in w
        ], "a real macro workbook must raise no extraction warning"

    def test_both_procedures_keep_their_declared_kind(self, graph):
        procs = {n["label"]: n for n in graph["nodes"] if n["kind"] == "vba"}
        assert set(procs) == {"Module1.Refresh", "Module1.Rate"}
        assert procs["Module1.Refresh"]["procKind"] == "Sub"
        assert procs["Module1.Rate"]["procKind"] == "Function"

    def test_the_call_edge_survives_the_round_trip(self, graph):
        calls = [
            (e["source"], e["target"]) for e in graph["edges"] if e["kind"] == "call"
        ]
        assert calls == [("vp:Module1.Refresh", "vp:Module1.Rate")]

    def test_the_sheet_reads_and_writes_reach_their_ranges(self, graph):
        by_id = {n["id"]: n for n in graph["nodes"]}
        written = {
            by_id[e["target"]]["label"]
            for e in graph["edges"]
            if e["kind"] == "vba-write"
        }
        read = {
            by_id[e["source"]]["label"]
            for e in graph["edges"]
            if e["kind"] == "vba-read"
        }
        assert "Synthesis!B10" in written
        assert "Params!A1" in read

    def test_a_reference_with_no_sheet_does_not_guess_one(self, graph):
        """``Cells(3, 2)`` acts on whatever sheet is active at run time.

        A static reader cannot know which, so it must not pick one: the target
        is an external-reference node, not a cell of some plausible sheet.
        """
        by_id = {n["id"]: n for n in graph["nodes"]}
        targets = [
            by_id[e["target"]]
            for e in graph["edges"]
            if e["kind"] == "vba-write" and e["source"] == "vp:Module1.Refresh"
        ]
        unqualified = [n for n in targets if n["kind"] == "opaque"]
        assert [n["label"] for n in unqualified] == ["VBA:?!B3"]

    def test_the_extracted_source_is_carried_into_the_report(self, graph):
        refresh = next(n for n in graph["nodes"] if n["label"] == "Module1.Refresh")
        assert "Worksheets(" in refresh["code"]
        assert refresh["module"] == "Module1"


class TestPowerQueryWorkbook:
    """A workbook fed by Get & Transform, read on a file Excel itself wrote.

    ``power_query.xlsx`` holds two M queries. ``BusyProducts`` reads the
    ``SalesTable`` of the ``Source`` sheet and lands on ``Loaded``;
    ``TinyConnectionOnly`` is computed and loaded nowhere. Both halves matter:
    the lineage has to cross the query, and a query that loads nothing must
    not invent a destination.
    """

    WORKBOOK = FIXTURES / "power_query.xlsx"

    def test_a_mashup_workbook_analyses_without_error(self):
        result = analyze(self.WORKBOOK)
        assert result.sheets == ["Loaded", "Source"]
        assert result.stats["totalFormulas"] == 0

    def test_each_query_is_a_node_carrying_its_m_source(self):
        result = analyze(self.WORKBOOK)
        queries = {n["label"]: n for n in result.nodes if n["kind"] == "query"}
        assert set(queries) == {"BusyProducts", "TinyConnectionOnly"}
        assert "Table.SelectRows" in queries["BusyProducts"]["code"]

    def test_the_lineage_crosses_the_query_from_source_to_destination(self):
        """``Source!A1:B4`` → ``BusyProducts`` → ``Loaded!A1:B3``."""
        result = analyze(self.WORKBOOK)
        ids = {n["label"]: n["id"] for n in result.nodes}
        edges = {(e["source"], e["target"]): e["kind"] for e in result.graph["edges"]}
        assert edges[ids["Source!A1:B4"], ids["BusyProducts"]] == "query"
        assert edges[ids["BusyProducts"], ids["Loaded!A1:B3"]] == "query-load"

    def test_a_query_says_which_table_it_reads(self):
        result = analyze(self.WORKBOOK)
        busy = next(n for n in result.nodes if n["label"] == "BusyProducts")
        assert busy["sources"] == [
            {
                "kind": "table",
                "target": "SalesTable",
                "function": "Excel.CurrentWorkbook",
            }
        ]
        assert busy["loadedTo"][0]["sheet"] == "Loaded"

    def test_a_connection_only_query_claims_no_destination(self):
        result = analyze(self.WORKBOOK)
        tiny = next(n for n in result.nodes if n["label"] == "TinyConnectionOnly")
        assert tiny["loadedTo"] == []
        assert not [e for e in result.graph["edges"] if e["source"] == tiny["id"]]

    def test_the_stats_count_the_queries_and_the_loaded_ones(self):
        result = analyze(self.WORKBOOK)
        assert result.stats["queries"] == 2
        assert result.stats["queriesLoaded"] == 1

    def test_the_warning_says_how_many_queries_feed_the_workbook(self):
        result = analyze(self.WORKBOOK)
        (warning,) = [w for w in result.warnings if "Power Query" in w]
        assert "2 Power Query queries" in warning


def _workbook_with_dynamic_table() -> bytes:
    """Workbook with a formal Excel Table object and a formula column."""
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(["Product", "Revenue", "Cost", "Margin"])
    ws.append(["Widget", 100, 60, "=B2-C2"])
    ws.append(["Gadget", 200, 120, "=B3-C3"])
    ws.append(["Gizmo", 150, 90, "=B4-C4"])
    tab = Table(displayName="SalesTable", ref="A1:D4")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tab)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _workbook_with_static_table() -> bytes:
    """Workbook with no formal table — just a header row over contiguous data."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Stats"
    ws.append(["Month", "Sales", "Region"])
    ws.append(["Jan", 1000, "North"])
    ws.append(["Feb", 1200, "South"])
    ws.append(["Mar", 900, "East"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestTableDetection:
    """Excel tables (dynamic and static) enrich the lineage graph."""

    def test_detect_dynamic_table(self):
        from openpyxl import load_workbook

        from linexcel.insights import detect_tables

        data = _workbook_with_dynamic_table()
        wb = load_workbook(io.BytesIO(data), read_only=False)
        tables = detect_tables(wb["Sales"])
        wb.close()
        assert len(tables) == 1
        t = tables[0]
        assert t["kind"] == "dynamic"
        assert t["name"] == "SalesTable"
        assert t["ref"] == "A1:D4"
        assert t["headers"] == ["Product", "Revenue", "Cost", "Margin"]
        assert t["data_rows"] == 3
        assert t["header_row"] == 1
        assert t["first_row"] == 2
        assert t["last_row"] == 4

    def test_detect_static_table(self):
        from openpyxl import load_workbook

        from linexcel.insights import detect_tables

        data = _workbook_with_static_table()
        wb = load_workbook(io.BytesIO(data), read_only=False)
        tables = detect_tables(wb["Stats"])
        wb.close()
        assert len(tables) == 1
        t = tables[0]
        assert t["kind"] == "static"
        assert t["ref"] == "A1:C4"
        assert t["headers"] == ["Month", "Sales", "Region"]
        assert t["data_rows"] == 3

    def test_static_table_formula_header_falls_back_to_column(self):
        """A formula cell in the header row must not leak into table_column.

        openpyxl returns formula text as a str; without the guard, a header
        like ``=SUM(...)`` becomes the table_column for every data cell in
        that column. The column letter is used as a readable fallback instead.
        """
        from openpyxl import Workbook, load_workbook

        from linexcel.insights import detect_tables

        wb = Workbook()
        ws = wb.active
        ws.title = "Sens"
        ws.append(["Month", "=SUM(B2:B3)", "Region"])
        ws.append(["Jan", 1000, "North"])
        ws.append(["Feb", 1200, "South"])
        buf = io.BytesIO()
        wb.save(buf)
        wb2 = load_workbook(io.BytesIO(buf.getvalue()), read_only=False)
        tables = detect_tables(wb2["Sens"])
        wb2.close()
        assert len(tables) == 1
        t = tables[0]
        assert t["headers"] == ["Month", "Column B", "Region"]

    def test_no_table_detected_on_blank_sheet(self):
        from openpyxl import load_workbook

        from linexcel.insights import detect_tables

        wb = load_workbook(io.BytesIO(_workbook_with_static_table()), read_only=False)
        ws = wb.active
        ws.delete_rows(1, ws.max_row)
        assert detect_tables(ws) == []
        wb.close()

    def test_workbook_context_carries_tables(self):
        data = _workbook_with_dynamic_table()
        result = analyze(data, filename="tables.xlsx")
        ctx = result.workbook_context
        sheet = ctx["sheets"][0]
        assert len(sheet["tables"]) == 1
        assert sheet["tables"][0]["name"] == "SalesTable"

    def test_graph_node_enriched_with_table(self):
        """A formula cell inside a table carries table_name/column/row."""
        data = _workbook_with_dynamic_table()
        result = analyze(data, filename="tables.xlsx")
        assert result.stats["tables"] == 1
        # The three Margin formulas (=B-C) group into one node on D2.
        enriched = [n for n in result.nodes if n.get("table_name") == "SalesTable"]
        assert enriched, "at least one node must be tagged with the table"
        node = enriched[0]
        assert node["table_column"] == "Margin"
        assert node["table_row"] == 0  # representative cell D2 → first data row

    def test_static_table_enriches_graph(self):
        data = _workbook_with_static_table()
        result = analyze(data, filename="static.xlsx")
        assert result.stats["tables"] >= 1
        # No formulas, but the graph stats should still report the table.
        assert result.stats["tables"] == 1

    def test_table_enrichment_does_not_break_empty_workbook(self):
        from openpyxl import Workbook

        wb = Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        result = analyze(buf.getvalue(), filename="empty.xlsx")
        assert result.stats["tables"] == 0
        assert not any("table_name" in n for n in result.nodes)

    def test_input_node_inside_table_is_enriched(self):
        """A referenced input cell that sits inside a table is tagged too."""
        from openpyxl import Workbook
        from openpyxl.worksheet.table import Table

        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["A", "B", "Result"])
        ws.append([10, 20, "=A2+B2"])  # single formula → no stretching
        ws.add_table(Table(displayName="T", ref="A1:C2"))
        buf = io.BytesIO()
        wb.save(buf)
        result = analyze(buf.getvalue(), filename="t.xlsx")
        # A2 and B2 are single-cell input precedents of C2; tagged with table T.
        tagged = [n for n in result.nodes if n.get("table_name") == "T"]
        columns = {n.get("table_column") for n in tagged}
        assert "A" in columns, "input cell A2 should carry its column header"
        assert "B" in columns, "input cell B2 should carry its column header"


class TestScanCeiling:
    """The formula sweep is bounded per sheet, and says what it left out.

    The ceiling exists for a file that *declares* far more than it holds — one
    stray cell at XFD1048576 makes the used range 17 billion cells — not to
    keep an honest workbook quick: sweeping costs about 0.7 µs per cell.
    """

    @staticmethod
    def sheet_of(rows: int, cols: int) -> bytes:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        for r in range(1, rows + 1):
            ws.cell(row=r, column=1, value=r)
            for c in range(2, cols + 1):
                ws.cell(row=r, column=c, value=f"=A{r}*{c}")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_a_sheet_under_the_ceiling_is_swept_whole(self):
        graph = analyze_workbook(self.sheet_of(30, 3), "scan.xlsx")["graph"]
        assert graph["meta"]["stats"]["totalFormulas"] == 60
        assert graph["meta"]["warnings"] == []

    def test_the_last_chunk_is_clipped_to_the_ceiling_not_dropped(self, monkeypatch):
        """Dropping it stopped a 4,000,000-cell budget at 3,600,000."""
        from linexcel import analyzer

        monkeypatch.setattr(analyzer, "MAX_CELLS_PER_SHEET", 30)
        monkeypatch.setattr(analyzer, "SCAN_CHUNK_ROWS", 100)
        graph = analyze_workbook(self.sheet_of(30, 3), "scan.xlsx")["graph"]
        # 30 cells of budget over 3 columns: rows 1-10, two formulas each
        assert graph["meta"]["stats"]["totalFormulas"] == 20

    def test_the_warning_names_the_first_row_left_out(self, monkeypatch):
        from linexcel import analyzer

        monkeypatch.setattr(analyzer, "MAX_CELLS_PER_SHEET", 30)
        monkeypatch.setattr(analyzer, "SCAN_CHUNK_ROWS", 100)
        graph = analyze_workbook(self.sheet_of(30, 3), "scan.xlsx")["graph"]
        (warning,) = graph["meta"]["warnings"]
        assert "scanned to row 10 of 30" in warning
        assert "missing from the lineage" in warning

    def test_a_wide_sheet_is_chunked_by_cells_rather_than_rows(self):
        """A 20,000-row chunk of a 16,384-column sheet is 327M strings."""
        from linexcel.analyzer import SCAN_CHUNK_CELLS, SCAN_CHUNK_ROWS, _chunk_rows

        assert _chunk_rows(3) == SCAN_CHUNK_ROWS
        assert _chunk_rows(16_384) == SCAN_CHUNK_CELLS // 16_384
        assert _chunk_rows(16_384) * 16_384 <= SCAN_CHUNK_CELLS
        # never zero, however wide the sheet
        assert _chunk_rows(SCAN_CHUNK_CELLS * 2) == 1
