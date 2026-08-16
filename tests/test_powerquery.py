"""Tests for reading Power Query out of a workbook.

The M source is the one thing in a workbook written in a real programming
language, and it is read here without being interpreted. These check the line
between the two: what a query plainly names is found, what an expression would
have to compute is not claimed, and neither case costs the analysis a run.
"""

import io
from pathlib import Path

import pytest
from openpyxl import Workbook

from linexcel import analyze
from linexcel.powerquery import (
    Query,
    QuerySource,
    parse_section,
    read_destinations,
    read_queries,
    read_section,
    scan_sources,
)

FIXTURE = Path("tests/fixtures/power_query.xlsx")


@pytest.fixture(scope="module")
def data() -> bytes:
    return FIXTURE.read_bytes()


def plain_workbook() -> bytes:
    wb = Workbook()
    wb.active["A1"] = 1
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestSectionParsing:
    def test_each_shared_member_is_one_query(self):
        section = "section Section1;\nshared A = 1;\nshared B = 2;"
        assert parse_section(section) == {"A": "1", "B": "2"}

    def test_the_section_header_is_not_a_query(self):
        assert "Section1" not in parse_section("section Section1;\nshared A = 1;")

    def test_a_semicolon_inside_a_string_ends_nothing(self):
        """``Csv.Document`` takes its delimiter as a string, often a ``;``."""
        section = 'shared A = Csv.Document(f, [Delimiter=";"]);\nshared B = 2;'
        parsed = parse_section(section)
        assert set(parsed) == {"A", "B"}
        assert parsed["A"].endswith("])")

    def test_a_semicolon_inside_a_comment_ends_nothing(self):
        section = "shared A = // one; two\n  1;\nshared B = 2;"
        assert set(parse_section(section)) == {"A", "B"}

    def test_a_name_with_spaces_keeps_its_spaces(self):
        assert parse_section('shared #"Raw Sales" = 1;') == {"Raw Sales": "1"}

    def test_metadata_before_a_member_is_not_part_of_the_name(self):
        section = '[Description="x"]\nshared A = 1;'
        assert parse_section(section) == {"A": "1"}

    def test_an_empty_section_yields_nothing(self):
        assert parse_section("") == {}


class TestSourceScanning:
    @pytest.mark.parametrize(
        ("m_source", "expected"),
        [
            (
                'Excel.CurrentWorkbook(){[Name="SalesTable"]}[Content]',
                ("table", "SalesTable", "Excel.CurrentWorkbook"),
            ),
            (
                'Csv.Document(File.Contents("C:\\data\\x.csv"))',
                ("file", "C:\\data\\x.csv", "File.Contents"),
            ),
            (
                'Web.Contents("https://api.example/v1")',
                ("web", "https://api.example/v1", "Web.Contents"),
            ),
            (
                'Sql.Database("srv01", "Finance")',
                ("database", "srv01", "Sql.Database"),
            ),
            (
                'Folder.Files("\\\\share\\reports")',
                ("folder", "\\\\share\\reports", "Folder.Files"),
            ),
        ],
    )
    def test_a_named_source_is_read_off_the_m(self, m_source, expected):
        (source,) = scan_sources(m_source, set())
        assert (source.kind, source.target, source.function) == expected

    def test_a_partner_connector_is_found_by_its_shape(self):
        """No list can enumerate every connector; the naming is the rule."""
        (source,) = scan_sources('Acme.Databases("tenant-7")', set())
        assert (source.kind, source.function) == ("database", "Acme.Databases")

    def test_transform_steps_are_not_mistaken_for_sources(self):
        m_source = (
            'Table.SelectRows(x, each [K] = "y")\n'
            'Text.Replace(a, "b", "c")\n'
            'Table.AddColumn(t, "New", each 1)'
        )
        assert scan_sources(m_source, set()) == []

    def test_the_same_source_named_twice_is_listed_once(self):
        m_source = 'a = File.Contents("p.csv"), b = File.Contents("p.csv")'
        assert len(scan_sources(m_source, set())) == 1

    def test_a_query_reading_another_names_it(self):
        m_source = 'let x = #"Raw Sales" in x'
        (source,) = scan_sources(m_source, {"Raw Sales"})
        assert (source.kind, source.target) == ("query", "Raw Sales")

    def test_a_local_step_shadowing_a_query_name_is_not_a_reference(self):
        """``Source`` here is the step just above, not the query next door."""
        m_source = "let Source = 1, Next = Source + 1 in Next"
        assert scan_sources(m_source, {"Source"}) == []

    def test_a_namespace_is_not_a_query_of_the_same_name(self):
        m_source = "Table.Sort(x, {})"
        assert scan_sources(m_source, {"Table"}) == []

    def test_a_query_name_inside_a_string_is_not_a_reference(self):
        m_source = 'let x = "Raw Sales" in x'
        assert scan_sources(m_source, {"Raw Sales"}) == []


class TestReadingTheFile:
    def test_the_m_source_is_recovered_from_the_mashup(self, data):
        section = read_section(data)
        assert section is not None
        assert "shared BusyProducts" in section

    def test_a_workbook_without_power_query_has_no_mashup(self):
        assert read_section(plain_workbook()) is None
        assert read_queries(plain_workbook()) == []

    def test_the_queries_come_back_with_their_source_and_destination(self, data):
        queries = {q.name: q for q in read_queries(data)}
        assert set(queries) == {"BusyProducts", "TinyConnectionOnly"}
        busy = queries["BusyProducts"]
        assert busy.sources == [
            QuerySource("table", "SalesTable", "Excel.CurrentWorkbook")
        ]
        assert busy.loaded is True
        assert busy.loaded_to[0].sheet == "Loaded"
        assert busy.loaded_to[0].ref == "A1:B3"

    def test_a_connection_only_query_loads_nowhere(self, data):
        (tiny,) = [q for q in read_queries(data) if q.name == "TinyConnectionOnly"]
        assert tiny.loaded is False
        assert tiny.loaded_to == []

    def test_the_connection_is_tied_to_the_range_it_fills(self, data):
        destinations = read_destinations(data)
        assert destinations["busyproducts"][0].ref == "A1:B3"

    def test_a_file_that_is_not_a_workbook_is_not_a_crash(self):
        assert read_queries(b"not a zip at all") == []
        assert read_section(b"") is None


class TestChainedQueries:
    """One query reading another is the shape most real mashups have."""

    def test_the_edge_runs_from_the_upstream_query(self, monkeypatch):
        from linexcel import analyzer

        section = (
            "section Section1;\n"
            'shared Raw = Excel.CurrentWorkbook(){[Name="T"]}[Content];\n'
            "shared Clean = Table.SelectRows(Raw, each true);"
        )
        monkeypatch.setattr(analyzer, "read_queries", _queries_from(section))
        graph = analyzer.analyze_workbook(plain_workbook(), "main.xlsx")["graph"]
        edges = {(e["source"], e["target"]) for e in graph["edges"]}
        assert ("q:Raw", "q:Clean") in edges

    def test_two_queries_differing_only_in_case_stay_apart(self, monkeypatch):
        """M is case-sensitive; folding the name would merge the two nodes."""
        from linexcel import analyzer

        section = "section Section1;\nshared Sales = 1;\nshared sales = 2;"
        monkeypatch.setattr(analyzer, "read_queries", _queries_from(section))
        graph = analyzer.analyze_workbook(plain_workbook(), "main.xlsx")["graph"]
        labels = {n["label"] for n in graph["nodes"] if n["kind"] == "query"}
        assert labels == {"Sales", "sales"}


def _queries_from(section: str):
    """Stand in for the mashup reader with M written here rather than in a file."""
    members = parse_section(section)
    names = set(members)
    queries = [
        Query(name=name, source=body, sources=scan_sources(body, names - {name}))
        for name, body in members.items()
    ]
    return lambda data: queries


class TestOutsideSources:
    def test_what_is_in_the_workbook_is_not_counted_as_outside(self):
        query = Query(
            name="Q",
            source="",
            sources=[
                QuerySource("table", "SalesTable", "Excel.CurrentWorkbook"),
                QuerySource("query", "Raw", "let"),
                QuerySource("file", "C:\\x.csv", "File.Contents"),
            ],
        )
        assert [s.target for s in query.outside_sources()] == ["C:\\x.csv"]


class TestWorkbooksWithoutQueries:
    def test_no_query_node_and_no_warning(self):
        result = analyze(plain_workbook(), filename="plain.xlsx")
        assert result.stats["queries"] == 0
        assert not [n for n in result.nodes if n["kind"] == "query"]
        assert not [w for w in result.warnings if "Power Query" in w]
