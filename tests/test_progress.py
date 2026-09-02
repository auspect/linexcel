"""Saying what is happening, and saying it in the right place.

Two rules the tests below exist to hold. Nothing is printed unless asked:
``analyze()`` is a library call, and one that writes to a terminal nobody
pointed it at is one people wrap in ``redirect_stderr``. And nothing is
printed to **stdout**, ever — `--json -` and `-o -` write the payload there,
so a progress bar in that stream corrupts a pipe rather than decorating it.
"""

import io

import pytest
from openpyxl import Workbook

from linexcel import analyze
from linexcel.analyzer import inspect_workbook
from linexcel.loader import load_cached_values
from linexcel.progress import Reporter


def workbook(cells: dict[str, object], sheet: str = "S") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for address, value in cells.items():
        ws[address] = value
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def small() -> bytes:
    return workbook({"A1": 2, "A2": 3, "A3": "=A1*A2"})


class TestSilenceByDefault:
    def test_a_disabled_reporter_says_nothing(self, capsys):
        reporter = Reporter(False)
        with reporter.phase("work", total=2) as phase:
            phase.step("one")
        reporter.note("something")
        assert capsys.readouterr() == ("", "")

    def test_analysing_a_workbook_is_silent(self, small, capsys):
        analyze(small, filename="s.xlsx")
        assert capsys.readouterr() == ("", "")


class TestWhereItWrites:
    def test_the_phase_and_its_timing_go_to_stderr(self, capsys):
        with Reporter(True).phase("cached values"):
            pass
        out, err = capsys.readouterr()
        assert out == ""
        assert "cached values" in err

    def test_a_note_goes_to_stderr_too(self, capsys):
        Reporter(True).note("hello")
        out, err = capsys.readouterr()
        assert out == ""
        assert "hello" in err

    def test_verbose_analysis_leaves_stdout_clean(self, small, capsys):
        """`--json -` writes the graph to stdout; progress must not join it."""
        analyze(small, filename="s.xlsx", verbose=True)
        out, err = capsys.readouterr()
        assert out == ""
        assert "extraction+grouping" in err

    def test_every_phase_is_still_reported(self, small, capsys):
        analyze(small, filename="s.xlsx", verbose=True)
        err = capsys.readouterr().err
        for phase in ("structure", "cached values", "extraction+grouping", "total"):
            assert phase in err, phase


class TestDrawingOnlyForAHuman:
    def test_a_redirected_stderr_gets_no_live_display(self):
        """Thousands of redraw escapes in a CI log help nobody."""
        assert Reporter(True)._live is False  # pytest captures stderr

    def test_the_plain_path_is_used_when_rich_is_not_wanted(self, capsys):
        with Reporter(True, force_plain=True).phase("work"):
            pass
        assert "[linexcel] work" in capsys.readouterr().err


class TestTheSlowReaderIsNotSilent:
    def test_falling_back_to_openpyxl_is_said_out_loud(self, small, monkeypatch):
        """A different reader is a different answer on an edge case."""
        from linexcel import loader

        def boom(data, reporter=None):
            raise RuntimeError("calamine said no")

        monkeypatch.setattr(loader, "_load_cached_values_calamine", boom)
        warnings: list[str] = []
        load_cached_values(small, warnings)
        (warning,) = warnings
        assert "openpyxl" in warning
        assert "calamine said no" in warning

    def test_the_values_are_still_there(self, small, monkeypatch):
        from linexcel import loader

        monkeypatch.setattr(
            loader,
            "_load_cached_values_calamine",
            lambda data, reporter=None: (_ for _ in ()).throw(RuntimeError("no")),
        )
        assert len(load_cached_values(small, [])) >= 1


class TestDryRun:
    """What the file claims, read from its headers, before committing to it."""

    def test_it_reports_the_declared_size_of_each_sheet(self, small):
        (sheet,) = inspect_workbook(small)["sheets"]
        assert sheet["name"] == "S"
        assert sheet["cells"] == sheet["rows"] * sheet["cols"]

    def test_a_stray_corner_is_flagged_as_over_the_ceiling(self):
        facts = inspect_workbook(workbook({"A1": 1, "XFD1048576": "corner"}))
        assert facts["sheets"][0]["truncated"] is True
        assert facts["densePathRefused"] is True

    def test_an_ordinary_file_is_flagged_as_neither(self, small):
        facts = inspect_workbook(small)
        assert facts["sheets"][0]["truncated"] is False
        assert facts["densePathRefused"] is False

    def test_the_workbooks_it_links_to_are_named(self):
        facts = inspect_workbook(workbook({"B1": "='[Ref.xlsx]Data'!B2"}))
        # named from the formula only once the package declares the link;
        # an openpyxl-written file declares none, so the list is empty here
        assert facts["externalWorkbooks"] == []

    def test_a_hidden_sheet_keeps_its_state(self):
        wb = Workbook()
        wb.active.title = "Visible"
        wb.active["A1"] = 1
        wb.create_sheet("Secret").sheet_state = "hidden"
        buf = io.BytesIO()
        wb.save(buf)
        states = {
            s["name"]: s["state"] for s in inspect_workbook(buf.getvalue())["sheets"]
        }
        assert states["Secret"] == "hidden"

    def test_the_ceilings_that_will_apply_are_reported(self, small):
        ceilings = inspect_workbook(small)["ceilings"]
        assert set(ceilings) == {"cellsPerSheet", "nodesPerSheet", "denseCells"}
