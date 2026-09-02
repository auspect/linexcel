"""Files that used to take the process down, and what happens instead now.

The rest of the suite asks whether the lineage is right. These ask whether an
answer comes back at all, on the workbooks a real user turns out to have: one
touched cell in the far corner, a linked file that is a CSV in disguise, a
folder that does not hold what the formulas name. Someone whose file kills the
tool does not open an issue — they stop using it.

``scripts/stress_scenarios.py`` covers the same ground end to end, in a
subprocess, and reports timings; this is the part that belongs in CI.
"""

import io

import pytest
from openpyxl import Workbook

from linexcel import analyze
from linexcel.external import read_workbook_values
from linexcel.loader import (
    MAX_CELLS_PER_SHEET,
    MAX_DENSE_CELLS,
    declared_cells,
    load_cached_values,
)


def workbook(cells: dict[str, object], sheet: str = "S") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for address, value in cells.items():
        ws[address] = value
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture(scope="module")
def stray_corner() -> bytes:
    """Three cells, and one someone touched once at the far corner.

    Excel never shrinks a used range back on its own, so this is what a
    workbook looks like after a stray click: `A1:XFD1048576`, 17 billion cells
    declared, four of them holding anything.
    """
    return workbook({"A1": 2, "A2": "=A1+1", "XFD1048576": "corner"})


@pytest.fixture(scope="module")
def corner_read(stray_corner):
    """One reading of the corner workbook, shared: it costs seconds.

    Sweeping up to the ceiling is the price of not trusting the declaration,
    and it is paid per call — so the assertions below share one call rather
    than each buying their own.
    """
    warnings: list[str] = []
    return load_cached_values(stray_corner, warnings), warnings


@pytest.fixture(scope="module")
def corner_result(stray_corner):
    """Likewise for the full analysis."""
    return analyze(stray_corner, filename="corner.xlsx")


class TestDeclaredRange:
    """What a sheet *claims* to use, read before anything reads the cells."""

    def test_the_corner_cell_is_seen_without_reading_the_sheet(self, stray_corner):
        assert declared_cells(stray_corner) == 16_384 * 1_048_576

    def test_an_ordinary_sheet_declares_what_it_holds(self):
        assert declared_cells(workbook({"A1": 1, "B2": 2})) == 4

    def test_a_single_cell_range_is_read_too(self):
        """``<dimension ref="A1"/>`` has no colon and still means one cell."""
        assert declared_cells(workbook({"A1": 1})) == 1

    def test_something_that_is_not_a_package_declares_nothing(self):
        assert declared_cells(b"not a zip at all") == 0


class TestTheFastPathIsSkippedRatherThanCrashing:
    """python-calamine builds a dense rows × columns array before returning.

    For a declared `A1:XFD1048576` that is a 512 GiB allocation, and a failed
    allocation in Rust aborts the process — no exception, nothing to catch. So
    the size is checked first and openpyxl, which is lazy, reads the file.
    """

    @pytest.mark.slow
    def test_the_workbook_analyses_instead_of_taking_the_process_with_it(
        self, corner_result
    ):
        assert corner_result.stats["totalFormulas"] == 1

    def test_the_values_are_still_read(self, corner_read):
        cached, _ = corner_read
        assert len(cached) >= 1

    def test_the_report_says_the_used_range_is_the_problem(self, corner_read):
        _, warnings = corner_read
        (warning,) = warnings
        assert "17,179,869,184 cells" in warning

    def test_and_says_what_to_do_about_it(self, corner_read):
        _, warnings = corner_read
        assert "delete" in warnings[0] and "save" in warnings[0]

    def test_an_ordinary_workbook_is_not_slowed_down_by_the_check(self):
        warnings: list[str] = []
        cached = load_cached_values(workbook({"A1": 1, "A2": "=A1"}), warnings)
        assert warnings == []
        assert len(cached) >= 1

    @pytest.mark.slow
    def test_the_warning_reaches_the_result(self, corner_result):
        assert any("used range" in w for w in corner_result.warnings)


class TestALinkedWorkbookThatCannotBeRead:
    """The same hazard, one file away — and there, refusing is the answer."""

    def test_a_declared_range_too_large_is_refused_by_name(
        self, tmp_path, stray_corner
    ):
        path = tmp_path / "Ref.xlsx"
        path.write_bytes(stray_corner)
        with pytest.raises(ValueError, match="used range"):
            read_workbook_values(path)

    def test_the_refusal_becomes_a_warning_naming_the_file(
        self, tmp_path, stray_corner
    ):
        (tmp_path / "Ref.xlsx").write_bytes(stray_corner)
        graph = analyze(
            workbook({"B1": "='[Ref.xlsx]Data'!B2"}),
            filename="main.xlsx",
            refs_dir=tmp_path,
        )
        (warning,) = [w for w in graph.warnings if "could not be read" in w]
        assert "Ref.xlsx" in warning
        assert "used range" in warning

    def test_a_workbook_of_ordinary_size_is_read(self, tmp_path):
        (tmp_path / "Ref.xlsx").write_bytes(workbook({"B2": 21}, sheet="Data"))
        assert read_workbook_values(tmp_path / "Ref.xlsx")[("Data", 2, 2)] == 21

    @pytest.mark.parametrize(
        ("name", "content"),
        [
            ("truncated", workbook({"A1": 1})[:200]),
            ("a CSV in disguise", b"Region,Amount\nNorth,12\n"),
            ("a real .xls", b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 512),
            ("empty", b""),
        ],
    )
    def test_an_unreadable_linked_file_costs_a_warning_not_the_run(
        self, tmp_path, name, content
    ):
        (tmp_path / "Ref.xlsx").write_bytes(content)
        result = analyze(
            workbook({"A1": 2, "B1": "='[Ref.xlsx]Data'!B2 * A1"}),
            filename="main.xlsx",
            refs_dir=tmp_path,
        )
        assert result.stats["externalWorkbooksRead"] == 0
        assert any("Ref.xlsx" in w for w in result.warnings), name


class TestTheTwoCeilings:
    """They bound different things, and conflating them misreports honest files."""

    def test_the_sweep_budget_is_past_any_real_sheet_and_still_bounded(self):
        """A full-height sheet 60 columns wide fits; the corner case does not."""
        assert 1_048_576 * 60 < MAX_CELLS_PER_SHEET < 16_384 * 1_048_576

    def test_the_dense_budget_is_a_memory_bound_so_it_sits_lower(self):
        """At ~32 bytes a cell, this is what calamine may try to allocate."""
        assert MAX_DENSE_CELLS < MAX_CELLS_PER_SHEET
        assert MAX_DENSE_CELLS * 32 < 1_000_000_000

    def test_an_honest_sheet_over_the_sweep_budget_is_not_called_a_liar(
        self, monkeypatch
    ):
        """Shrinking the sweep budget must not make the fast path complain."""
        from linexcel import analyzer

        monkeypatch.setattr(analyzer, "MAX_CELLS_PER_SHEET", 30)
        warnings: list[str] = []
        load_cached_values(workbook({"A1": 1, "J40": 2}), warnings)
        assert warnings == []


class TestTheDecompositionIsBoundedInTime:
    """A count of evaluations cannot bound a run. Only a clock can.

    `MAX_SCRATCH_EVALS` was meant to stop the step decomposition running away,
    and `preload_steps` — added later to batch the calls — never took from it.
    Worse, the count was the wrong unit: each evaluation asks the engine to
    walk the dirty dependency graph, so on a workbook of running totals one
    call costs O(graph) and four thousand of them outlast anybody's patience.
    Someone waited three hours on an estimate of four minutes.
    """

    @staticmethod
    def running_totals(rows: int = 300) -> bytes:
        """Every cell sums everything above it, and one reference is dead."""
        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        ws["Z1"] = "=Ghost!A1"
        for r in range(1, rows + 1):
            ws.cell(row=r, column=1, value=f"=SUM(A1:A{max(1, r - 1)}) + Z1")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_a_ceiling_of_zero_still_produces_a_report(self):
        rows_plus_the_dead_one = 301
        result = analyze(self.running_totals(), filename="x.xlsx", step_seconds=0)
        assert result.stats["totalFormulas"] == rows_plus_the_dead_one
        assert len(result.nodes) > 1

    def test_the_breakdown_is_what_gets_dropped(self):
        """Not the analysis: the graph, the edges and the values are all there.

        On a workbook the engine *can* compute, so that the difference the
        ceiling makes is visible rather than hidden behind a dead reference.
        """
        book = workbook({"A1": 2, "A2": 3, "A3": "=(A1+A2)*2"})
        cut = analyze(book, filename="x.xlsx", step_seconds=0)
        whole = analyze(book, filename="x.xlsx")

        assert len(cut.nodes) == len(whole.nodes)
        assert len(cut.edges) == len(whole.edges)
        # the value of the cell survives; only its decomposition does not
        assert self.value_of(cut, "A3") == self.value_of(whole, "A3") == 10
        cut_root = next(n["steps"] for n in cut.nodes if n.get("addr") == "A3")
        # the root step keeps its value — it is the cell's own, and needs no
        # scratch pass — while the sub-steps below it go
        assert self.evaluated_steps(cut) < self.evaluated_steps(whole)
        assert cut_root["evaluated"] is True

    @staticmethod
    def value_of(result, addr):
        return next(n["value"] for n in result.nodes if n.get("addr") == addr)

    @staticmethod
    def evaluated_steps(result) -> int:
        def walk(step):
            if not step:
                return 0
            return bool(step.get("evaluated")) + sum(
                walk(c) for c in step.get("children", [])
            )

        return sum(walk(n.get("steps")) for n in result.nodes)

    def test_the_report_says_it_stopped_and_why(self):
        result = analyze(self.running_totals(), filename="x.xlsx", step_seconds=0)
        (warning,) = [w for w in result.warnings if "decomposition stopped" in w]
        assert "--time-budget" in warning

    def test_the_batch_path_takes_from_the_budget(self):
        """The call it slipped past for two releases."""
        from linexcel.resolver import _Budget

        budget = _Budget(10)
        assert budget.take(4) is True
        assert budget.left == 6
        assert budget.take(20) is False
        assert budget.spent == "calls"

    def test_a_batch_too_big_is_refused_without_poisoning_the_rest(self):
        """One node asking for more than is left must not starve the next."""
        from linexcel.resolver import _Budget

        budget = _Budget(1)
        assert budget.take(5) is False
        assert budget.take(1) is True

    def test_a_deadline_stops_the_decomposition_and_not_the_values(self):
        """The distinction the zero-deadline test was written to hold."""
        from linexcel.resolver import _Budget

        budget = _Budget(10_000, seconds=0)
        assert budget.expired is True
        assert budget.spent == "time"
        # value recovery goes on: it is the answer, not the detail
        assert budget.take() is True

    def test_a_budget_with_room_says_nothing(self):
        from linexcel.resolver import _Budget

        budget = _Budget(10, seconds=60)
        assert budget.take(3) is True
        assert budget.warning() is None
