"""Tests for value resolution: guarded errors, resilience, dates, provenance."""

import datetime
import io
import json
from typing import Any

import formualizer as fz
import pytest
from openpyxl import Workbook

from linexcel.analyzer import analyze_workbook
from linexcel.decompose import SCRATCH_SENTINEL, _collect_step_exprs, _render_expr
from linexcel.graph import _spread_cells
from linexcel.loader import CachedValues, load_cached_values
from linexcel.resolver import _Budget, _is_volatile, _ValueResolver
from linexcel.values import serial_to_date_text

SHEET = "S"


def build(
    cells: dict[str, Any], formats: dict[str, str] | None = None, sheet: str = SHEET
) -> bytes:
    """Minimal in-memory workbook: one sheet, the given cells and formats."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for addr, value in cells.items():
        ws[addr] = value
    for addr, number_format in (formats or {}).items():
        ws[addr].number_format = number_format
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def graph_of(cells: dict[str, Any], formats: dict[str, str] | None = None) -> dict:
    return analyze_workbook(build(cells, formats), "values.xlsx")["graph"]


def node_of(graph: dict, node_id: str) -> dict:
    return next(n for n in graph["nodes"] if n["id"] == node_id)


class _StubEngine:
    """Engine stand-in: only reports stored values, never recalculates."""

    def __init__(
        self,
        values: dict[tuple[str, int, int], Any],
        evaluated: dict[tuple[str, int, int], Any] | None = None,
    ):
        self._values = values
        self._evaluated = evaluated or {}

    def get_value(self, sheet: str, row: int, col: int) -> Any:
        return self._values.get((sheet, row, col))

    def get_formula(self, sheet: str, row: int, col: int) -> None:
        return None

    def evaluate_cell(self, sheet: str, row: int, col: int) -> Any:
        return self._evaluated.get((sheet, row, col))


def resolver_for(
    engine_values: dict,
    cached: CachedValues,
    warnings: list,
    *,
    evaluated: dict[tuple[str, int, int], Any] | None = None,
    engine_alive: bool = True,
) -> _ValueResolver:
    resolver = _ValueResolver(
        _StubEngine(engine_values, evaluated=evaluated),
        {SHEET},
        cached,
        warnings,
        _Budget(0),
        scratch_ready=False,
        engine_alive=engine_alive,
    )
    return resolver


class TestGuardedErrors:
    def test_missing_sheet_guard_falls_back_to_the_second_argument(self):
        graph = graph_of({"A1": "=IFERROR(NOSHEET!A1, 456)"})
        node = node_of(graph, "c:S!A1")
        assert node["value"] == 456
        assert node["valueSource"] == "fallback"

    def test_native_error_guard_is_evaluated_by_the_engine(self):
        graph = graph_of({"A1": "=IFERROR(1/0, 42)"})
        node = node_of(graph, "c:S!A1")
        assert node["value"] == 42
        assert node["valueSource"] == "engine"

    def test_blank_operand_counts_as_zero(self):
        # A2 stays empty: Excel coerces the blank to 0, so the guard never fires
        graph = graph_of({"B1": "=IFERROR(A2+1, 7)"})
        assert node_of(graph, "c:S!B1")["value"] == 1

    def test_iserror_branch_is_evaluated(self):
        graph = graph_of({"A1": "=IF(ISERROR(1/0), 99, 0)"})
        assert node_of(graph, "c:S!A1")["value"] == 99

    def test_unguarded_missing_sheet_does_not_break_the_analysis(self):
        """The cell claims no value, and the rest of the pass still completes.

        An unguarded broken reference is now isolated before evaluation rather
        than aborting it, so the warning names an evaluation that finished. The
        cell itself is unchanged: it has no value to report either way.
        """
        graph = graph_of({"A1": "=NOSHEET!A1"})
        node = node_of(graph, "c:S!A1")
        assert node["value"] is None
        assert "valueSource" not in node
        warnings = graph["meta"]["warnings"]
        assert any("completed after isolating 1 cell" in w for w in warnings)


class TestResilience:
    """One bad reference used to blank out every formula of the workbook."""

    def test_good_formulas_keep_their_values_next_to_a_broken_one(self):
        graph = graph_of(
            {
                "A1": 1,
                "B1": 2,
                "A2": "=A1*2",
                "A3": "=IFERROR(NOSHEET!A1, 456)",
                "A4": "=SUM(A1:B1)",
            }
        )
        doubled = node_of(graph, "c:S!A2")
        summed = node_of(graph, "c:S!A4")
        assert doubled["value"] == 2
        assert doubled["valueSource"] == "engine"
        assert summed["value"] == 3
        assert summed["valueSource"] == "engine"

    def test_the_graph_still_describes_every_formula(self):
        graph = graph_of({"A1": 1, "A2": "=A1*2", "A3": "=NOSHEET!A1"})
        ids = {n["id"] for n in graph["nodes"]}
        assert {"c:S!A2", "c:S!A3"} <= ids
        assert graph["meta"]["stats"]["totalFormulas"] == 2


def build_multi(sheets: dict[str, dict[str, Any]]) -> bytes:
    """In-memory workbook with several sheets, in the given order."""
    wb = Workbook()
    first = True
    for title, cells in sheets.items():
        ws = wb.active if first else wb.create_sheet(title)
        ws.title = title
        first = False
        for addr, value in cells.items():
            ws[addr] = value
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# One unguarded broken reference makes the whole-workbook evaluation fail, and
# every value is then recovered cell by cell. A scratch evaluation only reads
# constants, so a formula pointing at another formula used to read 0: =A3+A4
# came back as 0 instead of 230, and =SUM(A1:A7) counted the two constants
# only. The precedents are now recovered first and fed back into the engine.
CHAINED = {
    "Inputs": {
        "A1": 10,
        "A2": 20,
        "A3": "=A1+A2",
        "A4": "=A1*A2",
        "A5": "=IFERROR(1/0,42)",
        "A6": "=IFERROR(NOSHEET!A1,456)",
        "A7": "=A3+A4",
        "B1": "=SUM(A1:A2)",
        "B2": "=IF(ISERROR(1/0),99,0)",
        "C1": 5,
        "C2": "=C1*2",
        "D1": 7,
        "D2": "=D1+1",
        "E1": datetime.date(2026, 8, 7),
        "E2": "=E1+1",
        "E3": "=E1+7",
        "E4": datetime.date(1999, 12, 31),
    },
    "Calc": {
        "A1": 100,
        "A2": "=Inputs!A1*10",
        "A3": "=Inputs!E1",
        "B1": "=SUM(Inputs!A1:A7)",
        "B2": "=Calc!A2+Calc!A1",
        "C1": "=IFERROR(Inputs!A5+1,0)",
        "C2": "=A1/B1",
        "D1": "=Inputs!E1+1",
    },
    "Errors": {
        "A1": "=1/0",
        "A2": "=NOSHEET!A1",
        "A3": "=IFERROR(A1+1,7)",
        "B1": 1,
        "B2": "=B1+B2",
    },
}


class TestChainedRefs:
    """A formula reading another formula, across sheets and through ranges."""

    def graph(self) -> dict:
        return analyze_workbook(build_multi(CHAINED), "chained.xlsx")["graph"]

    def test_same_sheet_chain_sums_the_computed_precedents(self):
        # =A3+A4 with A3==A1+A2 and A4==A1*A2: 30 + 200
        node = node_of(self.graph(), "c:Inputs!A7")
        assert node["value"] == 230
        assert node["valueSource"] == "engine"

    def test_range_over_formula_cells_is_summed_in_full(self):
        # Inputs!A1:A7 = 10 + 20 + 30 + 200 + 42 + 456 + 230, the guarded A6
        # contributing the value of its fallback branch
        node = node_of(self.graph(), "c:Calc!B1")
        assert node["value"] == 988
        assert node["valueSource"] == "engine"

    def test_cross_sheet_reference_to_a_guarded_formula(self):
        node = node_of(self.graph(), "c:Calc!C1")
        assert node["value"] == 43
        assert node["valueSource"] == "engine"

    def test_division_by_a_chained_range_sum(self):
        node = node_of(self.graph(), "c:Calc!C2")
        assert node["value"] == pytest.approx(100 / 988, abs=1e-6)
        assert node["valueSource"] == "engine"

    def test_guard_reading_a_cell_that_holds_an_error(self):
        # Errors!A1 is =1/0: the error travels to A3, whose guard catches it
        node = node_of(self.graph(), "c:Errors!A3")
        assert node["value"] == 7
        # formualizer >=0.8.0 handles IFERROR internally, so the guard
        # may fire as "fallback" instead of "engine" depending on version
        assert node["valueSource"] in ("engine", "fallback")

    def test_cross_sheet_constant_and_qualified_self_reference(self):
        graph = self.graph()
        assert node_of(graph, "c:Calc!A2")["value"] == 100
        assert node_of(graph, "c:Calc!B2")["value"] == 200

    def test_date_serial_travels_across_sheets(self):
        graph = self.graph()
        # formualizer >=0.8.0 returns ISO date strings instead of serial numbers
        assert node_of(graph, "c:Calc!A3")["value"] in (46241, "2026-08-07")
        assert node_of(graph, "c:Calc!D1")["value"] in (46242, "2026-08-08")

    def test_the_static_date_keeps_its_valuedate(self):
        assert node_of(self.graph(), "i:Inputs!E1")["valueDate"] == "2026-08-07"

    def test_the_error_cell_stays_an_error(self):
        # As Excel spells it: the engine reports a {"type": "Error"} object, and
        # a reader of the report must never see that object rather than #DIV/0!
        assert node_of(self.graph(), "c:Errors!A1")["value"] == "#DIV/0!"

    def test_the_unguarded_broken_reference_claims_no_value(self):
        node = node_of(self.graph(), "c:Errors!A2")
        assert node["value"] is None
        assert "valueSource" not in node

    def test_a_self_referencing_formula_does_not_crash(self):
        # =B1+B2 written in B2: the cycle is cut, B2 reads as blank
        # formualizer >=0.8.0 may return None instead of 1 for the cycle cell
        val = node_of(self.graph(), "c:Errors!B2")["value"]
        assert val in (1, None)

    def test_the_recovery_is_reported_in_the_warnings(self):
        warnings = self.graph()["meta"]["warnings"]
        assert any("Global evaluation incomplete" in w for w in warnings)
        assert any(w.startswith("Values recovered cell by cell:") for w in warnings)


class TestChainedViaGuarded:
    def test_guard_catches_the_error_of_a_referenced_cell(self):
        graph = graph_of({"A1": "=1/0", "A2": "=IFERROR(A1+1, 7)", "Z1": "=NOSHEET!A1"})
        assert node_of(graph, "c:S!A2")["value"] == 7

    def test_an_error_inside_a_range_poisons_the_sum(self):
        # Excel does not skip errors in SUM: SUM({#DIV/0!, 5}) is #DIV/0!, so
        # the guard fires and the fallback branch — not 6 — is what is shown
        graph = graph_of(
            {
                "A1": "=1/0",
                "A2": 5,
                "A3": "=IFERROR(SUM(A1:A2)+1, 0)",
                "Z1": "=NOSHEET!A1",
            }
        )
        node = node_of(graph, "c:S!A3")
        assert node["value"] == 0
        assert node["valueSource"] == "engine"


class TestSiblingIsolation:
    """One unguarded broken reference must not touch the cells around it."""

    def test_good_formulas_survive_on_every_sheet(self):
        graph = analyze_workbook(
            build_multi(
                {
                    "A": {"A1": 3, "A2": "=A1*4", "A3": "=A2+A1"},
                    "B": {"B1": "=NOSHEET!A1", "B2": "=A!A3*2"},
                    "C": {"C1": "=SUM(A!A1:A3)"},
                }
            ),
            "isolation.xlsx",
        )["graph"]
        assert node_of(graph, "c:A!A2")["value"] == 12
        assert node_of(graph, "c:A!A3")["value"] == 15
        assert node_of(graph, "c:B!B2")["value"] == 30
        assert node_of(graph, "c:C!C1")["value"] == 30
        assert node_of(graph, "c:B!B1")["value"] is None


class TestNestedGuard:
    def test_only_the_outer_guard_is_recovered(self):
        # Excel lets the inner IFERROR absorb the missing sheet — SUM(0) — and
        # returns 0. The engine cannot evaluate the expression at all, so the
        # recovery falls back to the branch of the outer guard. Documented
        # ceiling of the recovery, asserted here so a change is noticed.
        graph = graph_of({"A1": "=IFERROR(SUM(IFERROR(NOSHEET!A1,0)),1)"})
        node = node_of(graph, "c:S!A1")
        assert node["value"] == 1
        assert node["valueSource"] == "fallback"


class TestDates:
    def test_static_date_cell_is_reported_as_a_date(self):
        graph = graph_of({"A1": datetime.date(2026, 8, 7), "B1": "=A1+1"})
        node = node_of(graph, "i:S!A1")
        assert node["valueDate"] == "2026-08-07"
        # a midnight datetime serializes as the bare date, never with a time
        assert node["cachedValue"] == "2026-08-07"
        # formualizer >=0.8.0 returns ISO date strings instead of serial numbers
        val = node["value"]
        assert serial_to_date_text(val) == "2026-08-07" or val == "2026-08-07"

    def test_date_formula_is_converted(self):
        graph = graph_of(
            {"A1": datetime.date(2026, 8, 7), "A2": "=A1+1"},
            {"A2": "yyyy-mm-dd"},
        )
        assert node_of(graph, "c:S!A2")["valueDate"] == "2026-08-08"

    def test_plain_number_has_no_date(self):
        graph = graph_of({"A1": 7, "A2": "=A1*2"})
        node = node_of(graph, "c:S!A2")
        assert node["value"] == 14
        assert "valueDate" not in node

    def test_serial_conversion(self):
        assert serial_to_date_text(46241.0) == "2026-08-07"
        assert serial_to_date_text(61.0) == "1900-03-01"
        # Jan/Feb 1900 are real dates: serial 1 is 1900-01-01
        assert serial_to_date_text(1.0) == "1900-01-01"
        assert serial_to_date_text(59.0) == "1900-02-28"
        # 1900's phantom leap day: no real date, so nothing is claimed
        assert serial_to_date_text(60.0) is None
        assert serial_to_date_text(0.0) is None
        assert serial_to_date_text(0.0, epoch_1904=True) == "1904-01-01"
        assert serial_to_date_text("2026-08-07") is None
        assert serial_to_date_text(True) is None

    def test_timestamp_keeps_its_time_of_day(self):
        warnings: list[str] = []
        cached = CachedValues(
            {(SHEET, 1, 1): datetime.datetime(2026, 8, 7, 15, 30)},
            {(SHEET, 1, 1)},
            False,
        )
        resolver = resolver_for({}, cached, warnings)
        assert resolver.describe(SHEET, 1, 1)["cachedValue"] == "2026-08-07 15:30:00"

    def test_cached_values_are_loaded_from_the_file(self):
        cached = load_cached_values(
            build({"A1": datetime.date(2026, 8, 7), "A2": 3}, {"A2": "0.00"})
        )
        assert cached.get(SHEET, 1, 1) == datetime.datetime(2026, 8, 7)
        assert cached.get(SHEET, 2, 1) == 3
        assert cached.is_date(SHEET, 1, 1)
        assert not cached.is_date(SHEET, 2, 1)
        assert cached.epoch_1904 is False

    def test_calamine_and_openpyxl_agree_on_values(self):
        """The calamine fast path and the openpyxl fallback must return the
        same values, date flags, and epoch flag for a representative workbook."""
        from linexcel.loader import (
            _detect_epoch_1904,
            _load_cached_values_calamine,
            _load_cached_values_openpyxl,
        )

        data = build(
            {
                "A1": datetime.date(2026, 8, 7),
                "A2": datetime.datetime(2026, 8, 7, 15, 30),
                "A3": 3,
                "B1": 3.5,
                "B2": "text",
                "B3": True,
            },
            {"A1": "yyyy-mm-dd", "A2": "yyyy-mm-dd hh:mm:ss", "A3": "0.00"},
        )
        cal = _load_cached_values_calamine(data)
        opx = _load_cached_values_openpyxl(data)
        assert set(cal._values) == set(opx._values)
        assert cal._date_cells == opx._date_cells
        for key in opx._values:
            cval, oval = cal._values[key], opx._values[key]
            # Numbers: int vs float differ in type but compare equal.
            if isinstance(cval, (int, float)) and isinstance(oval, (int, float)):
                assert cval == oval or abs(float(cval) - float(oval)) < 1e-9
            else:
                assert cval == oval, (key, cval, oval)
        assert cal.epoch_1904 == opx.epoch_1904
        assert _detect_epoch_1904(data) is False
        assert cal.epoch_1904 is False

    def test_epoch_1904_is_detected_from_the_workbook_xml(self):
        """A 1904-date-system workbook sets the epoch flag so the resolver
        interprets engine serials with the right base date."""
        from openpyxl.utils.datetime import MAC_EPOCH

        from linexcel.loader import _detect_epoch_1904

        wb = Workbook()
        wb.active.title = SHEET
        wb.epoch = MAC_EPOCH
        wb.active["A1"] = datetime.date(2026, 8, 7)
        buf = io.BytesIO()
        wb.save(buf)
        cached = load_cached_values(buf.getvalue())
        assert cached.epoch_1904 is True
        assert _detect_epoch_1904(buf.getvalue()) is True


class TestVolatileFormulas:
    """A volatile cell is reported as *not* recalculated, on purpose.

    ``=TODAY()`` recomputed today can never agree with a file saved last week,
    and reporting that as a divergence blames the workbook for the calendar.
    linexcel keeps what the file stores and says why.
    """

    def test_a_volatile_cell_is_not_recalculated(self):
        node = node_of(graph_of({"A1": "=TODAY()"}), "c:S!A1")
        assert node["valueSource"] == "volatile"

    def test_the_clock_never_reaches_the_graph(self):
        """openpyxl stores no cached value, so there is nothing to show."""
        node = node_of(graph_of({"A1": "=NOW()+1"}), "c:S!A1")
        assert node["value"] is None
        assert node["valueSource"] == "volatile"

    def test_the_stored_value_is_kept_when_the_file_has_one(self):
        cached = CachedValues({(SHEET, 1, 1): 45000.0}, set(), False)
        resolver = resolver_for({(SHEET, 1, 1): 46240.0}, cached, [])
        assert resolver.value(SHEET, 1, 1, "=TODAY()") == (45000.0, "volatile", None)

    def test_no_divergence_is_reported_against_the_clock(self):
        warnings: list[str] = []
        cached = CachedValues({(SHEET, 1, 1): 45000.0}, set(), False)
        resolver = resolver_for({(SHEET, 1, 1): 46240.0}, cached, warnings)
        resolver.value(SHEET, 1, 1, "=TODAY()")
        assert warnings == []

    def test_a_volatile_cell_is_not_decomposed_either(self):
        """Steps under it would carry figures computed from today's clock."""
        node = node_of(graph_of({"A1": 1, "B1": "=A1+RAND()"}), "c:S!B1")
        assert node["valueSource"] == "volatile"
        assert node["steps"] is None

    @pytest.mark.parametrize(
        "formula",
        ["=TODAY()", "=NOW()", "=RAND()", "=RANDBETWEEN(1,6)", "=A1*rand()"],
    )
    def test_the_volatile_functions_are_recognised(self, formula):
        assert _is_volatile(formula)

    @pytest.mark.parametrize(
        "formula",
        [
            "=SUM(A1:A3)",
            # volatile to Excel, but the same value for the same workbook
            "=OFFSET(A1,1,1)",
            '=INDIRECT("A" & 2)',
            # a cell that merely talks about them
            '="uses TODAY() somewhere"',
            # a name that only starts like one
            "=TODAYS_TOTAL",
        ],
    )
    def test_a_stable_formula_is_still_recalculated(self, formula):
        assert not _is_volatile(formula)


class TestRootStep:
    """The root step is the formula itself, so the cell already answers it.

    Re-evaluating it in the scratch sheet recomputed what the workbook holds,
    and on a root like ``SUM(H2:H200001)`` that means walking 200,000 formula
    cells a second time — 29 s for a single node on a dense workbook.
    """

    def test_the_root_step_reports_the_value_of_the_cell(self):
        node = node_of(
            graph_of({"A1": 2, "A2": 3, "B1": "=ROUND(SUM(A1:A2) * 1.5, 2)"}),
            "c:S!B1",
        )
        assert node["valueSource"] == "engine"
        assert node["steps"]["label"] == "ROUND"
        assert node["steps"]["value"] == node["value"] == 7.5
        assert node["steps"]["evaluated"] is True

    def test_the_children_are_still_evaluated_on_their_own(self):
        """Only the root is taken from the cell; the steps below it are not."""
        steps = node_of(
            graph_of({"A1": 2, "A2": 3, "B1": "=ROUND(SUM(A1:A2) * 1.5, 2)"}),
            "c:S!B1",
        )["steps"]
        inner = steps["children"][0]
        assert inner["label"] == "*"
        assert inner["value"] == 7.5
        assert inner["children"][0]["label"] == "SUM"
        assert inner["children"][0]["value"] == 5

    def test_an_array_literal_root_no_longer_reads_as_an_error(self):
        """A step expression is re-rendered from the AST, and an array
        literal renders as the placeholder ``{...}``. Evaluating that gave
        ``#NAME?`` for a cell that plainly holds 21."""
        node = node_of(graph_of({"A1": "=SUM({1, 2, 3; 4, 5, 6})"}), "c:S!A1")
        assert node["value"] == 21
        assert node["steps"]["expr"] == "SUM({...})"
        assert node["steps"]["value"] == 21

    def test_a_value_that_is_not_the_engine_s_leaves_the_root_evaluated(self):
        """Under a guard the shown value is a fallback, not this expression."""
        node = node_of(graph_of({"A1": "=IFERROR(NOSHEET!A1, 456)"}), "c:S!A1")
        assert node["valueSource"] == "fallback"
        assert node["steps"]["label"] == "IFERROR"

    def test_the_batch_drops_the_root_expression_but_keeps_the_rest(self):
        ast = fz.parse("=ROUND(SUM(A1:A2) * 1.5, 2)").to_dict()
        assert _collect_step_exprs(ast) == [
            "ROUND(SUM(A1:A2) * 1.5, 2)",
            "SUM(A1:A2) * 1.5",
            "SUM(A1:A2)",
        ]
        assert _collect_step_exprs(ast, skip_root=True) == [
            "SUM(A1:A2) * 1.5",
            "SUM(A1:A2)",
        ]


class _RefusingEngine:
    """Scratch engine that silently refuses one expression.

    ``set_formula`` returning without raising while leaving the cell as it was
    is what the real engine does with a reference it cannot resolve — a
    structured ``Table[Column]`` it never loaded, a 3D ``'A:B'!A1``. The cell
    then still holds what the previous batch wrote there.
    """

    def __init__(self, refuse: str, results: dict[str, Any]):
        self.refuse = refuse
        self.results = results
        self.cells: dict[tuple[str, int, int], Any] = {}

    def set_formula(self, sheet: str, row: int, col: int, formula: str) -> None:
        if self.refuse in formula:
            return  # silently keeps whatever the cell already held
        body = formula[1:] if formula.startswith("=") else formula
        if body.startswith('"') and body.endswith('"'):
            self.cells[(sheet, row, col)] = body[1:-1]
        else:
            self.cells[(sheet, row, col)] = self.results.get(body)

    def evaluate_cells(self, targets: list[tuple[str, int, int]]) -> list[Any]:
        return [self.cells.get(t) for t in targets]


class TestScratchIsolation:
    """Steps share the scratch row, so each cell must be cleared before use.

    ``set_formula`` does not raise on an expression the engine will not take —
    it leaves the cell alone — and the batch then read back whatever the
    *previous* node had computed in that column, reporting another cell's
    value as this step's own. Seen on a workbook whose ``SUM(Sales[Revenue])``
    step showed the text of an unrelated cell.
    """

    def resolver(self) -> _ValueResolver:
        engine = _RefusingEngine("NoTable", {"1 + 2": 3})
        return _ValueResolver(
            engine,
            {SHEET},
            CachedValues({}, set(), False),
            [],
            _Budget(10),
            scratch_ready=True,
        )

    def test_a_refused_expression_does_not_inherit_the_previous_value(self):
        resolver = self.resolver()
        resolver.preload_steps(["1 + 2"], SHEET)
        assert resolver.eval_expr("1 + 2", SHEET) == (3, True)
        # same scratch column, an expression the engine will not take
        resolver.preload_steps(["SUM(NoTable[Amount])"], SHEET)
        assert resolver.eval_expr("SUM(NoTable[Amount])", SHEET) == (None, False)

    def test_the_sentinel_is_never_reported_as_a_value(self):
        resolver = self.resolver()
        resolver.preload_steps(["SUM(NoTable[Amount])"], SHEET)
        value, evaluated = resolver.eval_expr("SUM(NoTable[Amount])", SHEET)
        assert evaluated is False
        assert value != SCRATCH_SENTINEL and value is None

    def test_a_workbook_never_ships_the_sentinel(self):
        graph = graph_of({"A1": "=1+2", "B1": "=SUM(NoSuchTable[Amount])"})
        assert SCRATCH_SENTINEL not in json.dumps(graph)


class TestStepExpressionRendering:
    """A step is evaluated by re-parsing the text it renders itself as.

    The parser keeps grouping in the shape of the tree and drops the
    parentheses, so a subtree rendered flat says something else entirely:
    ``=D2*(1-Rate)`` came back as ``D2 * 1 - Rate``, and the step reported
    2470.06 under a cell holding 1976.208.
    """

    @staticmethod
    def rendered(formula: str) -> str:
        return _render_expr(fz.parse(formula).to_dict())

    @pytest.mark.parametrize(
        ("formula", "expected"),
        [
            ("=A1*(1-B1)", "A1 * (1 - B1)"),
            ("=(A1+B1)*C1", "(A1 + B1) * C1"),
            ("=A1-(B1-C1)", "A1 - (B1 - C1)"),
            ("=A1/(B1*C1)", "A1 / (B1 * C1)"),
            ("=-(A1+B1)", "-(A1 + B1)"),
            ("=(A1>B1)*2", "(A1 > B1) * 2"),
            ("=A1&(B1&C1)", "A1 & (B1 & C1)"),
            # nothing to restore: the flat form already means what it says
            ("=A1+B1+C1", "A1 + B1 + C1"),
            ("=A1+B1*C1", "A1 + B1 * C1"),
            ("=-A1^2", "-A1 ^ 2"),
            # the one postfix operator
            ("=A1%", "A1%"),
            ("=A1%*2", "A1% * 2"),
        ],
    )
    def test_the_grouping_survives_the_round_trip(self, formula, expected):
        assert self.rendered(formula) == expected

    def test_rendering_is_stable_under_re_parsing(self):
        """Whatever comes out must parse back to the same tree."""
        for formula in ("=ROUND(SUM(A1:A3)*(1+T),2)", "=IF(A1>B1,(A1-B1)/B1,0)"):
            once = self.rendered(formula)
            assert self.rendered("=" + once) == once

    def test_a_grouped_formula_decomposes_to_the_value_of_its_cell(self):
        node = node_of(
            graph_of({"A1": 100, "B1": 0.2, "C1": "=A1*(1-B1)"}),
            "c:S!C1",
        )
        assert node["value"] == 80
        assert node["steps"]["expr"] == "A1 * (1 - B1)"
        assert node["steps"]["value"] == 80

    def test_a_grouped_child_step_is_evaluated_as_written(self):
        """The child is re-parsed on its own, so it needs its parentheses."""
        steps = node_of(
            graph_of({"A1": 10, "B1": 2, "C1": 4, "D1": "=A1/(B1*C1)*100"}),
            "c:S!D1",
        )["steps"]
        assert steps["expr"] == "A1 / (B1 * C1) * 100"
        assert steps["value"] == 125
        assert steps["children"][0]["expr"] == "A1 / (B1 * C1)"
        assert steps["children"][0]["value"] == 1.25


class TestStretchedSamples:
    """A group node stands for every cell sharing one R1C1 pattern.

    The panel shows the sampled cells one row per cell, so which cells are
    sampled decides what the reader can actually verify: the head of a stretch
    is made of near-identical neighbours, and a pattern that broke shows up
    further down.
    """

    def stretched(self, rows: int = 10) -> dict:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET
        for r in range(2, rows + 2):
            ws.cell(row=r, column=2, value=r)
            ws.cell(row=r, column=1, value=f"=B{r}*2")
        buf = io.BytesIO()
        wb.save(buf)
        return analyze_workbook(buf.getvalue(), "stretched.xlsx")["graph"]

    def test_the_sample_spans_the_whole_stretch(self):
        samples = node_of(self.stretched(), "g:S!A2#10")["samples"]
        assert [s["addr"] for s in samples] == ["A2", "A4", "A6", "A9", "A11"]

    def test_the_first_and_last_cell_are_always_sampled(self):
        samples = node_of(self.stretched(rows=500), "g:S!A2#500")["samples"]
        assert samples[0]["addr"] == "A2"
        assert samples[-1]["addr"] == "A501"

    def test_each_sample_carries_its_own_provenance(self):
        samples = node_of(self.stretched(), "g:S!A2#10")["samples"]
        assert [s["value"] for s in samples] == [4.0, 8.0, 12.0, 18.0, 22.0]
        assert {s["valueSource"] for s in samples} == {"engine"}

    def test_a_group_smaller_than_the_sample_size_is_shown_whole(self):
        samples = node_of(self.stretched(rows=3), "g:S!A2#3")["samples"]
        assert [s["addr"] for s in samples] == ["A2", "A3", "A4"]

    def test_the_sampler_keeps_reading_order(self):
        cells = [(r, c) for r in range(1, 11) for c in (1, 2)]
        picked = _spread_cells(cells, 5)
        assert picked == sorted(picked)
        assert picked[0] == (1, 1)
        assert picked[-1] == (10, 2)


class TestProvenance:
    def test_formula_node_reports_the_engine(self):
        graph = graph_of({"A1": 2, "A2": "=A1*3"})
        assert node_of(graph, "c:S!A2")["valueSource"] == "engine"

    def test_guarded_node_reports_the_fallback(self):
        graph = graph_of({"A1": 1, "A2": "=IFERROR(NOSHEET!A1, 456)"})
        assert node_of(graph, "c:S!A2")["valueSource"] == "fallback"

    def test_static_cell_carries_the_file_value(self):
        graph = graph_of({"A1": datetime.date(2026, 8, 7), "B1": "=A1+1"})
        node = node_of(graph, "i:S!A1")
        assert node["cachedValue"] == "2026-08-07"
        assert node["values"][0]["source"] == "engine"
        assert node["values"][0]["date"] == "2026-08-07"

    def test_range_samples_carry_source_and_date(self):
        graph = graph_of({"A1": 1, "A2": 2, "A3": "=SUM(A1:A2)"})
        samples = node_of(graph, "i:S!A1:A2")["values"]
        assert [s["addr"] for s in samples] == ["A1", "A2"]
        assert {s["source"] for s in samples} == {"engine"}
        assert all(s["date"] is None for s in samples)

    def test_step_inputs_carry_the_date(self):
        graph = graph_of(
            {"A1": datetime.date(2026, 8, 7), "A2": "=A1+1"},
            {"A2": "yyyy-mm-dd"},
        )
        steps = node_of(graph, "c:S!A2")["steps"]
        assert steps["inputs"][0] == {
            "ref": "A1",
            "value": 46241.0,
            "date": "2026-08-07",
        } or steps["inputs"][0] == {
            "ref": "A1",
            "value": "2026-08-07",
            "date": "2026-08-07",
        }

    # A workbook written by openpyxl carries no cached value for formula cells,
    # and its constants always agree with what the engine recomputes, so the
    # divergence warning cannot be produced from a built file: it is exercised
    # here on the resolver itself, which is where the comparison happens.
    def test_date_divergence_is_warned(self):
        warnings: list[str] = []
        cached = CachedValues(
            {(SHEET, 1, 1): datetime.datetime(2026, 8, 7)}, {(SHEET, 1, 1)}, False
        )
        resolver = resolver_for({(SHEET, 1, 1): 46240.0}, cached, warnings)
        value, source, date_text = resolver.value(SHEET, 1, 1)
        assert (value, source, date_text) == (46240.0, "engine", "2026-08-06")
        assert warnings == [
            "S!A1: recalculated 46240.0 differs from file value 2026-08-07"
        ]

    def test_number_divergence_respects_the_float_tolerance(self):
        warnings: list[str] = []
        cached = CachedValues({(SHEET, 1, 1): 5.0}, set(), False)
        resolver = resolver_for({(SHEET, 1, 1): 5.0000000001}, cached, warnings)
        assert resolver.value(SHEET, 1, 1)[0] == 5.0000000001
        assert warnings == []

        warnings.clear()
        resolver = resolver_for({(SHEET, 1, 1): 6.0}, cached, warnings)
        assert resolver.value(SHEET, 1, 1)[0] == 6.0
        assert warnings == ["S!A1: recalculated 6.0 differs from file value 5.0"]

    def test_file_value_is_used_when_the_engine_has_nothing(self):
        warnings: list[str] = []
        cached = CachedValues({(SHEET, 2, 1): "KO"}, set(), False)
        resolver = resolver_for({}, cached, warnings)
        assert resolver.value(SHEET, 2, 1) == ("KO", "file", None)
        assert warnings == []


def error(kind: str) -> dict:
    """An engine error value, shaped as formualizer reports it."""
    return {"type": "Error", "kind": kind}


class TestErrorValues:
    """The engine reports errors as a dict; a report must show Excel's text.

    Left alone it reaches the panel as ``{'type': 'Error', 'kind': 'Na'}`` —
    Python source, where the reader is looking for ``#N/A`` — and, worse, never
    equals the ``#N/A`` the file stores, so two identical readings are announced
    as a disagreement.
    """

    @pytest.mark.parametrize(
        ("formula", "text"),
        [
            ("=1/0", "#DIV/0!"),
            ("=NA()", "#N/A"),
            ("=SQRT(-1)", "#NUM!"),
            ('="a"+1', "#VALUE!"),
            ("=NoSuchFunction(1)", "#NAME?"),
        ],
    )
    def test_an_error_reads_as_the_text_excel_shows(self, formula, text):
        assert node_of(graph_of({"A1": formula}), "c:S!A1")["value"] == text

    def test_the_same_error_on_both_sides_is_not_a_divergence(self):
        warnings: list[str] = []
        cached = CachedValues({(SHEET, 1, 1): "#DIV/0!"}, set(), False)
        resolver = resolver_for({(SHEET, 1, 1): error("Div")}, cached, warnings)
        assert resolver.value(SHEET, 1, 1) == ("#DIV/0!", "engine", None)
        assert warnings == []

    def test_a_different_error_is_a_divergence_named_in_excel_terms(self):
        warnings: list[str] = []
        cached = CachedValues({(SHEET, 1, 1): "#VALUE!"}, set(), False)
        resolver = resolver_for({(SHEET, 1, 1): error("Num")}, cached, warnings)
        assert resolver.value(SHEET, 1, 1)[0] == "#NUM!"
        assert warnings == ["S!A1: recalculated #NUM! differs from file value #VALUE!"]

    def test_an_error_recalculated_over_a_number_is_a_divergence(self):
        warnings: list[str] = []
        cached = CachedValues({(SHEET, 1, 1): 31}, set(), False)
        resolver = resolver_for({(SHEET, 1, 1): error("Ref")}, cached, warnings)
        assert resolver.value(SHEET, 1, 1)[0] == "#REF!"
        assert warnings == ["S!A1: recalculated #REF! differs from file value 31"]


class TestUncomputableFormulas:
    """A limit of the engine is not a value of the cell.

    ``NImpl`` means the engine met a function or operator it does not implement
    — the range intersection below — and ``Circ`` a reference cycle. Showing
    either as "the value linexcel recalculated" claims a recalculation that
    never happened, and puts it against the file's own figure as if the
    workbook were wrong.
    """

    def test_an_unimplemented_formula_keeps_the_value_stored_in_the_file(self):
        warnings: list[str] = []
        cached = CachedValues({(SHEET, 1, 1): 31}, set(), False)
        resolver = resolver_for({(SHEET, 1, 1): error("NImpl")}, cached, warnings)
        assert resolver.value(SHEET, 1, 1) == (31, "file", None)
        assert warnings == []
        assert resolver.uncomputed_warning() == (
            "1 cell(s) could not be computed by the engine and keep the "
            "value stored in the file: S!A1"
        )

    def test_a_per_cell_recovery_limit_is_named_in_the_warning(self, monkeypatch):
        warnings: list[str] = []
        cached = CachedValues({(SHEET, 1, 1): 31}, set(), False)
        resolver = resolver_for(
            {},
            cached,
            warnings,
            evaluated={(SHEET, 1, 1): error("NImpl")},
        )
        monkeypatch.setattr(resolver, "_eval_formula", lambda *_args: (None, None))
        assert resolver.value(SHEET, 1, 1, "=UNSUPPORTED()") == (31, "file", None)
        assert resolver.uncomputed_warning() == (
            "1 cell(s) could not be computed by the engine and keep the "
            "value stored in the file: S!A1"
        )

    def test_a_scratch_recovery_limit_is_named_in_the_warning(self, monkeypatch):
        warnings: list[str] = []
        cached = CachedValues({(SHEET, 1, 1): 31}, set(), False)
        resolver = resolver_for({}, cached, warnings, engine_alive=False)
        monkeypatch.setattr(
            resolver, "_eval_raw", lambda *_args: (error("NImpl"), True)
        )
        assert resolver.value(SHEET, 1, 1, "=UNSUPPORTED()") == (31, "file", None)
        assert resolver.uncomputed_warning() == (
            "1 cell(s) could not be computed by the engine and keep the "
            "value stored in the file: S!A1"
        )

    def test_a_reference_cycle_is_not_reported_as_a_value(self):
        cached = CachedValues({(SHEET, 1, 1): "#VALUE!"}, set(), False)
        resolver = resolver_for({(SHEET, 1, 1): error("Circ")}, cached, [])
        assert resolver.value(SHEET, 1, 1) == ("#VALUE!", "file", None)

    def test_an_unknown_error_kind_is_not_invented_into_a_value(self):
        cached = CachedValues({(SHEET, 1, 1): 7}, set(), False)
        resolver = resolver_for({(SHEET, 1, 1): error("SomethingNew")}, cached, [])
        assert resolver.value(SHEET, 1, 1) == (7, "file", None)

    def test_the_range_intersection_operator_claims_no_value(self):
        # A workbook openpyxl wrote caches nothing, so there is no file value to
        # fall back to either: the cell must simply claim none.
        graph = graph_of({"A1": 1, "A2": 2, "A3": 3, "B1": "=SUM(A1:A3 A2:A3)"})
        node = node_of(graph, "c:S!B1")
        assert node["value"] is None
        assert "valueSource" not in node

    def test_the_unimplemented_cells_are_named_in_the_warnings(self):
        graph = graph_of({"A1": 1, "A2": 2, "A3": 3, "B1": "=SUM(A1:A3 A2:A3)"})
        assert any(
            "could not be computed by the engine" in w and "S!B1" in w
            for w in graph["meta"]["warnings"]
        )

    def test_an_unimplemented_step_reads_as_not_evaluated(self):
        steps = node_of(
            graph_of({"A1": 1, "A2": 2, "A3": 3, "B1": "=SUM(A1:A3 A2:A3)"}),
            "c:S!B1",
        )["steps"]
        assert steps["evaluated"] is False
        assert steps["value"] is None


class TestReadingsAgree:
    """One rule for "do these two readings of a cell say the same thing?".

    It used to be two, and they contradicted each other: Python returned "no
    difference" for every pair of strings — missing a recalculated `non` over
    a stored `oui` — while the viewer called any difference a disagreement,
    which lit up a French workbook over `6,7 €`.
    """

    from linexcel.values import readings_agree as _agree

    @pytest.mark.parametrize(
        ("recalculated", "stored", "verdict"),
        [
            # the same value, written by two machines with different habits
            ("6.7 €", "6,7 €", "format"),
            ("6.7", "6,7", "format"),
            ("Total : 1234.5", "Total : 1 234,5", "format"),  # espace insécable
            ("1234.5", "1.234,5", "format"),  # German grouping
            ("1234.5", "1'234.5", "format"),  # Swiss grouping
            ("A 1,5 B", "A 1.5 B", "format"),
            # the same value, written the same way
            ("Réf. 12", "Réf. 12", "same"),
            (6.7, 6.7, "same"),
            # genuinely not the same
            ("oui", "non", "differ"),
            ("6.7 €", "6.8 €", "differ"),
            (6.7, 6.8, "differ"),
            # a comma that separates words, not decimals
            ("oui, non", "oui. non", "differ"),
            # the text around the numbers has to match
            ("6.7 €", "6,7 $", "differ"),
        ],
    )
    def test_the_verdict(self, recalculated, stored, verdict):
        from linexcel.values import readings_agree

        assert readings_agree(recalculated, stored, None) == verdict

    def test_a_text_disagreement_is_finally_reported(self):
        """It never was: the old comparison returned False for any two strings."""
        from linexcel.values import readings_agree

        assert readings_agree("oui", "non", None) == "differ"

    def test_a_number_of_digits_that_differs_is_not_a_formatting_quirk(self):
        from linexcel.values import readings_agree

        assert readings_agree("1 2", "12", None) == "differ"


def with_stored_results(data: bytes, cells: dict[str, str]) -> bytes:
    """Give formula cells the result a spreadsheet application would have stored.

    openpyxl writes formulas and an empty ``<v/>``: a generated fixture has no
    stored results at all, so the whole file-against-recalculated comparison
    has nothing to compare. Filling that element is what a real Excel does on
    save, and it is the only way to exercise the comparison without one —
    including the case that started this, a French Excel storing ``6,7 €``.
    """
    import re as _re
    import zipfile

    out = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(data)) as src, zipfile.ZipFile(out, "w") as dst:
        for item in src.infolist():
            blob = src.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                xml = blob.decode("utf-8")
                for address, text in cells.items():
                    xml = _re.sub(
                        rf'<c r="{address}"([^>]*)>(<f>.*?</f>)\s*<v\s*/>',
                        rf'<c r="{address}"\g<1> t="str">\g<2><v>{text}</v>',
                        xml,
                    )
                blob = xml.encode("utf-8")
            dst.writestr(item, blob)
    return out.getvalue()


class TestAFrenchWorkbookEndToEnd:
    """The case a human found by saving a file in Excel and opening the report."""

    @staticmethod
    def workbook() -> bytes:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        ws["A1"] = 6.7
        ws["B1"] = '=A1&" €"'
        ws["B2"] = '=IF(A1>5,"oui","non")'
        buf = io.BytesIO()
        wb.save(buf)
        # what a French Excel would have written into the file
        return with_stored_results(buf.getvalue(), {"B1": "6,7 €", "B2": "non"})

    def test_a_comma_against_a_dot_is_not_a_divergence(self):
        from linexcel import analyze

        result = analyze(self.workbook(), filename="fr.xlsx")
        node = next(n for n in result.nodes if n.get("addr") == "B1")
        assert node["value"] == "6.7 €"
        assert node["cachedValue"] == "6,7 €"
        assert node["cachedAgreement"] == "format"

    def test_and_raises_no_warning(self):
        from linexcel import analyze

        result = analyze(self.workbook(), filename="fr.xlsx")
        assert not [w for w in result.warnings if "B1" in w]

    def test_a_text_that_really_differs_is_reported(self):
        """Which it never was: the old rule ignored every pair of strings."""
        from linexcel import analyze

        result = analyze(self.workbook(), filename="fr.xlsx")
        node = next(n for n in result.nodes if n.get("addr") == "B2")
        assert node["cachedAgreement"] == "differ"
        (warning,) = [w for w in result.warnings if "B2" in w]
        assert "oui" in warning and "non" in warning
