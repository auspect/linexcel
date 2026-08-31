"""Génère des fixtures Excel adversariales (.xlsx/.xlsm) pour linexcel.

Cas hostiles couverts (cf. ADVERSARIAL_SPEC.md, axe A) :
- macros VBA inoffensives (.xlsm avec vbaProject.bin sain,
  donneur : tests/fixtures/macros.xlsm)
- macro « hostile » : vbaProject.bin tronqué/corrompu (erreur de parsing OLE)
- macro security block : classeur macro-enabled dont le vbaProject.bin a été
  stripé par la politique de sécurité (content-type xlsm, pas de bin)
- étiquettes de confidentialité : docProps (core/app/custom) avec
  company/classification, style étiquette MIP manipulée par xlwings
- dates système 1904 (workbookPr date1904="1")
- refs cassées =NOSHEET!A1, =IFERROR(NOSHEET!A1,-1), =SUM(...) sur formules,
  cycles auto-référents, formules croisées multi-feuilles
- enrichissements : chaîne de dépendances > MAX_CHAIN_DEPTH, refs externes
  [Budget.xlsx] sans refs_dir, formules volatiles (TODAY/NOW), feuilles au
  nom unicode avec espaces, noms définis pointant vers une feuille absente,
  zip tronqué et fichier non-zip à extension .xlsx (frontière publique)

Usage :
    uv run python tools/gen_fixtures.py [--rows N] [--cols N] [--sheets N]
                                        [--out DIR] [--only nom1,nom2]
"""

from __future__ import annotations

import argparse
import datetime as dt
import io
import zipfile
from collections.abc import Callable
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_OUT = ROOT / "tests" / "fixtures" / "adversarial"
VBA_DONOR = ROOT / "tests" / "fixtures" / "macros.xlsm"

CT_XLSX_MAIN = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"
)
CT_XLSM_MAIN = "application/vnd.ms-excel.sheet.macroEnabled.main+xml"
CT_VBA_BIN = "application/vnd.ms-office.vbaProject"
REL_VBA = "http://schemas.microsoft.com/office/2006/relationships/vbaProject"


def _grid(wb: Workbook, rows: int, cols: int, sheets: int, prefix: str = "S") -> None:
    """Remplit `wb` : grille de valeurs + formules, formules croisées."""
    for s in range(sheets):
        ws = wb.create_sheet(f"{prefix}{s + 1}")
        for r in range(1, rows + 1):
            for c in range(1, cols + 1):
                cell = ws.cell(row=r, column=c)
                if r == 1:
                    cell.value = r * 100 + c  # en-tête de valeurs
                elif c == 1:
                    # ref croisée vers la feuille précédente, valeur sur la 1re
                    cell.value = f"={prefix}{s}!A{r}+1" if s > 0 else r * 10
                elif c == cols and cols > 1:
                    col_letter = cell.offset(column=-1).column_letter
                    cell.value = f"=SUM(A{r}:{col_letter}{r})"
                else:
                    cell.value = f"=A{r}*{c}"


def _fresh_wb() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _save(wb: Workbook, path: Path) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    path.write_bytes(data)
    return data


def _rewrite_zip(data: bytes, edits: dict[str, bytes | None]) -> bytes:
    """Réécrit un zip OOXML : remplace (bytes) ou supprime (None) des parts."""
    out = io.BytesIO()
    src = zipfile.ZipFile(io.BytesIO(data))
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            if item.filename in edits:
                replacement = edits.pop(item.filename)
                if replacement is not None:
                    dst.writestr(item.filename, replacement)
            else:
                dst.writestr(item, src.read(item.filename))
        for name, content in edits.items():
            if content is not None:
                dst.writestr(name, content)
    return out.getvalue()


def _donor_vba() -> bytes:
    with zipfile.ZipFile(VBA_DONOR) as z:
        return z.read("xl/vbaProject.bin")


def _to_xlsm(data: bytes, vba_bin: bytes | None) -> bytes:
    """Convertit un .xlsx en .xlsm : content-type macro-enabled + vbaProject.

    ``vba_bin=None`` simule le strip de sécurité (container xlsm sans bin).
    """
    src = zipfile.ZipFile(io.BytesIO(data))
    ct = src.read("[Content_Types].xml").decode()
    ct = ct.replace(CT_XLSX_MAIN, CT_XLSM_MAIN)
    rels = src.read("xl/_rels/workbook.xml.rels").decode()
    edits: dict[str, bytes | None] = {}
    if vba_bin is not None:
        if 'Extension="bin"' not in ct:
            ct = ct.replace(
                '<Default Extension="xml"',
                f'<Default Extension="bin" ContentType="{CT_VBA_BIN}"/>'
                '<Default Extension="xml"',
            )
        rels = rels.replace(
            "</Relationships>",
            '<Relationship Id="rIdVba" Type="' + REL_VBA + '" '
            'Target="vbaProject.bin"/></Relationships>',
        )
        edits["xl/vbaProject.bin"] = vba_bin
    edits["[Content_Types].xml"] = ct.encode()
    edits["xl/_rels/workbook.xml.rels"] = rels.encode()
    return _rewrite_zip(data, edits)


def _inject_confidentiality(data: bytes) -> bytes:
    """Ajoute des étiquettes de confidentialité dans docProps."""
    src = zipfile.ZipFile(io.BytesIO(data))
    core = src.read("docProps/core.xml").decode()
    core = core.replace(
        "</cp:coreProperties>",
        "<cp:keywords>classification:SECRET; internal-only</cp:keywords>"
        "<cp:category>Confidential</cp:category></cp:coreProperties>",
    )
    edits: dict[str, bytes | None] = {"docProps/core.xml": core.encode()}
    names = src.namelist()
    if "docProps/app.xml" in names:
        app = src.read("docProps/app.xml").decode()
        app = app.replace(
            "</Properties>",
            "<Company>ACME Corp — Classification: SECRET</Company></Properties>",
        )
        edits["docProps/app.xml"] = app.encode()
    # Étiquette MIP-like en propriété personnalisée
    custom = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/'
        'custom-properties" xmlns:vt="http://schemas.openxmlformats.org/'
        'officeDocument/2006/docPropsVTypes">'
        '<property fmtid="{D5CDD505-2E9C-101B-9397-08002B2CF9AE}" pid="2" '
        'name="MSIP_Label_a1b2c3_Enabled"><vt:bool>true</vt:bool></property>'
        '<property fmtid="{D5CDD505-2E9C-101B-9397-08002B2CF9AE}" pid="3" '
        'name="MSIP_Label_a1b2c3_Name"><vt:lpwstr>SECRET</vt:lpwstr></property>'
        "</Properties>"
    )
    edits["docProps/custom.xml"] = custom.encode()
    ct = src.read("[Content_Types].xml").decode()
    ct = ct.replace(
        "</Types>",
        '<Override PartName="/docProps/custom.xml" ContentType="application/'
        'vnd.openxmlformats-officedocument.custom-properties+xml"/></Types>',
    )
    edits["[Content_Types].xml"] = ct.encode()
    rels = src.read("_rels/.rels").decode()
    rels = rels.replace(
        "</Relationships>",
        '<Relationship Id="rIdCustom" Type="http://schemas.openxmlformats.org/'
        'officeDocument/2006/relationships/custom-properties" '
        'Target="docProps/custom.xml"/></Relationships>',
    )
    edits["_rels/.rels"] = rels.encode()
    return _rewrite_zip(data, edits)


# ---------------------------------------------------------------- fixtures


def make_realiste(rows: int, cols: int, sheets: int) -> Workbook:
    wb = _fresh_wb()
    _grid(wb, rows, cols, sheets)
    return wb


def make_formules_croisees(rows: int, cols: int, sheets: int) -> Workbook:
    wb = _fresh_wb()
    _grid(wb, max(rows, 5), max(cols, 3), max(sheets, 3), prefix="Feuil")
    ws = wb.create_sheet("Synthese")
    for s in range(1, wb.sheetnames.index("Synthese")):
        ws.cell(row=s, column=1).value = f"=SUM(Feuil{s}!A1:A{max(rows, 5)})"
    ws.cell(row=99, column=1).value = "=Feuil1!B2*Feuil3!B2"
    return wb


def make_nosheet(rows: int, cols: int, sheets: int) -> Workbook:
    wb = make_realiste(max(rows, 8), max(cols, 4), max(sheets, 1))
    ws = wb[wb.sheetnames[0]]
    ws["D2"] = "=NOSHEET!A1"
    ws["D3"] = "=NOSHEET!A1+Data!B2"
    ws["D4"] = "=SUM(NOSHEET!A1:A10)"
    return wb


def make_iferror_nosheet(rows: int, cols: int, sheets: int) -> Workbook:
    wb = make_realiste(max(rows, 8), max(cols, 4), max(sheets, 1))
    ws = wb[wb.sheetnames[0]]
    ws["E2"] = "=IFERROR(NOSHEET!A1,-1)"
    ws["E3"] = '=IFERROR(VLOOKUP(NOSHEET!A1,A1:B8,2,FALSE),"absent")'
    return wb


def make_sum_sur_formules(rows: int, cols: int, sheets: int) -> Workbook:
    wb = _fresh_wb()
    ws = wb.create_sheet("Calc")
    n = max(rows, 7)
    for r in range(1, n + 1):
        ws.cell(row=r, column=1).value = f"=B{r}*2"  # A dépend de B
        ws.cell(row=r, column=2).value = r  # B : valeurs
    ws.cell(row=n + 2, column=1).value = f"=SUM(A1:A{n})"
    ws.cell(row=n + 3, column=1).value = f"=SUM(A1:A{n})+SUM(B1:B{n})"
    return wb


def make_cycles(rows: int, cols: int, sheets: int) -> Workbook:
    wb = make_realiste(max(rows, 6), max(cols, 4), max(sheets, 1))
    ws = wb[wb.sheetnames[0]]
    ws["F1"] = "=F1+1"  # auto-référence directe
    ws["F2"] = "=F3+1"  # cycle mutuel F2 <-> F3
    ws["F3"] = "=F2+1"
    ws["G1"] = "=F1*2"  # consommateur du cycle
    return wb


def make_dates_1904(rows: int, cols: int, sheets: int) -> Workbook:
    wb = make_realiste(max(rows, 6), max(cols, 4), max(sheets, 1))
    wb.epoch = dt.datetime(1904, 1, 1)  # workbookPr date1904="1"
    ws = wb[wb.sheetnames[0]]
    ws["H1"] = dt.datetime(2026, 8, 31)
    ws["H2"] = "=H1+30"
    ws["H3"] = dt.date(1900, 3, 1)  # ambiguïté 1900/1904
    return wb


def make_confidentiel(rows: int, cols: int, sheets: int) -> Workbook:
    wb = make_realiste(max(rows, 6), max(cols, 4), max(sheets, 1))
    wb.properties.creator = "Direction Financière"
    wb.properties.title = "Budget prévisionnel — NE PAS DIFFUSER"
    return wb


def make_vba_base(rows: int, cols: int, sheets: int) -> Workbook:
    wb = make_realiste(max(rows, 6), max(cols, 4), max(sheets, 1))
    ws = wb.create_sheet("MacroData")
    ws["A1"] = "=MacroFunc(B1)"
    ws["B1"] = 42
    return wb


def make_chaine_profonde(rows: int, cols: int, sheets: int) -> Workbook:
    """Chaîne de dépendances plus profonde que MAX_CHAIN_DEPTH (24)."""
    wb = _fresh_wb()
    ws = wb.create_sheet("Chaine")
    depth = max(rows, 60)
    ws["A1"] = 1
    for r in range(2, depth + 1):
        ws.cell(row=r, column=1).value = f"=A{r - 1}+1"
    ws["B1"] = f"=A{depth}*2"  # consommateur en bout de chaîne
    return wb


def make_refs_externes(rows: int, cols: int, sheets: int) -> Workbook:
    wb = make_realiste(max(rows, 6), max(cols, 4), max(sheets, 1))
    ws = wb[wb.sheetnames[0]]
    ws["J1"] = "='[Budget.xlsx]Annual'!B4"
    ws["J2"] = "='[Budget.xlsx]Annual'!B4*2+A2"
    return wb


def make_volatiles(rows: int, cols: int, sheets: int) -> Workbook:
    wb = make_realiste(max(rows, 6), max(cols, 4), max(sheets, 1))
    ws = wb[wb.sheetnames[0]]
    ws["K1"] = "=TODAY()"  # pas de clock -> 25569 (1970) attendu
    ws["K2"] = "=NOW()"
    ws["K3"] = "=K1+365"
    return wb


def make_feuilles_unicode(rows: int, cols: int, sheets: int) -> Workbook:
    wb = _fresh_wb()
    ws1 = wb.create_sheet("Données 2026")
    ws2 = wb.create_sheet("Synthèse été")
    for r in range(1, max(rows, 8) + 1):
        ws1.cell(row=r, column=1).value = r
    ws2["A1"] = "=SUM('Données 2026'!A1:A8)"
    ws2["A2"] = "='Données 2026'!A3*2"
    return wb


def make_noms_definis_casses(rows: int, cols: int, sheets: int) -> Workbook:
    from openpyxl.workbook.defined_name import DefinedName

    wb = make_realiste(max(rows, 6), max(cols, 4), max(sheets, 1))
    wb.defined_names["TauxOK"] = DefinedName("TauxOK", attr_text="S1!$B$2")
    wb.defined_names["TauxKO"] = DefinedName("TauxKO", attr_text="NOSHEET!$A$1")
    ws = wb[wb.sheetnames[0]]
    ws["L1"] = "=TauxOK*2"
    ws["L2"] = "=TauxKO*2"
    return wb


# (nom fichier, builder, post-traitement)
FIXTURES: list[tuple[str, Callable[[int, int, int], Workbook], str]] = [
    ("realiste.xlsx", make_realiste, "xlsx"),
    ("formules_croisees.xlsx", make_formules_croisees, "xlsx"),
    ("refs_cassees_nosheet.xlsx", make_nosheet, "xlsx"),
    ("iferror_nosheet.xlsx", make_iferror_nosheet, "xlsx"),
    ("sum_sur_formules.xlsx", make_sum_sur_formules, "xlsx"),
    ("cycles.xlsx", make_cycles, "xlsx"),
    ("dates_1904.xlsx", make_dates_1904, "xlsx"),
    ("chaine_profonde.xlsx", make_chaine_profonde, "xlsx"),
    ("refs_externes.xlsx", make_refs_externes, "xlsx"),
    ("volatiles.xlsx", make_volatiles, "xlsx"),
    ("feuilles_unicode.xlsx", make_feuilles_unicode, "xlsx"),
    ("noms_definis_casses.xlsx", make_noms_definis_casses, "xlsx"),
    ("confidentialite.xlsx", make_confidentiel, "confidentiel"),
    ("macros_inoffensives.xlsm", make_vba_base, "vba_sain"),
    ("macro_hostile_corrompue.xlsm", make_vba_base, "vba_corrompu"),
    ("macro_security_block.xlsm", make_vba_base, "vba_stripe"),
    ("fichier_tronque.xlsx", make_realiste, "tronque"),
    ("pas_un_zip.xlsx", make_realiste, "brut"),
]


def generate(
    out: Path, rows: int, cols: int, sheets: int, only: set[str] | None
) -> list[Path]:
    out.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []
    vba_bin = _donor_vba()
    for name, builder, kind in FIXTURES:
        stem = Path(name).stem
        if only and stem not in only:
            continue
        path = out / name
        data = _save(builder(rows, cols, sheets), path)
        if kind == "confidentiel":
            path.write_bytes(_inject_confidentiality(data))
        elif kind == "vba_sain":
            path.write_bytes(_to_xlsm(data, vba_bin))
        elif kind == "vba_corrompu":
            path.write_bytes(_to_xlsm(data, vba_bin[: len(vba_bin) // 3]))
        elif kind == "vba_stripe":
            path.write_bytes(_to_xlsm(data, None))
        elif kind == "tronque":
            # zip coupé au milieu du flux : BadZipFile côté loader
            path.write_bytes(data[: len(data) // 2])
        elif kind == "brut":
            # extension .xlsx mais contenu quelconque (CSV renommé, par ex.)
            path.write_bytes(b"col1;col2\n1;2\n3;4\n")
        written.append(path)
    return written


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--rows", type=int, default=50, help="lignes par feuille")
    parser.add_argument("--cols", type=int, default=12, help="colonnes par feuille")
    parser.add_argument("--sheets", type=int, default=3, help="nombre de feuilles")
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    parser.add_argument(
        "--only", type=str, default=None, help="stems séparés par virgule"
    )
    args = parser.parse_args()
    only = set(args.only.split(",")) if args.only else None
    written = generate(args.out, args.rows, args.cols, args.sheets, only)
    for path in written:
        try:
            shown = path.relative_to(ROOT)
        except ValueError:  # --out hors du dépôt
            shown = path
        print(f"  {shown} ({path.stat().st_size} o)")
    print(f"{len(written)} fixtures générées dans {args.out}")


if __name__ == "__main__":
    main()
