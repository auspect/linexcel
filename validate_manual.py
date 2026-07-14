#!/usr/bin/env python3
import io
import os
import shutil
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.workbook.defined_name import DefinedName
except ImportError:
    print("❌ Error: openpyxl is not installed. Run 'uv run validate_manual.py'")
    sys.exit(1)


# Add src/ to sys.path so we can import the local linexcel package directly
sys.path.insert(0, str(Path(__file__).parent / "src"))
import linexcel


def build_sample_workbook() -> bytes:
    """Generate a rich Excel workbook with styles, charts, and shifted cells."""
    wb = Workbook()

    # --- STYLES DEFINITION ---
    font_title = Font(name="Segoe UI", size=14, bold=True, color="1F4E79")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=11)
    font_bold = Font(name="Segoe UI", size=11, bold=True)

    fill_header = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    fill_zebra = PatternFill(
        start_color="F2F5F8", end_color="F2F5F8", fill_type="solid"
    )

    border_thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    border_header = Border(
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
        top=Side(style="medium", color="1F4E79"),
        bottom=Side(style="medium", color="1F4E79"),
    )

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # --- 1. SHEET: Ventes ---
    ws = wb.active
    ws.title = "Ventes"

    # Leaving empty Row 1 & 2, and empty Column A. Table starts at B3.
    ws["B2"] = "Rapport de Ventes Hebdomadaire"
    ws["B2"].font = font_title

    headers = ["Produit", "Qté", "Prix", "CA"]
    for col_idx, text in enumerate(headers, start=2):
        cell = ws.cell(row=3, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_header

    for r in range(4, 104):
        p_cell = ws.cell(row=r, column=2, value=f"P{r - 3}")
        q_cell = ws.cell(row=r, column=3, value=r % 7 + 1)
        pr_cell = ws.cell(row=r, column=4, value=10.5 + (r % 13))
        ca_cell = ws.cell(
            row=r, column=5, value=f"=C{r}*D{r}"
        )  # Formula relative to C and D

        row_fill = fill_zebra if r % 2 == 0 else None
        for col_idx, cell in enumerate([p_cell, q_cell, pr_cell, ca_cell], start=2):
            cell.font = font_data
            cell.border = border_thin
            if row_fill:
                cell.fill = row_fill

        p_cell.alignment = align_left
        q_cell.alignment = align_right
        pr_cell.alignment = align_right
        ca_cell.alignment = align_right

        pr_cell.number_format = "$#,##0.00"
        ca_cell.number_format = "$#,##0.00"

    # Freeze row 3 & column A (so we specify B4)
    ws.freeze_panes = "B4"
    ws.column_dimensions["D"].hidden = True  # Hide column D (Prix)

    # Configure print settings: Landscape, paper size A3 (large), fit to width
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.merge_cells("G3:H3")
    ws["G3"] = "Contexte de Présentation"
    ws["G3"].font = font_bold
    ws["G3"].alignment = align_center
    ws["G3"].border = border_thin

    # A1 comment was shifted to B3 (new table start)
    ws["B3"].comment = Comment("Exported product category", "Data team")

    # B width adjustment for better visibility
    ws.column_dimensions["B"].width = 30

    # Add Chart to "Ventes"
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Chiffre d'Affaires par Produit (Top 10)"
    chart.y_axis.title = "CA ($)"
    chart.x_axis.title = "Produit"

    # References for chart (CA is col 5, Product is col 2)
    data = Reference(ws, min_col=5, min_row=3, max_row=13)
    cats = Reference(ws, min_col=2, min_row=4, max_row=13)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 16
    chart.height = 10
    ws.add_chart(chart, "G5")

    # --- 2. SHEET: Synthese ---
    syn = wb.create_sheet("Synthese")
    syn["B2"] = "Synthèse des Performances"
    syn["B2"].font = font_title

    # Headers
    syn.cell(row=4, column=2, value="Métrique").font = font_header
    syn.cell(row=4, column=2).fill = fill_header
    syn.cell(row=4, column=2).border = border_header
    syn.cell(row=4, column=2).alignment = align_left

    syn.cell(row=4, column=3, value="Valeur").font = font_header
    syn.cell(row=4, column=3).fill = fill_header
    syn.cell(row=4, column=3).border = border_header
    syn.cell(row=4, column=3).alignment = align_right

    # Formulas reference Ventes!E4:E103 (the CA column E)
    m_cells = [
        ("Total CA", "=SUM(Ventes!E4:E103)", "$#,##0.00"),
        (
            "CA Moyen par Produit",
            "=ROUND(AVERAGE(Ventes!E4:E103), 2)",
            "$#,##0.00",
        ),
        (
            "Statut Objectif",
            "=IF(SUM(Ventes!E4:E103)>TauxCible, "
            'CONCATENATE("OK: ", ROUND(C5/1000,1), "k"), "KO")',
            None,
        ),
    ]

    for idx, (label, formula, num_fmt) in enumerate(m_cells, start=5):
        lbl_cell = syn.cell(row=idx, column=2, value=label)
        lbl_cell.font = font_bold
        lbl_cell.border = border_thin

        val_cell = syn.cell(row=idx, column=3, value=formula)
        val_cell.font = font_data
        val_cell.border = border_thin
        val_cell.alignment = align_right
        if num_fmt:
            val_cell.number_format = num_fmt

    # Configure print settings: Landscape, paper size A3 (large), fit to width
    syn.page_setup.orientation = syn.ORIENTATION_LANDSCAPE
    syn.page_setup.paperSize = syn.PAPERSIZE_A3
    syn.page_setup.fitToWidth = 1
    syn.page_setup.fitToHeight = 0
    syn.sheet_properties.pageSetUpPr.fitToPage = True

    # --- 3. SHEET: Params ---
    params = wb.create_sheet("Params")
    params["B2"] = "Paramètres de Simulation"
    params["B2"].font = font_title

    lbl_cell = params.cell(row=4, column=2, value="Seuil CA Cible")
    lbl_cell.font = font_bold
    lbl_cell.border = border_thin

    val_cell = params.cell(row=4, column=3, value=5000)
    val_cell.font = font_data
    val_cell.border = border_thin
    val_cell.number_format = "$#,##0.00"
    val_cell.alignment = align_right

    wb.defined_names.add(DefinedName("TauxCible", attr_text="Params!$C$4"))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def main():
    print("--- 📊 Validation Manuelle de linexcel ---")
    data = build_sample_workbook()
    excel_path = Path("validation_demo.xlsx")
    excel_path.write_bytes(data)
    print(f"   Classeur de test enregistré sous : {excel_path.resolve()}")

    print("1. Analyse du classeur avec linexcel...")
    result = linexcel.analyze(data, filename="validation_demo.xlsx")

    print("   Structure détectée :")
    print(f"     - Feuilles : {', '.join(result.sheets)}")
    print(f"     - Stats : {result.stats}")

    # Check for Gemini API key
    api_key = os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY")
    docs_fr, docs_en = None, None
    wb_doc_fr, wb_doc_en = None, None

    if api_key:
        print(
            "\n2. 🔑 Clé API Gemini trouvée. "
            "Génération de la documentation de calculs par l'IA..."
        )
        try:
            print("   - Génération de la version française...")
            docs_fr = result.document(api_key=api_key, language="fr")
            wb_doc_fr = result.document_workbook(api_key=api_key, language="fr")

            print("   - Génération de la version anglaise...")
            docs_en = result.document(api_key=api_key, language="en")
            wb_doc_en = result.document_workbook(api_key=api_key, language="en")
            print("   - Documentations générées avec succès.")
        except Exception as e:
            print(f"   ❌ Erreur lors de la génération IA : {e}")
            print("   Poursuite de la génération sans les données IA.")
    else:
        print("\n2. ⚠️ Pas de clé API Gemini trouvée (export sans IA).")
        print(
            "   Définissez la variable d'environnement GEMINI_API_KEY pour tester l'IA."
        )
        print('   Exemple: export GEMINI_API_KEY="votre_cle_ici"')

    print("\n3. 💬 Extraction des métadonnées et commentaires de l'Excel...")
    context = result.workbook_context
    for sheet in context["sheets"]:
        print(f"   Feuille '{sheet['name']}' :")
        if sheet.get("freeze_panes"):
            print(f"     - Volet figé : {sheet['freeze_panes']}")
        if sheet.get("hidden_columns"):
            print(f"     - Colonnes masquées : {', '.join(sheet['hidden_columns'])}")
        if sheet.get("merged_ranges"):
            print(f"     - Cellules fusionnées : {', '.join(sheet['merged_ranges'])}")
        comments = sheet.get("comments", [])
        if comments:
            print("     - Commentaires :")
            for c in comments:
                print(f"       * {c['cell']} ({c['author']}) : {c['text'].strip()}")
        else:
            print("     - Aucun commentaire.")

    print("\n4. 📸 Génération des captures d'écran (screenshots)...")
    screenshots_payload = None
    try:
        screenshots_dir = Path("validation_screenshots")
        if screenshots_dir.exists():
            shutil.rmtree(screenshots_dir)
        screenshots_dir.mkdir(exist_ok=True)
        screenshots = result.save_screenshots(screenshots_dir, dpi=200)
        print(
            f"   ✅ {len(screenshots)} capture(s) d'écran enregistrée(s) "
            f"dans {screenshots_dir.name}/ :"
        )
        for s in screenshots:
            print(f"     - {s.name}")

        if screenshots:
            sorted_screens = sorted(screenshots, key=lambda x: x.name)
            if len(sorted_screens) >= 5:
                screenshots_payload = {
                    "Ventes": sorted_screens[0:3],
                    "Synthese": [sorted_screens[3]],
                    "Params": [sorted_screens[4]],
                }
            else:
                screenshots_payload = sorted_screens
    except Exception as e:
        print(f"   ⚠️ Impossible de générer les captures d'écran : {e}")
        print(
            "      (Nécessite LibreOffice headless et 'pdftoppm' installés "
            "sur le système)"
        )

    print("\n5. Enregistrement des fichiers HTML de visualisation...")
    output_fr = Path("validate_out_fr.html")
    output_en = Path("validate_out_en.html")

    result.save_html(
        output_fr,
        docs=docs_fr,
        workbook_doc=wb_doc_fr,
        screenshots=screenshots_payload,
        language="fr",
    )
    result.save_html(
        output_en,
        docs=docs_en,
        workbook_doc=wb_doc_en,
        screenshots=screenshots_payload,
        language="en",
    )

    print("\n🎉 Succès ! Les fichiers de visualisation ont été écrits ici :")
    print(f"   👉 Version française : {output_fr.resolve()}")
    print(f"   👉 Version anglaise  : {output_en.resolve()}")
    print("\nPour inspecter manuellement les résultats :")
    print("   1. Ouvrez l'un des fichiers HTML dans votre navigateur web.")
    if api_key:
        print(
            "   2. Regardez l'onglet 'Synthèse' / 'Workbook overview' pour "
            "voir la description globale par l'IA."
        )
        print(
            "   3. Sélectionnez des nœuds de formule pour voir la "
            "documentation IA correspondante."
        )
    print(
        "   4. Vérifiez les dossiers 'validation_screenshots/' pour "
        "voir le rendu visuel."
    )


if __name__ == "__main__":
    main()
