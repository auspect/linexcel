"""Workbooks used by ``validate_manual.py``.

Two fixtures, with different jobs:

``build_sales_workbook`` is a small, plausible French sales report. It is what
the README screenshots are captured from, so it stays readable: a table that
does not start at A1, a hidden column, a merged range, a chart, a comment, a
defined name and cross-sheet aggregation. Enough to show every tab of the report
doing something, and little enough to take in at a glance.

``build_stress_workbook`` is the opposite. It is an English workbook built to
break things: every Excel error value, references the analyser cannot resolve,
sheet names that need quoting, a circular pair, formulas past the dossier
truncation limit, and cell text chosen to escape the generated HTML if the
escaping is wrong. Nothing in it is meant to look like a real business file —
each sheet is a category of failure, named after what it attacks.

Not covered here: VBA. ``openpyxl`` can preserve an existing ``vbaProject.bin``
but cannot author one, so exercising ``linexcel.vba`` needs a real ``.xlsm``
recorded by Excel. ``tests/test_lineage.py`` covers that path with injected
modules instead.
"""

from __future__ import annotations

import datetime
import io

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

# ──────────────────────────────────────────────
# 1. Sales report — the readable one
# ──────────────────────────────────────────────


def build_sales_workbook() -> bytes:
    """Generate a rich Excel workbook with styles, charts, and shifted cells."""
    wb = Workbook()

    # --- STYLES DEFINITION ---
    font_title = Font(name="Segoe UI", size=14, bold=True, color="1F4E79")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=11)
    font_bold = Font(name="Segoe UI", size=11, bold=True)

    fill_header = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    fill_zebra = PatternFill(
        start_color="F2F5F8", end_color="F2F5F8", fill_type="solid"
    )

    border_thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    border_header = Border(
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
        top=Side(style="medium", color="1F4E79"),
        bottom=Side(style="medium", color="1F4E79"),
    )

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # --- 1. SHEET: Ventes ---
    ws = wb.active
    ws.title = "Ventes"

    # Leaving empty Row 1 & 2, and empty Column A. Table starts at B3.
    ws["B2"] = "Rapport de Ventes Hebdomadaire"
    ws["B2"].font = font_title

    headers = ["Produit", "Qté", "Prix", "CA"]
    for col_idx, text in enumerate(headers, start=2):
        cell = ws.cell(row=3, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_header

    for r in range(4, 104):
        p_cell = ws.cell(row=r, column=2, value=f"P{r - 3}")
        q_cell = ws.cell(row=r, column=3, value=r % 7 + 1)
        pr_cell = ws.cell(row=r, column=4, value=10.5 + (r % 13))
        ca_cell = ws.cell(
            row=r, column=5, value=f"=C{r}*D{r}"
        )  # Formula relative to C and D

        row_fill = fill_zebra if r % 2 == 0 else None
        for col_idx, cell in enumerate([p_cell, q_cell, pr_cell, ca_cell], start=2):
            cell.font = font_data
            cell.border = border_thin
            if row_fill:
                cell.fill = row_fill

        p_cell.alignment = align_left
        q_cell.alignment = align_right
        pr_cell.alignment = align_right
        ca_cell.alignment = align_right

        pr_cell.number_format = "$#,##0.00"
        ca_cell.number_format = "$#,##0.00"

    # Freeze row 3 & column A (so we specify B4)
    ws.freeze_panes = "B4"
    ws.column_dimensions["D"].hidden = True  # Hide column D (Prix)

    # Configure print settings: Landscape, paper size A3 (large), fit to width
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.merge_cells("G3:H3")
    ws["G3"] = "Contexte de Présentation"
    ws["G3"].font = font_bold
    ws["G3"].alignment = align_center
    ws["G3"].border = border_thin

    # A1 comment was shifted to B3 (new table start)
    ws["B3"].comment = Comment("Exported product category", "Data team")

    # B width adjustment for better visibility
    ws.column_dimensions["B"].width = 30

    # Add Chart to "Ventes"
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Chiffre d'Affaires par Produit (Top 10)"
    # openpyxl's axis descriptor coerces a str into a Title; ty only sees the
    # declared Title type.
    chart.y_axis.title = "CA ($)"  # ty: ignore[invalid-assignment]
    chart.x_axis.title = "Produit"  # ty: ignore[invalid-assignment]

    # References for chart (CA is col 5, Product is col 2)
    data = Reference(ws, min_col=5, min_row=3, max_row=13)
    cats = Reference(ws, min_col=2, min_row=4, max_row=13)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 16
    chart.height = 10
    ws.add_chart(chart, "G5")

    # --- 2. SHEET: Synthese ---
    syn = wb.create_sheet("Synthese")
    syn["B2"] = "Synthèse des Performances"
    syn["B2"].font = font_title

    # Headers
    syn.cell(row=4, column=2, value="Métrique").font = font_header
    syn.cell(row=4, column=2).fill = fill_header
    syn.cell(row=4, column=2).border = border_header
    syn.cell(row=4, column=2).alignment = align_left

    syn.cell(row=4, column=3, value="Valeur").font = font_header
    syn.cell(row=4, column=3).fill = fill_header
    syn.cell(row=4, column=3).border = border_header
    syn.cell(row=4, column=3).alignment = align_right

    # Formulas reference Ventes!E4:E103 (the CA column E)
    m_cells = [
        ("Total CA", "=SUM(Ventes!E4:E103)", "$#,##0.00"),
        (
            "CA Moyen par Produit",
            "=ROUND(AVERAGE(Ventes!E4:E103), 2)",
            "$#,##0.00",
        ),
        (
            "Statut Objectif",
            "=IF(SUM(Ventes!E4:E103)>TauxCible, "
            'CONCATENATE("OK: ", ROUND(C5/1000,1), "k"), "KO")',
            None,
        ),
    ]

    for idx, (label, formula, num_fmt) in enumerate(m_cells, start=5):
        lbl_cell = syn.cell(row=idx, column=2, value=label)
        lbl_cell.font = font_bold
        lbl_cell.border = border_thin

        val_cell = syn.cell(row=idx, column=3, value=formula)
        val_cell.font = font_data
        val_cell.border = border_thin
        val_cell.alignment = align_right
        if num_fmt:
            val_cell.number_format = num_fmt

    # Configure print settings: Landscape, paper size A3 (large), fit to width
    syn.page_setup.orientation = syn.ORIENTATION_LANDSCAPE
    syn.page_setup.paperSize = syn.PAPERSIZE_A3
    syn.page_setup.fitToWidth = 1
    syn.page_setup.fitToHeight = 0
    syn.sheet_properties.pageSetUpPr.fitToPage = True

    # --- 3. SHEET: Params ---
    params = wb.create_sheet("Params")
    params["B2"] = "Paramètres de Simulation"
    params["B2"].font = font_title

    lbl_cell = params.cell(row=4, column=2, value="Seuil CA Cible")
    lbl_cell.font = font_bold
    lbl_cell.border = border_thin

    val_cell = params.cell(row=4, column=3, value=5000)
    val_cell.font = font_data
    val_cell.border = border_thin
    val_cell.number_format = "$#,##0.00"
    val_cell.alignment = align_right

    wb.defined_names.add(DefinedName("TauxCible", attr_text="Params!$C$4"))

    return _to_bytes(wb)


# ──────────────────────────────────────────────
# 2. Stress workbook — the hostile one
# ──────────────────────────────────────────────

#: Rows in the stretched table. Large enough that grouping must collapse it to a
#: single node, small enough that LibreOffice still renders it quickly.
STRESS_ROWS = 250

#: Strings that must survive extraction, JSON embedding and HTML rendering
#: unchanged. Each one has broken something in some report generator: the first
#: two close the script tag the graph is embedded in, the next two collide with
#: the viewer's own template placeholders, ``$&`` is a regex replacement
#: reference, and U+2028 is a newline to a JavaScript string literal but not to
#: JSON.
HOSTILE_STRINGS = (
    "</script><script>alert(1)</script>",
    '"><img src=x onerror=alert(1)>',
    "__TITLE__",
    "__GRAPH_JSON__",
    "value is $& and $1",
    "line\u2028separator",
    "quote\" apostrophe' backslash\\ tab\t",
    "=cmd|'/c calc'!A1",  # CSV-injection shape, harmless as a stored string
)


def build_stress_workbook() -> bytes:
    """Generate an English workbook that attacks the analysis pipeline.

    Every sheet is one category of problem, so a report that renders all of them
    has exercised: error propagation, unresolvable references, quoting, deep and
    oversized formulas, hidden and empty sheets, and HTML escaping.
    """
    wb = Workbook()
    _stress_data(wb)
    _stress_lookups(wb)
    _stress_errors(wb)
    _stress_volatile(wb)
    _stress_cross_refs(wb)
    _stress_quoting(wb)
    _stress_hidden(wb)
    _stress_hostile(wb)
    wb.create_sheet("Empty Sheet")  # no cells at all: every bound is degenerate
    _stress_names(wb)
    return _to_bytes(wb)


def _stress_data(wb: Workbook) -> None:
    """A wide table of mixed types, and the stretched formulas over it."""
    ws = wb.active
    ws.title = "Sales Data"

    headers = ["Region", "Rep", "Booked", "Units", "Price", "Revenue", "Flag"]
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.font = Font(bold=True)

    epoch = datetime.date(2026, 1, 1)
    for r in range(2, STRESS_ROWS + 2):
        ws.cell(row=r, column=1, value=["North", "South", "East", "West"][r % 4])
        ws.cell(row=r, column=2, value=f"Rep {r - 1:03d}")
        ws.cell(row=r, column=3, value=epoch + datetime.timedelta(days=r)).number_format
        ws.cell(row=r, column=3).number_format = "yyyy-mm-dd"
        # Every 10th row is blank and every 17th holds a number stored as text:
        # both make an aggregate disagree with a naive row count.
        if r % 10 == 0:
            pass
        elif r % 17 == 0:
            ws.cell(row=r, column=4, value=str(r % 9 + 1))
        else:
            ws.cell(row=r, column=4, value=r % 9 + 1)
        ws.cell(row=r, column=5, value=round(4.25 + (r % 23) * 1.5, 2))
        ws.cell(row=r, column=6, value=f"=IFERROR(D{r}*E{r}, 0)")
        ws.cell(row=r, column=7, value=f'=IF(F{r}>60, "HIGH", "LOW")')

    # Dates before 1900-03-01, where Excel's deliberate leap-year bug puts the
    # serial one day off every other implementation.
    ws["I1"] = "Lotus-bug window"
    for offset, row in enumerate(range(2, 8)):
        ws.cell(row=row, column=9, value=datetime.date(1900, 1, 1 + offset))
        ws.cell(row=row, column=9).number_format = "yyyy-mm-dd"

    ws["K1"] = "Whole-column aggregate"
    ws["K2"] = "=SUM(F:F)"
    ws["K3"] = "=COUNT(D:D)"
    ws["K4"] = "=COUNTA(D:D)"  # disagrees with K3: the text-numbers count here

    # A real table, so the structured reference on 'Cross Refs' has a referent.
    table = Table(displayName="SalesTable", ref=f"A1:G{STRESS_ROWS + 1}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(table)

    ws.freeze_panes = "B2"
    ws.row_dimensions[3].hidden = True
    ws.column_dimensions["E"].hidden = True
    ws.merge_cells("M1:P1")
    ws["M1"] = "Merged banner spanning four columns"
    ws["A1"].comment = Comment(
        "Region codes come from the CRM export, not from finance. "
        "Réf. ticket #4821 — вопрос про кодировку.",
        "Data Steward",
    )
    ws.add_data_validation(_units_validation(ws))
    ws.conditional_formatting.add(
        f"F2:F{STRESS_ROWS + 1}",
        CellIsRule(operator="greaterThan", formula=["60"], font=Font(bold=True)),
    )
    _print_wide(ws)


def _units_validation(ws) -> DataValidation:
    validation = DataValidation(
        type="whole", operator="between", formula1=1, formula2=9
    )
    validation.add(f"D2:D{STRESS_ROWS + 1}")
    return validation


def _stress_lookups(wb: Workbook) -> None:
    """Lookup and aggregate functions, plus formulas built to be too big."""
    ws = wb.create_sheet("Lookups")
    ws["A1"] = "Function"
    ws["B1"] = "Result"
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)

    cases = [
        ("VLOOKUP exact", "=VLOOKUP(\"Rep 007\", 'Sales Data'!B:F, 5, FALSE)"),
        ("VLOOKUP missing", "=VLOOKUP(\"Rep 999\", 'Sales Data'!B:F, 5, FALSE)"),
        (
            "INDEX/MATCH",
            "=INDEX('Sales Data'!F:F, MATCH(\"Rep 012\", 'Sales Data'!B:B, 0))",
        ),
        (
            "XLOOKUP",
            "=XLOOKUP(\"Rep 020\", 'Sales Data'!B:B, 'Sales Data'!F:F, \"none\")",
        ),
        ("SUMIFS", "=SUMIFS('Sales Data'!F:F, 'Sales Data'!A:A, \"North\")"),
        (
            "COUNTIFS",
            "=COUNTIFS('Sales Data'!A:A, \"East\", 'Sales Data'!G:G, \"HIGH\")",
        ),
        ("SUMPRODUCT", "=SUMPRODUCT('Sales Data'!D2:D100, 'Sales Data'!E2:E100)"),
        ("TEXTJOIN", "=TEXTJOIN(\", \", TRUE, 'Sales Data'!A2:A6)"),
        (
            "LET",
            "=LET(total, SUM('Sales Data'!F:F), n, COUNT('Sales Data'!F:F), total/n)",
        ),
        ("Array constant", "=SUM({1,2,3;4,5,6})"),
        ("Intersection", "=SUM('Sales Data'!D2:D10 'Sales Data'!D5:D20)"),
    ]
    for row, (label, formula) in enumerate(cases, start=2):
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=formula)

    ws["A20"] = "Nested IF, 30 deep"
    ws["B20"] = _nested_if(30)
    ws["A21"] = "Formula past the dossier limit"
    ws["B21"] = _oversized_sum(700)
    ws["A22"] = "Deeply parenthesised"
    ws["B22"] = "=" + "(" * 40 + "1+2" + ")" * 40
    ws.column_dimensions["A"].width = 32
    _print_wide(ws)


def _nested_if(depth: int) -> str:
    """``=IF(A1>1, IF(A1>2, … , 0), 0)`` — a chain the decomposer must walk."""
    formula = "0"
    for level in range(depth, 0, -1):
        formula = f"IF('Sales Data'!K2>{level}, {formula}, {level})"
    return "=" + formula


def _oversized_sum(terms: int) -> str:
    """A formula whose dossier exceeds ``aidoc.MAX_DOSSIER_CHARS``.

    Excel's own ceiling is 8,192 characters, so this stays legal while forcing
    the truncation branch that replaces the decomposition with a marker.
    """
    return "=" + "+".join(
        f"ROUND('Sales Data'!E{i % 200 + 2}, 2)" for i in range(terms)
    )


def _stress_errors(wb: Workbook) -> None:
    """One cell per Excel error value, and the guards that swallow them."""
    ws = wb.create_sheet("Errors")
    ws["A1"] = "Error"
    ws["B1"] = "Raw"
    ws["C1"] = "Guarded"
    for cell in ("A1", "B1", "C1"):
        ws[cell].font = Font(bold=True)

    cases = [
        ("#DIV/0!", "=1/0"),
        ("#VALUE!", '="text"+1'),
        ("#N/A", "=NA()"),
        ("#NAME?", "=NoSuchFunction(1)"),
        ("#NUM!", "=SQRT(-1)"),
        ("#NULL!", "=SUM(A1:A2 B5:B6)"),
        ("#REF!", "=#REF!"),
    ]
    for row, (label, formula) in enumerate(cases, start=2):
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=formula)
        ws.cell(row=row, column=3, value=f'=IFERROR({formula[1:]}, "handled")')

    ws["A12"] = "Aggregate over errors"
    ws["B12"] = "=SUM(B2:B8)"  # itself an error: errors propagate through SUM
    ws["B13"] = "=AGGREGATE(9, 6, B2:B8)"  # 6 = ignore errors
    ws.column_dimensions["A"].width = 24
    _print_wide(ws)


def _stress_volatile(wb: Workbook) -> None:
    """Functions whose value depends on when — or on a string built at runtime."""
    ws = wb.create_sheet("Volatile")
    rows = [
        ("TODAY", "=TODAY()"),
        ("NOW", "=NOW()"),
        ("RAND", "=RAND()"),
        ("RANDBETWEEN", "=RANDBETWEEN(1, 100)"),
        # INDIRECT and OFFSET build their target at evaluation time, so no static
        # analysis can draw the edge — these must land as unresolved, not wrong.
        ("INDIRECT literal", "=INDIRECT(\"'Sales Data'!F2\")"),
        ("INDIRECT computed", "=INDIRECT(\"'Sales Data'!F\" & 2 + 3)"),
        ("OFFSET", "=OFFSET('Sales Data'!A1, 5, 5)"),
        ("INDIRECT of a name", '=INDIRECT("TaxRate")'),
    ]
    ws["A1"] = "Function"
    ws["B1"] = "Value"
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    for row, (label, formula) in enumerate(rows, start=2):
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=formula)
    ws.column_dimensions["A"].width = 22
    _print_wide(ws)


def _stress_cross_refs(wb: Workbook) -> None:
    """References that leave the sheet, the workbook, or logic itself."""
    ws = wb.create_sheet("Cross Refs")
    rows = [
        ("3-D reference", "=SUM('Sales Data:Lookups'!A1)"),
        ("External workbook", "='[Budget FY26.xlsx]Annual'!$B$4"),
        ("External, unopened path", "='C:\\Reports\\[Old.xlsx]Sheet1'!A1"),
        ("Structured reference", "=SUM(SalesTable[Revenue])"),
        ("Structured, this row", "=SalesTable[[#This Row],[Units]]"),
        ("Quoted sheet name", "='O''Brien''s Café'!B2"),
        ("Defined name", "=TaxRate"),
        ("Sheet-scoped name", "=LocalLimit"),
        ("Name holding a constant", "=CompanyName"),
        ("Name holding a formula", "=RevenueColumn"),
    ]
    ws["A1"] = "Reference kind"
    ws["B1"] = "Formula"
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    for row, (label, formula) in enumerate(rows, start=2):
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=formula)

    # A circular pair. Excel reports it rather than evaluating; the analyser must
    # come back with something instead of recursing until it dies.
    ws["A15"] = "Circular pair"
    ws["B15"] = "=B16+1"
    ws["B16"] = "=B15+1"
    ws["A17"] = "Self reference"
    ws["B17"] = "=B17"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 40
    _print_wide(ws)


def _stress_quoting(wb: Workbook) -> None:
    """A sheet whose own name has to be quoted everywhere it is mentioned."""
    ws = wb.create_sheet("O'Brien's Café")
    ws["A1"] = "Sheet names carry apostrophes, spaces and accents"
    ws["A1"].font = Font(bold=True)
    ws["B2"] = 42
    ws["B3"] = "=B2*2"
    ws["B4"] = "=SUM('Sales Data'!F2:F10) + B3"
    ws["A6"] = "Referenced from 'Cross Refs' as 'O''Brien''s Café'!B2"
    ws.column_dimensions["A"].width = 50
    _print_wide(ws)


def _stress_hidden(wb: Workbook) -> None:
    """Sheets the reader never sees, which still feed the visible ones."""
    config = wb.create_sheet("Hidden Config")
    config["A1"] = "Setting"
    config["B1"] = "Value"
    config["A2"] = "Tax rate"
    config["B2"] = 0.196
    config["A3"] = "Local limit"
    config["B3"] = 250_000
    config.sheet_state = "hidden"

    archive = wb.create_sheet("Very Hidden Archive")
    archive["A1"] = "Only reachable through the VBA project editor"
    archive["A2"] = "=SUM('Sales Data'!F:F)"
    archive.sheet_state = "veryHidden"


def _stress_hostile(wb: Workbook) -> None:
    """Cell text, comments and a defined name aimed at the HTML escaping.

    None of these should ever execute. If any does, the failure is visible as a
    dialog or a broken report rather than as a silent wrong answer.
    """
    ws = wb.create_sheet("Hostile Text")
    ws["A1"] = "Payload"
    ws["B1"] = "Stored"
    ws["C1"] = "Concatenated into a formula"
    for cell in ("A1", "B1", "C1"):
        ws[cell].font = Font(bold=True)

    for row, payload in enumerate(HOSTILE_STRINGS, start=2):
        ws.cell(row=row, column=1, value=f"case {row - 1}")
        ws.cell(row=row, column=2, value=payload)
        ws.cell(row=row, column=3, value=f'=B{row} & " (copied)"')

    ws["A12"] = "Very long single value"
    ws["B12"] = "long " * 2_000
    ws["A13"] = "Right-to-left text"
    ws["B13"] = "מכירות רבעוניות"
    ws["A14"] = "Emoji and combining marks"
    ws["B14"] = "revenue 📈 café e\u0301"

    ws["B2"].comment = Comment(
        "A comment that also closes the tag: </script><b>bold</b>",
        "</script>",
    )
    ws.column_dimensions["B"].width = 46
    _print_wide(ws)


def _stress_names(wb: Workbook) -> None:
    """Defined names of every shape the graph has to resolve."""
    wb.defined_names.add(DefinedName("TaxRate", attr_text="'Hidden Config'!$B$2"))
    wb.defined_names.add(DefinedName("CompanyName", attr_text='"Contoso Ltd"'))
    wb.defined_names.add(
        DefinedName("RevenueColumn", attr_text="'Sales Data'!$F$2:$F$251")
    )
    wb.defined_names.add(DefinedName("Café_Total", attr_text="'O''Brien''s Café'!$B$4"))
    # A name pointing at a range that no longer exists.
    wb.defined_names.add(DefinedName("BrokenName", attr_text="#REF!$A$1"))
    # Sheet-scoped: same short name could exist on another sheet with another
    # target, so resolution has to consider scope and not just the name.
    cross_refs = wb["Cross Refs"]
    cross_refs.defined_names.add(
        DefinedName("LocalLimit", attr_text="'Hidden Config'!$B$3")
    )


# ──────────────────────────────────────────────
# Shared helpers
# ──────────────────────────────────────────────


def _print_wide(ws) -> None:
    """Landscape A3, fitted to width — keeps the rendered PNGs legible."""
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def _to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
